from django.shortcuts import render

# Create your views here.
from django.shortcuts import render
from django.http import HttpResponse
# Create your views here.
from django.shortcuts import render
from .models import *
from .serializers import *
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser, AllowAny
from django.contrib.auth.hashers import check_password
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate, login
from rest_framework.parsers import JSONParser, FormParser, MultiPartParser
from decimal import Decimal
from rest_framework import status, permissions


from django.shortcuts import get_object_or_404

import secrets
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from rest_framework_simplejwt.tokens import RefreshToken, AccessToken
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.exceptions import AuthenticationFailed
from rest_framework_simplejwt.tokens import AccessToken

from .models import Admin
from .serializers import AdminSerializer


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.exceptions import AuthenticationFailed
from rest_framework_simplejwt.tokens import AccessToken

from .models import Admin
from .serializers import AdminSerializer

from django.core.mail import send_mail
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth.hashers import make_password
from django.db.models import Q

from django.db.models import Count, Q
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.pagination import PageNumberPagination


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db import transaction
from datetime import datetime
import re

from .models import Shipment_product, InventorySerial
from .serializers import ShipmentProductSerializer

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.pagination import PageNumberPagination

from .models import Employee
from .serializers import EmployeeSerializer


from django.db.models import Count, Q
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import Lead
import csv
import os
from datetime import datetime

from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import Lead, Employee
from .serializers import EmployeeLeadMonthlyReportSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator

from .models import DistributorInformation
from .serializers import (
    DistributorCreateSerializer,
    DistributorLoginSerializer
)
import io

import qrcode
import barcode
from barcode.writer import ImageWriter
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors

from reportlab.lib.pagesizes import landscape
from reportlab.lib.units import mm, inch
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from .models import InventorySerial, Product
from decimal import Decimal, InvalidOperation
from rest_framework.pagination import PageNumberPagination
from django.db import IntegrityError
import calendar
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from .models import Visit
from .serializers import VisitSerializer

import os
from django.conf import settings

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
import os
from django.conf import settings
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.exceptions import AuthenticationFailed
from rest_framework_simplejwt.tokens import AccessToken
from django.contrib.auth.hashers import make_password

from .models import DistributorInformation
from .serializers import DistributorCreateSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser
from django.shortcuts import get_object_or_404
from myapp.models import Warranty, Product, User
from myapp.serializers import WarrantySerializer

# views.py

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import Warranty, Product
from .serializers import WarrantySerializer

# views.py

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import Warranty, Product
from .serializers import WarrantySerializer

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from rest_framework.parsers import MultiPartParser, FormParser

from myapp.models import (
    Warranty,
    Product,
    Serial,
    DistributorInformation
)
from myapp.serializers import WarrantySerializer

from rest_framework.views import APIView
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.core.files.storage import default_storage  # ✅ THIS LINE

from .models import Warranty
from .serializers import WarrantySerializer
from rest_framework.views import APIView
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.core.files.storage import default_storage
from django.http import Http404
from .models import Warranty
from .serializers import WarrantySerializer

from django.core.mail import EmailMessage
from django.conf import settings
import os
import traceback
from django.conf import settings
from django.core.files.storage import default_storage
from django.core.mail import EmailMessage
from django.shortcuts import get_object_or_404

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser

from .models import Warranty
from .serializers import WarrantySerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.pagination import PageNumberPagination

from .models import InventorySerial, Product, Shipment_product
from .serializers import InventorySerialSerializer
from django.shortcuts import get_object_or_404, render
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import InventorySerial, InventorySerialHistory, DistributerOrders
from .utils import generate_inventory_barcode   # ✅ IMPORT ADDED






# utils.py
from .utils import create_folder

def my_view(request):
    create_folder()
    return HttpResponse("Done")


class AdminView(APIView):

    # 🔹 GET (single / all)
    def get(self, request, id=None):
        if id:
            try:
                admin = Admin.objects.get(id=id)
                serializer = AdminSerializer(admin)
                return Response({
                    "status": "success",
                    "data": serializer.data
                })
            except Admin.DoesNotExist:
                return Response({"status": "invalid id"}, status=status.HTTP_404_NOT_FOUND)

        admins = Admin.objects.all().order_by("-created_at")
        serializer = AdminSerializer(admins, many=True)
        return Response({
            "status": "success",
            "message":"Admin list fetched successfully",
            "data": serializer.data
        })

    # 🔹 POST
    def post(self, request):
        serializer = AdminSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "status": "success",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": "invalid data",
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # 🔹 PATCH (partial update)
    def patch(self, request, id=None):
        if not id:
            return Response({'status': 'error', 'message': 'ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            admin = Admin.objects.get(id=id)
        except Admin.DoesNotExist:
            return Response({"status": "invalid id"}, status=status.HTTP_404_NOT_FOUND)

        serializer = AdminSerializer(admin, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "status": "success",
                "data": serializer.data
            })

        return Response({
            "status": "invalid data",
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # 🔹 DELETE
    def delete(self, request, id=None):
        if not id:
            return Response({"status": "invalid id"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            admin = Admin.objects.get(id=id)
            admin.delete()
            return Response({"status": "deleted"})
        except Admin.DoesNotExist:
            return Response({"status": "invalid id"}, status=status.HTTP_404_NOT_FOUND)


class AdminLoginView(APIView):
    def post(self, request):
        serializer = AdminLoginSerializer(data=request.data)
        if serializer.is_valid():
            admin = serializer.validated_data

            refresh = RefreshToken.for_user(admin)

            return Response({
                "message": "Login successful",
                "refresh": str(refresh),
                "access": str(refresh.access_token),
                "admin_id": admin.id,
                "status": admin.status
            })

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# from rest_framework.permissions import IsAuthenticated
# class AdminProfile(APIView):
#     permission_classes = [IsAuthenticated]

#     def get(self, request):
#         admin = request.user
#         return Response({
#             "id": admin.id,
#             "name": admin.name,
#             "email": admin.email,
#             "status": admin.status
#         })



@method_decorator(csrf_exempt, name='dispatch')
class AdminLoginAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get("email")
        password = request.data.get("password")

        if not email or not password:
            return Response(
                {"status": "error", "message": "Email and password required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Verify admin
        try:
            admin = Admin.objects.get(email=email)
        except Admin.DoesNotExist:
            return Response(
                {"status": "error", "message": "Invalid email or password"},
                status=status.HTTP_401_UNAUTHORIZED,
            )


        # OR if password is plain text (NOT RECOMMENDED):
        if admin.password != password:
            return Response(
                {"status": "error", "message": "Invalid email or password"},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        # Generate JWT tokens
        refresh = RefreshToken()
        refresh['admin_id'] = admin.id
        refresh['email'] = admin.email
        refresh['is_admin'] = True  # Flag to identify admin tokens

        return Response({
            "status": "success",
            "message": "Login successful",
            "access_token": str(refresh.access_token),
            "refresh_token": str(refresh),
            "data": {
                "id": admin.id,
                "email": admin.email,
                "password": admin.password,
            }
        }, status=status.HTTP_200_OK)




class AdminProfileView(APIView):

    def get_admin(self, request):
        auth = request.headers.get("Authorization")

        if not auth or not auth.startswith("Bearer "):
            raise AuthenticationFailed("Authorization header missing")

        token_str = auth.split(" ")[1]

        try:
            token = AccessToken(token_str)
        except Exception:
            raise AuthenticationFailed("Invalid or expired token")

        admin_id = token.get("admin_id")
        if not admin_id:
            raise AuthenticationFailed("admin_id not found in token")

        try:
            return Admin.objects.get(id=admin_id)
        except Admin.DoesNotExist:
            raise AuthenticationFailed("Admin not found")

    def get(self, request):
        admin = self.get_admin(request)
        return Response({
            "status": "success",
            "data": AdminSerializer(admin).data
        })

    def patch(self, request):
        admin = self.get_admin(request)
        serializer = AdminSerializer(admin, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def delete(self, request):
        admin = self.get_admin(request)
        admin.delete()
        return Response({"message": "Admin deleted"})

class AdminForgotPasswordView(APIView):

    def post(self, request):
        email = request.data.get("email")

        if email is None or str(email).strip() == "":
            return Response(
                {
                    "status": "error",
                    "message": "Email is required"
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            admin = Admin.objects.get(email=email)
        except Admin.DoesNotExist:
            return Response(
                {
                    "status": "error",
                    "message": "Email not registered"
                },
                status=status.HTTP_404_NOT_FOUND
            )

        reset_obj = AdminPasswordReset.objects.create(admin=admin)

        return Response(
            {
                "status": "success",
                "message": "Password reset link sent",
                "token": str(reset_obj.token)  # 👈 for Postman testing
            },
            status=status.HTTP_200_OK
        )

class AdminResetPasswordView(APIView):

    def post(self, request):
        token = request.data.get("token")
        new_password = request.data.get("new_password")
        confirm_password = request.data.get("confirm_password")

        if not token or not new_password or not confirm_password:
            return Response(
                {
                    "status": "Invalid data",
                    "message": "Token, new password and confirm password are required"
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        if new_password != confirm_password:
            return Response(
                {
                    "status": "error",
                    "message": "Passwords do not match"
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            reset_obj = AdminPasswordReset.objects.get(token=token)
        except AdminPasswordReset.DoesNotExist:
            return Response(
                {
                    "status": "error",
                    "message": "Invalid or expired token"
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        admin = reset_obj.admin
        admin.password = new_password  # ✅ PLAIN TEXT
        admin.save()

        reset_obj.delete()

        return Response(
            {
                "status": "success",
                "message": "Password reset successfully"
            },
            status=status.HTTP_200_OK
        )

class CategoryView(APIView):
    """GET categories: all, single by ID, or search by name/ID"""

    def get(self, request, id=None):
        # 1️⃣ Single category by path ID
        if id:
            try:
                category = Category.objects.get(id=id)
                serializer = CategorySerializer(category)
                return Response({
                    "success": True,
                    "count": 1,
                    "data": serializer.data
                }, status=status.HTTP_200_OK)
            except Category.DoesNotExist:
                return Response({
                    "success": False,
                    "message": "Category not found"
                }, status=status.HTTP_404_NOT_FOUND)

        # 2️⃣ All categories / search
        search = request.query_params.get('search', '')

        # Clean search: remove spaces and trailing slashes
        search = search.strip().rstrip('/')

        categories = Category.objects.all()

        if search:
            if search.isdigit():
                # Search by ID OR by name if numeric
                categories = categories.filter(
                    Q(id=int(search)) | Q(name__icontains=search)
                )
            else:
                # Search by name only
                categories = categories.filter(
                    name__icontains=search
                )

        categories = categories.order_by('-id')
        serializer = CategorySerializer(categories, many=True)

        return Response({
            "success": True,
            "message":"Category list fetched successfully",
            "count": categories.count(),
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    # POST method
    def post(self, request):
        serializer = CategorySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                'status': 'success',
                'message': 'Category created successfully',
                'data': serializer.data
            }, status=status.HTTP_201_CREATED)
        return Response({
            'status': 'Invalid data',

            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # PATCH method (partial update)
    def patch(self, request, id=None):
        if not id:
            return Response({'status': 'error', 'message': 'ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            category = Category.objects.get(id=id)
        except Category.DoesNotExist:
            return Response({'status': 'error', 'message': 'Category not found'}, status=status.HTTP_404_NOT_FOUND)

        serializer = CategorySerializer(category, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({
                'status': 'success',
                'message': 'Category updated successfully',
                'data': serializer.data
            })
        return Response({
            'status': 'error',
            'message': 'Invalid data',
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # DELETE method
    def delete(self, request, id=None):
        if not id:
            return Response({
                'status': 'error',
                'message': 'ID is required'
                }, status=status.HTTP_400_BAD_REQUEST)

        try:
            category = Category.objects.get(id=id)
            category.delete()
            return Response({
            'status': 'success',
            'message': 'Category deleted successfully'
            })
        except Category.DoesNotExist:
            return Response({'status': 'error', 'message': 'Category not found'}, status=status.HTTP_404_NOT_FOUND)

class MaterialAPIView(APIView):

    def get(self, request, pk=None):
        title = request.query_params.get('title')
        status_param = request.query_params.get('status')

        # Get by PK
        if pk:
            material = get_object_or_404(Material, pk=pk)
            serializer = MaterialSerializer(material)
            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # Base queryset
        materials = Material.objects.all()

        # Filter by Title (Partial Match)
        if title:
            materials = materials.filter(title__icontains=title)

        # Filter by Status (Boolean)
        if status_param is not None:
            if status_param.lower() in ['true', '1']:
                materials = materials.filter(status=True)
            elif status_param.lower() in ['false', '0']:
                materials = materials.filter(status=False)
            else:
                return Response({
                    "success": False,
                    "message": "Invalid status value. Use true or false."
                }, status=status.HTTP_400_BAD_REQUEST)

        # No data found
        if not materials.exists():
            return Response({
                "success": False,
                "message": "No materials found with given filters"
            }, status=status.HTTP_404_NOT_FOUND)

        serializer = MaterialSerializer(materials, many=True)
        return Response({
            "success": True,
            "message":"Material list fetched successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)
    # ---------- POST ----------
    def post(self, request):
        serializer = MaterialSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Material created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- PATCH ----------
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "ID is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        material = get_object_or_404(Material, pk=pk)
        serializer = MaterialSerializer(material, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Material updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- DELETE ----------
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        material = get_object_or_404(Material, pk=pk)
        material.delete()

        return Response({
            "success": True,
            "message": "Material deleted successfully"
        }, status=status.HTTP_200_OK)



class ColourAPIView(APIView):

    # ---------- GET ----------
    def get(self, request, pk=None):
        if pk:
            colour = get_object_or_404(Colour, pk=pk)
            serializer = ColourSerializer(colour)
            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        colours = Colour.objects.all()

        # 🔍 Filters
        colour_name = request.query_params.get('colour_name')
        status_param = request.query_params.get('status')

        if colour_name:
            colours = colours.filter(colour_name__icontains=colour_name)

        if status_param is not None:
            # ✅ Convert string to boolean safely
            status_param = status_param.lower()

            if status_param in ['true', '1']:
                status_param = True
            elif status_param in ['false', '0']:
                status_param = False
            else:
                return Response({
                    "success": False,
                    "message": "Invalid status value. Use true/false or 1/0."
                }, status=status.HTTP_400_BAD_REQUEST)

            colours = colours.filter(status=status_param)

        serializer = ColourSerializer(colours, many=True)

        return Response({
            "success": True,
            "message":"Colour list fetched successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    # ---------- POST ----------
    def post(self, request):
        serializer = ColourSerializer(data=request.data)

        if serializer.is_valid():
            colour = serializer.save()
            return Response({
                "success": True,
                "message": "Colour created successfully",
                "data": {
                    "id": colour.id,
                    "colour_name": colour.colour_name,
                    "status": colour.status
                }
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- PATCH ----------
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "ID is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        colour = get_object_or_404(Colour, pk=pk)
        serializer = ColourSerializer(colour, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Colour updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- DELETE ----------
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "Id is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        colour = get_object_or_404(Colour, pk=pk)
        colour.delete()

        return Response({
            "success": True,
            "message": "Colour deleted successfully"
        }, status=status.HTTP_200_OK)


class ProductAPI(APIView):

    parser_classes = (MultiPartParser, FormParser)

    def get(self, request, pk=None):

        # 🔹 Get filter params
        colour_id = request.GET.get("colour_id")
        colour_name = request.GET.get("colour_name")

        material_id = request.GET.get("material_id")
        material_name = request.GET.get("material_name")

        product_name = request.GET.get("product_name")

        # ✅ NEW: Status filter
        status = request.GET.get("status")
        status_bool = None
        if status is not None:
            if status.lower() == "true":
                status_bool = True
            elif status.lower() == "false":
                status_bool = False

        # 🔹 Sequence-wise list (/products/sequence/)
        if request.path.endswith('/products/sequence/'):
            products = (
                Product.objects
                .annotate(
                    total_product=Count("inventory_serials"),
                    reserved_product=Count(
                        "inventory_serials",
                        filter=Q(inventory_serials__status="RESERVED")
                    ),
                    stock_available=Count(
                        "inventory_serials",
                        filter=Q(inventory_serials__status="IN_STOCK")
                    )
                )
                .order_by("product_sequence")
            )

            # ✅ Colour filter
            if colour_id:
                products = products.filter(colour_id=colour_id)

            if colour_name:
                products = products.filter(colour_id__name__icontains=colour_name)

            # ✅ Material filter
            if material_id:
                products = products.filter(material_id=material_id)

            if material_name:
                products = products.filter(material_id__name__icontains=material_name)

            # ✅ Product name filter
            if product_name:
                products = products.filter(product_name__icontains=product_name)


            # ✅ Status filter
            if status_bool is not None:
                products = products.filter(status=status_bool)

            return Response({
                "success": True,
                "count": products.count(),
                "data": ProductSerializer(products, many=True).data
            })

        # 🔹 Single by ID
        if pk:
            product = Product.objects.filter(id=pk)

            if product_name:
                product = product.filter(product_name__icontains=product_name)

            # ✅ Status filter (optional but added safely)
            if status_bool is not None:
                product = product.filter(status=status_bool)

            product = product.first()

            if not product:
                return Response({
                    "success": False,
                    "message": "Product not found"
                }, status=404)

            return Response({
                "success": True,
                "count": 1,
                "data": ProductSerializer(product).data
            })

        # 🔹 Normal list
        products = Product.objects.all().order_by("-id")

        # ✅ Colour filter
        if colour_id:
            products = products.filter(colour_id=colour_id)

        if colour_name:
            products = products.filter(colour_id__name__icontains=colour_name)

        # ✅ Material filter
        if material_id:
            products = products.filter(material_id=material_id)

        if material_name:
            products = products.filter(material_id__name__icontains=material_name)

        # ✅ Product name filter
        if product_name:
            products = products.filter(product_name__icontains=product_name)


        # ✅ Status filter
        if status_bool is not None:
            products = products.filter(status=status_bool)

        return Response({
            "success": True,
            "message":"Product list fetched successfully",
            "count": products.count(),
            "data": ProductSerializer(products, many=True).data
        })

    def post(self, request):
        serializer = ProductSerializer(data=request.data)
        if serializer.is_valid():
            product = serializer.save()
            return Response(
                {"success": True, "data": ProductSerializer(product).data},
                status=201
            )
        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


    def patch(self, request, pk):
        product = Product.objects.filter(id=pk).first()
        if not product:
               return Response({'status': 'error', 'message': 'ID is required'}, status=status.HTTP_400_BAD_REQUEST)
        serializer = ProductSerializer(product, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"success": True})

        return Response(serializer.errors, status=400)

    def delete(self, request, pk):
        product = Product.objects.filter(id=pk).first()
        if not product:
            return Response({"message": "Not found"}, status=404)

        product.delete()
        return Response({"success": True})



class BannerAPIView(APIView):

    def get(self,request,pk=None):

        if pk:
            banner = get_object_or_404(Banner,pk=pk)
            serializer= BannerSerializer(banner)
            return Response({
                "success":True,
                "message":"banner created successfully",
                "data":serializer.data
            },status = status.HTTP_200_OK
            )
        banner = Banner.objects.all()
        serializer =BannerSerializer(banner, many =True)

        return Response(
            {
                "success":True,
                "message":"Banner list fetched successfully",
                # "count":Register.count(),
                "data":serializer.data
            },status=status.HTTP_200_OK)
    def post(self,request):
        serializer = BannerSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success":True,
                "message":"banner data are created successfully...",
                "data": serializer.data
            },status=status.HTTP_201_CREATED)
        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "sucess":False,
                "Errors":"ID is required..."
            },status=status.HTTP_400_BAD_REQUEST)

        banner = get_object_or_404(Banner, pk=pk)

        serializer = BannerSerializer(banner,data=request.data,partial=True)

        if serializer.is_valid():
            serializer.save()

            return Response({
                "sucess":True,
                "message":"Sponsored content updated successfully"
            },status=status.HTTP_200_OK)

        return Response({
            "Success":False,
            "errors":serializer.errors
        },status=status.HTTP_400_BAD_REQUEST)
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "Id is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        banner = get_object_or_404(Banner, pk=pk)
        banner.delete()
        return Response({
            "success": True,
            "message": "Banner content deleted successfully"
        }, status=status.HTTP_200_OK)


class ImportShipmentPagination(PageNumberPagination):
    page_size = 10  # default items per page
    page_size_query_param = 'page_size'  # allow dynamic page size
    max_page_size = 100

class ImportShipmentAPI(APIView):

    def get(self, request, pk=None):
        # 🔹 Get single record
        if pk:
            try:
                obj = Import_shipment.objects.get(pk=pk)
            except Import_shipment.DoesNotExist:
                return Response(
                    {"success": False, "message": "Import shipment not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            serializer = ImportShipmentSerializer(obj)
            return Response({"success": True, "data": serializer.data})

        # 🔹 Queryset for all records
        qs = Import_shipment.objects.all().order_by('-id')


        # 🔹 Get query params individually
        supplier_name = request.query_params.get("supplier_name", "").strip()
        supplier_invoice_no = request.query_params.get("supplier_invoice_no", "").strip()
        invoice_currency = request.query_params.get("invoice_currency", "").strip()
        arrival_date = request.query_params.get("arrival_date", "").strip()  # YYYY-MM-DD

        # 🔹 Apply filters if param is provided
        if supplier_name:
            qs = qs.filter(supplier_name__icontains=supplier_name)
        if supplier_invoice_no:
            qs = qs.filter(supplier_invoice_no__icontains=supplier_invoice_no)
        if invoice_currency:
            qs = qs.filter(invoice_currency__icontains=invoice_currency)
        if arrival_date:
            qs = qs.filter(arrival_date=arrival_date)  # exact match for date

        # 🔹 Apply Pagination
        paginator = ImportShipmentPagination()
        paginated_qs = paginator.paginate_queryset(qs, request)

        serializer = ImportShipmentSerializer(qs, many=True)
        return Response({
            "success": True,
            "message":"Import Shipment fetched successfully",
            "count": qs.count(),
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        })


    def post(self, request):
        serializer = ImportShipmentSerializer(data=request.data)
        if serializer.is_valid():
            Import_shipment.objects.create(**serializer.validated_data)
            return Response(
                {
                    "success": True,
                    "message": "Import shipment created successfully",
                    "data": serializer.data
                },
                status=status.HTTP_201_CREATED
            )

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


    def patch(self, request, pk=None):
        if not pk:
            return Response(
                {"success": False, "message": "ID is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            obj = Import_shipment.objects.get(pk=pk)
        except Import_shipment.DoesNotExist:
            return Response(
                {"success": False, "message": "Import shipment not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = ImportShipmentSerializer(obj, data=request.data, partial=True)
        if serializer.is_valid():
            for key, value in serializer.validated_data.items():
                setattr(obj, key, value)
            obj.save()

            return Response({
                "success": True,
                "message": "Import shipment updated successfully",
                "data": serializer.data
            })

        return Response(
            {"success": False, "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )

    def delete(self, request, pk=None):
        if not pk:
            return Response(
                {"success": False, "message": "Id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            obj = Import_shipment.objects.get(pk=pk)
        except Import_shipment.DoesNotExist:
            return Response(
                {"success": False, "message": "Import shipment not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        obj.delete()
        return Response(
            {"success": True, "message": "Import shipment deleted successfully"},
            status=status.HTTP_200_OK
        )

class ImportCostPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class ImportCostAPI(APIView):

    def get(self, request, pk=None):

        # 🔹 Single record (no pagination)
        if pk:
            try:
                cost = Import_cost.objects.get(pk=pk)
            except Import_cost.DoesNotExist:
                return Response(
                    {"success": False, "message": "Cost not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            serializer = ImportCostSerializer(cost)
            return Response({
                "success": True,
                "data": serializer.data
            })

        # 🔹 Queryset
        qs = Import_cost.objects.all()

        # 🔹 (Optional Filters - if you want like shipment)
        shipment_id = request.query_params.get("shipment_id", "").strip()
        cost_type = request.query_params.get("cost_type", "").strip()

        if shipment_id:
            qs = qs.filter(shipment_id=shipment_id)

        if cost_type:
            qs = qs.filter(cost_type__icontains=cost_type)

        # 🔹 Apply Pagination
        paginator = ImportCostPagination()
        paginated_qs = paginator.paginate_queryset(qs, request)

        serializer = ImportCostSerializer(paginated_qs, many=True)

        return Response({
            "success": True,
            "message":"Import Costs fetched successfully",
            "count": qs.count(),
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        })


    def post(self, request):
        serializer = ImportCostSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"success": True, "message": "Cost created successfully", "data": serializer.data},
                            status=status.HTTP_201_CREATED)
        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk=None):
        if not pk:
            return Response({"success": False, "message": "ID is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            cost = Import_cost.objects.get(pk=pk)
        except Import_cost.DoesNotExist:
            return Response({"success": False, "message": "Cost not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = ImportCostSerializer(cost, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"success": True, "message": "Cost updated successfully", "data": serializer.data})
        return Response({"success": False, "errors": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk=None):
        if not pk:
            return Response({"success": False, "message": "pk is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            cost = Import_cost.objects.get(pk=pk)
        except Import_cost.DoesNotExist:
            return Response({"success": False, "message": "Cost not found"}, status=status.HTTP_404_NOT_FOUND)

        cost.delete()
        return Response({"success": True, "message": "Cost deleted successfully"}, status=status.HTTP_200_OK)


class ShipmentProductPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class ShipmentProductAPI(APIView):

    def get(self, request, pk=None):

        # 🔹 Single record (NO pagination)
        if pk:
            try:
                obj = Shipment_product.objects.get(pk=pk)
            except Shipment_product.DoesNotExist:
                return Response(
                    {"success": False, "message": "Shipment product not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            serializer = ShipmentProductSerializer(obj)
            return Response(
                {"success": True, "data": serializer.data},
                status=status.HTTP_200_OK
            )

        # 🔹 Queryset
        qs = Shipment_product.objects.all().order_by("-id")

        # 🔹 Apply Pagination
        paginator = ShipmentProductPagination()
        paginated_qs = paginator.paginate_queryset(qs, request)

        serializer = ShipmentProductSerializer(paginated_qs, many=True)

        return Response({
            "success": True,
            "message":"ShipmentProduct fetched successfully",

            "count": qs.count(),
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        })
    def post(self, request):
        serializer = ShipmentProductSerializer(data=request.data)

        if not serializer.is_valid():
            return Response(
                {"status": "Invalid data", "errors": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():  # rollback-safe
            shipment_product = serializer.save()

            product = shipment_product.product_id
            batch_data = shipment_product.batch_data
            quantity = shipment_product.quantity
            sku = product.sku

            # ✅ Get location and warehouse from shipment_product
            location = shipment_product.location_id
            warehouse = shipment_product.warehouse_id

            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

            # 🔢 last serial PER PRODUCT
            last_serial = (
                InventorySerial.objects
                .filter(product_id=product)
                .order_by("-id")
                .first()
            )

            start = 1
            if last_serial:
                match = re.search(r"-(\d+)$", last_serial.serial_number)
                if match:
                    start = int(match.group(1)) + 1

            # Create inventory serials with location and warehouse
            serial_objects = []
            for i in range(start, start + quantity):
                serial_number = f"{sku}-{batch_data}-{timestamp}-{i}"

                serial_objects.append(
                    InventorySerial(
                        product_id=product,
                        batch_id=batch_data,
                        serial_number=serial_number,
                        status="IN_STOCK",
                        location=location,
                        warehouse=warehouse
                    )
                )

            InventorySerial.objects.bulk_create(serial_objects)

        return Response(
            {
                "success": True,
                "message": "Shipment product created and inventory serials generated successfully",
                "shipment_product_id": shipment_product.id,
                "serials_created": quantity
            },
            status=status.HTTP_201_CREATED
        )

    def patch(self, request, pk=None):
        if not pk:
            return Response(
                {"success": False, "message": "ID is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            obj = Shipment_product.objects.get(pk=pk)
        except Shipment_product.DoesNotExist:
            return Response(
                {"success": False, "message": "Shipment product not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = ShipmentProductSerializer(obj, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Shipment product updated successfully",
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )

        return Response(
            {"success": False, "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )

    def delete(self, request, pk=None):
        if not pk:
            return Response(
                {"success": False, "message": "Id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            obj = Shipment_product.objects.get(pk=pk)
        except Shipment_product.DoesNotExist:
            return Response(
                {"success": False, "message": "Shipment product not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        obj.delete()

        return Response(
            {"success": True, "message": "Shipment product deleted successfully"},
            status=status.HTTP_200_OK
        )



class DepartmentAPI(APIView):

    # 🔹 GET (all departments or single department)
    def get(self, request, pk=None):
        if pk:
            department = Department.objects.filter(id=pk).first()
            if not department:
                return Response(
                    {"success": False, "message": "Department not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            serializer = DepartmentSerializer(department)
            return Response(
                {"success": True, "data": serializer.data},
                status=status.HTTP_200_OK
            )

        departments = Department.objects.all().order_by('-id')
        serializer = DepartmentSerializer(departments, many=True)

        return Response(
            {
                "success": True,
                "message":"Departments list fetched successfully",
                "count": departments.count(),
                "data": serializer.data
            },
            status=status.HTTP_200_OK
        )

    # 🔹 POST (Create department)
    def post(self, request):
        serializer = DepartmentSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Department created successfully",
                    "data": serializer.data
                },
                status=status.HTTP_201_CREATED
            )

        return Response(
            {"status":"Invalid data",
            "errors": serializer.errors
            },status=status.HTTP_400_BAD_REQUEST)

    # 🔹 PATCH (Partial update)
    def patch(self, request, pk=None):

        if not pk:
            return Response(
                {"success": False, "message": "id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        department = Department.objects.filter(id=pk).first()
        if not department:
              return Response({'status': 'error', 'message': 'ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = DepartmentSerializer(
            department, data=request.data, partial=True
        )

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Department updated successfully",
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )

        return Response(
            {"success": False, "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )

    # 🔹 DELETE
    def delete(self, request, pk):
        department = Department.objects.filter(id=pk).first()
        if not department:
            return Response(
                {"success": False, "message": "Department not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        department.delete()
        return Response(
            {"success": True, "message": "Department deleted successfully"},
            status=status.HTTP_200_OK
        )

class RolesAPI(APIView):

    # 🔹 GET (all roles or single role)
    def get(self, request, pk=None):
        if pk:
            role = Roles.objects.filter(id=pk).first()
            if not role:
                return Response(
                    {"success": False, "message": "Role not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            serializer = RolesSerializer(role)
            return Response(
                {"success": True, "data": serializer.data},
                status=status.HTTP_200_OK
            )

        roles = Roles.objects.all().order_by('-id')
        serializer = RolesSerializer(roles, many=True)

        return Response(
            {
                "success": True,
                "message":"Roles list fetched successfully",
                "count": roles.count(),
                "data": serializer.data
            },
            status=status.HTTP_200_OK
        )

    # 🔹 POST (Create role)
    def post(self, request):
        serializer = RolesSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Role created successfully",
                    "data": serializer.data
                },
                status=status.HTTP_201_CREATED
            )

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # 🔹 PATCH (Partial update)
    def patch(self, request, pk=None):

        if not pk:
            return Response(
                {"success": False, "message": "id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        role = Roles.objects.filter(id=pk).first()
        if not role:
               return Response({'status': 'error', 'message': 'ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = RolesSerializer(
            role, data=request.data, partial=True
        )

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Role updated successfully",
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )

        return Response(
            {"success": False, "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )

    # 🔹 DELETE
    def delete(self, request, pk):
        role = Roles.objects.filter(id=pk).first()
        if not role:
            return Response(
                {"success": False, "message": "Role not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        role.delete()
        return Response(
            {"success": True, "message": "Role deleted successfully"},
            status=status.HTTP_200_OK
        )


class OfficeBranchAPI(APIView):

    # 🔹 GET (all branches or single branch)
    def get(self, request, pk=None):
        if pk:
            branch = Office_branch.objects.filter(id=pk).first()
            if not branch:
                return Response(
                    {"success": False, "message": "Office branch not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            serializer = OfficeBranchSerializer(branch)
            return Response(
                {"success": True, "data": serializer.data},
                status=status.HTTP_200_OK
            )

        branches = Office_branch.objects.all().order_by('-id')
        serializer = OfficeBranchSerializer(branches, many=True)

        return Response(
            {
                "success": True,
                "message":"Office Branches list fetched successfully",
                "count": branches.count(),
                "data": serializer.data
            },
            status=status.HTTP_200_OK
        )

    # 🔹 POST (Create office branch)
    def post(self, request):
        serializer = OfficeBranchSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Office branch created successfully",
                    "data": serializer.data
                },
                status=status.HTTP_201_CREATED
            )

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # 🔹 PATCH (Partial update)
    def patch(self, request, pk=None):
            # ✅ Check ID required
        if not pk:
            return Response(
                {"success": False, "message": "id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        branch = Office_branch.objects.filter(id=pk).first()
        if not branch:
               return Response({'status': 'error', 'message': 'ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = OfficeBranchSerializer(
            branch, data=request.data, partial=True
        )

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Office branch updated successfully",
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )

        return Response(
            {"success": False, "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )


    # 🔹 DELETE
    def delete(self, request, pk):
        branch = get_object_or_404(Office_branch, pk=pk)
        branch.delete()

        return Response(
            {"success": True, "message": "Office branch deleted successfully"},
            status=status.HTTP_200_OK
        )


# class EmployeePagination(PageNumberPagination):
#     page_size = 10
#     page_size_query_param = 'page_size'
#     max_page_size = 100
    # def get_paginated_response(self, data):
    #     return Response({
    #         "status": "success",
    #         "count": self.page.paginator.count,
    #         "next": self.get_next_link(),
    #         "previous": self.get_previous_link(),
    #         "data": data   # ✅ NO "results"
    #     })



class EmployeePagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"

    def get_paginated_response(self, data):
        return Response({
            "success": data.get("success", True),
            "message": data.get("message", ""),
            "count": self.page.paginator.count,
            "total_pages": self.page.paginator.num_pages,   # ✅ total pages
            "current_page": self.page.number,               # ✅ current page
            "next": self.get_next_link(),
            "previous": self.get_previous_link(),
            "data": data.get("data", [])
        })


class EmployeeView(APIView):

    # 🔹 GET (single / all)
    def get(self, request, id=None):

        role_name = request.GET.get("role_name")

        # 🔹 Single record (NO pagination)
        if id:
            try:
                employee = Employee.objects.get(id=id)
                serializer = EmployeeSerializer(employee)
                return Response({
                    "success": True,
                    "message": "Employee details fetched successfully",
                    "count": 1,
                    "data": serializer.data
                }, status=status.HTTP_200_OK)

            except Employee.DoesNotExist:
                return Response({
                    "success": False,
                    "message": "Invalid employee id",
                    "count": 0,
                    "data": []
                }, status=status.HTTP_404_NOT_FOUND)

        # 🔹 Queryset
        employees = Employee.objects.all().order_by("-created_at")

        if role_name:
            employees = employees.filter(role_id__name__icontains=role_name)


        # 🔹 Apply Pagination
        paginator = EmployeePagination()
        paginated_employees = paginator.paginate_queryset(employees, request)

        if paginated_employees is not None:
            serializer = EmployeeSerializer(paginated_employees, many=True)
            return paginator.get_paginated_response({
                "success": True,
                "message": "Employee list fetched successfully",
                "data": serializer.data
            })

        # 🔹 Fallback if pagination not applied
        serializer = EmployeeSerializer(employees, many=True)
        return Response({
            "success": True,
            "message": "Employee list fetched successfully",
            "count": employees.count(),
            "next": None,
            "previous": None,
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    # 🔹 POST
    def post(self, request):
        serializer = EmployeeSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"status": "success", "data": serializer.data}, status=status.HTTP_201_CREATED)

        return Response({"status": "invalid data", "errors": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    # 🔹 PATCH
    def patch(self, request, id=None):
        if not id:
               return Response({'status': False, 'message': 'ID is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            employee = Employee.objects.get(id=id)
        except Employee.DoesNotExist:
            return Response({"status": False,
            "message":"Employee fatched successfully"}, status=status.HTTP_404_NOT_FOUND)

        serializer = EmployeeSerializer(employee, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"status": True,"message":"Employee updated sucessfully" ,"data": serializer.data})

        return Response({"status": False, "errors": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    # 🔹 DELETE
    def delete(self, request, id=None):
        if not id:
            return Response({"status": "error","message":"Id is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = Employee.objects.get(id=id)
            employee.delete()
            return Response({"status": True,"message":"Employee deleted successfully"})
        except Employee.DoesNotExist:
            return Response({"status": "invalid id"}, status=status.HTTP_404_NOT_FOUND)


@method_decorator(csrf_exempt, name='dispatch')
class EmployeeLoginAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        email = request.data.get("email")
        password = request.data.get("password")

        if not email or not password:
            return Response(
                {"status": "Invalid data", "message": "Email and password required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            employee = Employee.objects.get(email=email)
        except Employee.DoesNotExist:
            return Response(
                {"status": "error", "message": "Invalid email or password"},
                status=status.HTTP_401_UNAUTHORIZED
            )

        if employee.password != password:
            return Response(
                {"status": "error", "message": "Invalid email or password"},
                status=status.HTTP_401_UNAUTHORIZED
            )

        # JWT Token
        refresh = RefreshToken()
        refresh["employee_id"] = employee.id
        refresh["email"] = employee.email
        refresh["is_employee"] = True

        # 👇 Use serializer here
        employee_data = EmployeeSerializer(employee).data

        return Response({
            "status": "success",
            "message": "Login successful",
            "access_token": str(refresh.access_token),
            "refresh_token": str(refresh),
            "data": employee_data   # ✅ All fields except password
        }, status=status.HTTP_200_OK)



class EmployeeProfileView(APIView):

    def get_employee(self, request):
        auth = request.headers.get("Authorization")

        if not auth or not auth.startswith("Bearer "):
            raise AuthenticationFailed("Authorization header missing")

        token_str = auth.split(" ")[1]

        try:
            token = AccessToken(token_str)
        except Exception:
            raise AuthenticationFailed("Invalid or expired token")

        employee_id = token.get("employee_id")
        if not employee_id:
            raise AuthenticationFailed("employee_id not found in token")

        try:
            return Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            raise AuthenticationFailed("Employee not found")

    def get(self, request):
        employee = self.get_employee(request)
        return Response({
            "status": "success",
            "data": EmployeeSerializer(employee).data
        })

    def patch(self, request):
        employee = self.get_employee(request)
        serializer = EmployeeSerializer(employee, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def delete(self, request):
        employee = self.get_employee(request)
        employee.delete()
        return Response({"message": "Employee deleted"})
          
class LeadPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class LeadAPI(APIView):

    def get(self, request, pk=None):

        # ---------- 🔹 GET SINGLE LEAD ----------
        if pk:
            lead = Lead.objects.filter(id=pk).first()
            if not lead:
                return Response(
                    {"success": False, "message": "Lead not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            serializer = LeadSerializer(lead)
            return Response(
                {"success": True, "data": serializer.data},
                status=status.HTTP_200_OK
            )

        # ---------- 🔹 BASE QUERYSET ----------
        qs = Lead.objects.all().order_by("-id")

        # ---------- 🔹 FILTERS ----------
        lead_type = request.query_params.get("lead_type", "").strip()
        interest_level = request.query_params.get("interest_level", "").strip()
        lead_status = request.query_params.get("lead_status", "").strip()
        created_by = request.query_params.get("created_by", "").strip()

        assigned_to = request.query_params.get("assigned_to", "").strip()

        if lead_type:
            qs = qs.filter(lead_type__iexact=lead_type)

        if interest_level:
            qs = qs.filter(interest_level__iexact=interest_level)

        if lead_status:
            qs = qs.filter(lead_status__iexact=lead_status)

        if created_by.isdigit():
            qs = qs.filter(created_by_id=int(created_by))

        if assigned_to.isdigit():
            qs = qs.filter(assigned_to_id=int(assigned_to))

        # ---------- 🔹 PAGINATION ----------
        paginator = LeadPagination()
        paginated_qs = paginator.paginate_queryset(qs, request)

        serializer = LeadSerializer(paginated_qs, many=True)

        return Response({
            "success": True,
            "message":"Lead fetched successfully",

            "count": qs.count(),
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        })

    def post(self, request):
        serializer = LeadSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Lead created successfully",
                    "data": serializer.data
                },
                status=status.HTTP_201_CREATED
            )

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # 🔹 PATCH (Partial update)
    def patch(self, request, pk=None):

        if not pk:
            return Response(
                {"success": False, "message": "id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        lead = Lead.objects.filter(id=pk).first()
        if not lead:
            return Response(
                {"success": False, "message": "ID is required"},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = LeadSerializer(
            lead,
            data=request.data,
            partial=True
        )

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Lead updated successfully",
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )

        return Response(
            {"success": False, "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )

    # 🔹 DELETE
    def delete(self, request, pk=None):
        if not pk:
            return Response(
                {"success": False, "message": "id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        lead = Lead.objects.filter(id=pk).first()
        if not lead:
            return Response(
                {"success": False, "message": "Lead not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        lead.delete()
        return Response(
            {"success": True, "message": "Lead deleted successfully"},
            status=status.HTTP_200_OK
        )


class EmployeeLeadMonthlyReportAPI(APIView):

    def get_month_name(self, month_value):
        month_map = {
            1: "January",
            2: "February",
            3: "March",
            4: "April",
            5: "May",
            6: "June",
            7: "July",
            8: "August",
            9: "September",
            10: "October",
            11: "November",
            12: "December",
        }
        return month_map.get(int(month_value), str(month_value))

    def get(self, request):
        try:
            employee_id = request.GET.get("employee_id")
            month = request.GET.get("month")
            year = request.GET.get("year")

            if not month:
                return Response({
                    "success": False,
                    "message": "month is required"
                }, status=status.HTTP_400_BAD_REQUEST)

            try:
                month_int = int(month)
                if month_int < 1 or month_int > 12:
                    return Response({
                        "success": False,
                        "message": "month must be between 1 and 12"
                    }, status=status.HTTP_400_BAD_REQUEST)
            except ValueError:
                return Response({
                    "success": False,
                    "message": "month must be a valid number"
                }, status=status.HTTP_400_BAD_REQUEST)

            # optional year filter
            if year:
                try:
                    year_int = int(year)
                except ValueError:
                    return Response({
                        "success": False,
                        "message": "year must be a valid number"
                    }, status=status.HTTP_400_BAD_REQUEST)
            else:
                year_int = datetime.now().year

            month_name = self.get_month_name(month_int)

            reports_dir = os.path.join(settings.MEDIA_ROOT, "reports")
            os.makedirs(reports_dir, exist_ok=True)

            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

            # ---------------- Styles ----------------
            title_font = Font(size=16, bold=True, color="FFFFFF")
            subtitle_font = Font(size=11, bold=True)
            header_font = Font(size=11, bold=True, color="FFFFFF")
            normal_font = Font(size=10)

            title_fill = PatternFill("solid", fgColor="1F4E78")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            summary_fill = PatternFill("solid", fgColor="D9EAF7")

            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

            thin_border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )

            # =========================================================
            # 1) PARTICULAR EMPLOYEE REPORT (OLD FUNCTIONALITY SAME)
            # =========================================================
            if employee_id:
                try:
                    employee = Employee.objects.get(id=employee_id)
                except Employee.DoesNotExist:
                    return Response({
                        "success": False,
                        "message": "Employee not found"
                    }, status=status.HTTP_404_NOT_FOUND)

                queryset = Lead.objects.filter(
                    created_by_id=employee_id,
                    created_at__month=month_int,
                    created_at__year=year_int
                ).order_by("-id")

                if not queryset.exists():
                    return Response({
                        "success": False,
                        "message": f"No lead data found for this employee in {month_name} {year_int}"
                    }, status=status.HTTP_404_NOT_FOUND)

                total_leads = queryset.count()
                converted_leads = queryset.filter(lead_status__iexact="Converted").count()
                conversion_ratio = round((converted_leads / total_leads) * 100, 2) if total_leads > 0 else 0

                file_name = f"employee_{employee_id}_{month_name}_{year_int}_lead_report_{timestamp}.xlsx"
                file_path = os.path.join(reports_dir, file_name)

                wb = Workbook()
                ws = wb.active
                ws.title = "Lead Report"

                # Title
                ws.merge_cells("A1:L1")
                ws["A1"] = "EMPLOYEE MONTHLY LEAD REPORT"
                ws["A1"].font = title_font
                ws["A1"].fill = title_fill
                ws["A1"].alignment = center_align
                ws.row_dimensions[1].height = 28

                # Summary
                summary_data = [
                    ["Employee Name", f"{employee.first_name} {employee.last_name}".strip(), "Employee Code", getattr(employee, "employee_code", "")],
                    ["Month", month_name, "Year", year_int],
                    ["Total Leads", total_leads, "Converted Leads", converted_leads],
                    ["Conversion Ratio", f"{conversion_ratio}%", "", ""],
                ]

                start_row = 3
                for row_index, row_data in enumerate(summary_data, start=start_row):
                    for col_index, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=row_index, column=col_index, value=value)
                        cell.border = thin_border
                        cell.alignment = left_align
                        if col_index in [1, 3]:
                            cell.font = subtitle_font
                            cell.fill = summary_fill
                        else:
                            cell.font = normal_font

                # Table Header
                table_start_row = 9
                headers = [
                    "Lead ID",
                    "Lead Type",
                    "Business Name",
                    "Contact Person",
                    "Phone",
                    "Email",
                    "City",
                    "State",
                    "Lead Status",
                    "Interest Level",
                    "Region",
                    "Remarks",
                ]

                for col_index, header in enumerate(headers, start=1):
                    cell = ws.cell(row=table_start_row, column=col_index, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border

                # Table Data
                data_row = table_start_row + 1
                for lead in queryset:
                    row_values = [
                        lead.id,
                        getattr(lead, "lead_type", ""),
                        getattr(lead, "business_name", ""),
                        getattr(lead, "contact_person", ""),
                        getattr(lead, "phone", ""),
                        getattr(lead, "email", ""),
                        getattr(lead, "city", ""),
                        getattr(lead, "state", ""),
                        getattr(lead, "lead_status", ""),
                        getattr(lead, "interest_level", ""),
                        getattr(lead, "region", ""),
                        getattr(lead, "remarks", "") if getattr(lead, "remarks", None) else "",
                    ]

                    for col_index, value in enumerate(row_values, start=1):
                        cell = ws.cell(row=data_row, column=col_index, value=value)
                        cell.font = normal_font
                        cell.alignment = left_align
                        cell.border = thin_border

                    data_row += 1

                # Auto width
                for col_idx in range(1, 13):
                    column_letter = get_column_letter(col_idx)
                    max_length = 0
                    for row_num in range(1, ws.max_row + 1):
                        cell_value = ws.cell(row=row_num, column=col_idx).value
                        if cell_value is not None:
                            max_length = max(max_length, len(str(cell_value)))
                    ws.column_dimensions[column_letter].width = min(max_length + 3, 35)

                ws.freeze_panes = "A10"
                wb.save(file_path)

                download_url = request.build_absolute_uri(
                    settings.MEDIA_URL + "reports/" + file_name
                )

                response_data = {
                    "employee_id": employee.id,
                    "employee_name": f"{employee.first_name} {employee.last_name}".strip(),
                    "employee_code": getattr(employee, "employee_code", None),
                    "month": month_name,
                    "year": year_int,
                    "total_leads": total_leads,
                    "converted_leads": converted_leads,
                    "conversion_ratio": f"{conversion_ratio}%",
                    "download_url": download_url,
                }

                serializer = EmployeeLeadMonthlyReportSerializer(response_data)

                return Response({
                    "success": True,
                    "message": "Employee lead report generated successfully",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)

            # =========================================================
            # 2) ALL EMPLOYEES REPORT (NEW FUNCTIONALITY)
            # =========================================================
            queryset = Lead.objects.filter(
                created_at__month=month_int,
                created_at__year=year_int
            ).order_by("-id")

            if not queryset.exists():
                return Response({
                    "success": False,
                    "message": f"No lead data found for {month_name} {year_int}"
                }, status=status.HTTP_404_NOT_FOUND)

            total_leads = queryset.count()
            converted_leads = queryset.filter(lead_status__iexact="Converted").count()
            conversion_ratio = round((converted_leads / total_leads) * 100, 2) if total_leads > 0 else 0

            file_name = f"all_employees_{month_name}_{year_int}_lead_report_{timestamp}.xlsx"
            file_path = os.path.join(reports_dir, file_name)

            wb = Workbook()
            ws = wb.active
            ws.title = "All Employee Lead Report"

            # Title
            ws.merge_cells("A1:M1")
            ws["A1"] = "ALL EMPLOYEES MONTHLY LEAD REPORT"
            ws["A1"].font = title_font
            ws["A1"].fill = title_fill
            ws["A1"].alignment = center_align
            ws.row_dimensions[1].height = 28

            # Summary
            summary_data = [
                ["Month", month_name, "Year", year_int],
                ["Total Leads", total_leads, "Converted Leads", converted_leads],
                ["Conversion Ratio", f"{conversion_ratio}%", "", ""],
            ]

            start_row = 3
            for row_index, row_data in enumerate(summary_data, start=start_row):
                for col_index, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_index, column=col_index, value=value)
                    cell.border = thin_border
                    cell.alignment = left_align
                    if col_index in [1, 3]:
                        cell.font = subtitle_font
                        cell.fill = summary_fill
                    else:
                        cell.font = normal_font

            # Table Header
            table_start_row = 8
            headers = [
                "Lead ID",
                "Employee ID",
                "Employee Name",
                "Lead Type",
                "Business Name",
                "Contact Person",
                "Phone",
                "Email",
                "City",
                "State",
                "Lead Status",
                "Interest Level",
                "Region",
                "Remarks",
            ]

            for col_index, header in enumerate(headers, start=1):
                cell = ws.cell(row=table_start_row, column=col_index, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # Table Data
            data_row = table_start_row + 1
            for lead in queryset:
                employee = getattr(lead, "created_by", None)
                employee_name = ""
                employee_code = ""

                if employee:
                    employee_name = f"{getattr(employee, 'first_name', '')} {getattr(employee, 'last_name', '')}".strip()
                    employee_code = getattr(employee, "employee_code", "")

                row_values = [
                    lead.id,
                    getattr(employee, "id", "") if employee else "",
                    employee_name if employee_name else employee_code,
                    getattr(lead, "lead_type", ""),
                    getattr(lead, "business_name", ""),
                    getattr(lead, "contact_person", ""),
                    getattr(lead, "phone", ""),
                    getattr(lead, "email", ""),
                    getattr(lead, "city", ""),
                    getattr(lead, "state", ""),
                    getattr(lead, "lead_status", ""),
                    getattr(lead, "interest_level", ""),
                    getattr(lead, "region", ""),
                    getattr(lead, "remarks", "") if getattr(lead, "remarks", None) else "",
                ]

                for col_index, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=data_row, column=col_index, value=value)
                    cell.font = normal_font
                    cell.alignment = left_align
                    cell.border = thin_border

                data_row += 1

            # Auto width
            for col_idx in range(1, 15):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for row_num in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row_num, column=col_idx).value
                    if cell_value is not None:
                        max_length = max(max_length, len(str(cell_value)))
                ws.column_dimensions[column_letter].width = min(max_length + 3, 35)

            ws.freeze_panes = "A9"
            wb.save(file_path)

            download_url = request.build_absolute_uri(
                settings.MEDIA_URL + "reports/" + file_name
            )

            # per employee summary
            employee_summary = []
            employees = Employee.objects.all().order_by("id")

            for emp in employees:
                emp_queryset = queryset.filter(created_by_id=emp.id)
                if emp_queryset.exists():
                    emp_total = emp_queryset.count()
                    emp_converted = emp_queryset.filter(lead_status__iexact="Converted").count()
                    emp_ratio = round((emp_converted / emp_total) * 100, 2) if emp_total > 0 else 0

                    employee_summary.append({
                        "employee_id": emp.id,
                        "employee_name": f"{emp.first_name} {emp.last_name}".strip(),
                        "employee_code": getattr(emp, "employee_code", None),
                        "total_leads": emp_total,
                        "converted_leads": emp_converted,
                        "conversion_ratio": f"{emp_ratio}%",
                    })

            return Response({
                "success": True,
                "message": "All employees lead report generated successfully",
                "data": {
                    "month": month_name,
                    "year": year_int,
                    "total_leads": total_leads,
                    "converted_leads": converted_leads,
                    "conversion_ratio": f"{conversion_ratio}%",
                    "download_url": download_url,
                    "employees": employee_summary
                }
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "success": False,
                "message": "Something went wrong",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Employee, Lead


class EmployeeLeadMonthlyReportAPI(APIView):

    def get_month_name(self, month_value):
        month_map = {
            1: "January",
            2: "February",
            3: "March",
            4: "April",
            5: "May",
            6: "June",
            7: "July",
            8: "August",
            9: "September",
            10: "October",
            11: "November",
            12: "December",
        }
        return month_map.get(int(month_value), str(month_value))

    def get(self, request):

        try:
            employee_id = request.GET.get("employee_id")
            month = request.GET.get("month")
            year = request.GET.get("year")

            if not month:
                return Response({
                    "success": False,
                    "message": "month is required"
                }, status=status.HTTP_400_BAD_REQUEST)

            month_int = int(month)

            if month_int < 1 or month_int > 12:
                return Response({
                    "success": False,
                    "message": "month must be between 1 and 12"
                }, status=status.HTTP_400_BAD_REQUEST)

            if year:
                year_int = int(year)
            else:
                year_int = datetime.now().year

            month_name = self.get_month_name(month_int)

            queryset = Lead.objects.filter(
                created_at__month=month_int,
                created_at__year=year_int
            ).order_by("-id")

            if employee_id:
                queryset = queryset.filter(created_by_id=employee_id)

            if not queryset.exists():
                return Response({
                    "success": False,
                    "message": "No lead data found"
                }, status=status.HTTP_404_NOT_FOUND)

            # =========================
            # Create Workbook
            # =========================
            wb = Workbook()
            ws = wb.active
            ws.title = "Lead Report"

            # =========================
            # Styles
            # =========================
            title_font = Font(size=16, bold=True, color="FFFFFF")
            header_font = Font(size=11, bold=True, color="FFFFFF")
            normal_font = Font(size=10)

            title_fill = PatternFill("solid", fgColor="1F4E78")
            header_fill = PatternFill("solid", fgColor="4F81BD")

            center_align = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            left_align = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True
            )

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # =========================
            # Title
            # =========================
            ws.merge_cells("A1:N1")

            if employee_id:
                title = f"EMPLOYEE MONTHLY LEAD REPORT - {month_name} {year_int}"
            else:
                title = f"ALL EMPLOYEE MONTHLY LEAD REPORT - {month_name} {year_int}"

            ws["A1"] = title
            ws["A1"].font = title_font
            ws["A1"].fill = title_fill
            ws["A1"].alignment = center_align

            # =========================
            # Headers
            # =========================
            headers = [
                "Lead ID",
                "Employee ID",
                "Employee Name",
                "Lead Type",
                "Business Name",
                "Contact Person",
                "Phone",
                "Email",
                "City",
                "State",
                "Lead Status",
                "Interest Level",
                "Region",
                "Remarks",
            ]

            header_row = 3

            for col_index, header in enumerate(headers, start=1):

                cell = ws.cell(
                    row=header_row,
                    column=col_index,
                    value=header
                )

                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # =========================
            # Data Rows
            # =========================
            data_row = header_row + 1

            for lead in queryset:

                employee = getattr(lead, "created_by", None)

                employee_name = ""

                if employee:
                    employee_name = (
                        f"{employee.first_name} "
                        f"{employee.last_name}"
                    ).strip()

                row_values = [
                    lead.id,
                    employee.id if employee else "",
                    employee_name,
                    getattr(lead, "lead_type", ""),
                    getattr(lead, "business_name", ""),
                    getattr(lead, "contact_person", ""),
                    getattr(lead, "phone", ""),
                    getattr(lead, "email", ""),
                    getattr(lead, "city", ""),
                    getattr(lead, "state", ""),
                    getattr(lead, "lead_status", ""),
                    getattr(lead, "interest_level", ""),
                    getattr(lead, "region", ""),
                    getattr(lead, "remarks", ""),
                ]

                for col_index, value in enumerate(row_values, start=1):

                    cell = ws.cell(
                        row=data_row,
                        column=col_index,
                        value=value
                    )

                    cell.font = normal_font
                    cell.alignment = left_align
                    cell.border = thin_border

                data_row += 1

            # =========================
            # Auto Column Width
            # =========================
            for col in ws.columns:

                max_length = 0
                column = col[0].column_letter

                for cell in col:

                    try:
                        if cell.value:
                            max_length = max(
                                max_length,
                                len(str(cell.value))
                            )
                    except:
                        pass

                adjusted_width = min(max_length + 5, 40)

                ws.column_dimensions[column].width = adjusted_width

            # =========================
            # Freeze Header
            # =========================
            ws.freeze_panes = "A4"

            # =========================
            # Save File
            # =========================
            reports_dir = os.path.join(
                settings.MEDIA_ROOT,
                "reports"
            )

            os.makedirs(reports_dir, exist_ok=True)

            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

            file_name = (
                f"lead_report_{month_name}_{year_int}_{timestamp}.xlsx"
            )

            file_path = os.path.join(
                reports_dir,
                file_name
            )

            wb.save(file_path)

            # =========================
            # Download Response
            # =========================
            response = FileResponse(
                open(file_path, "rb"),
                as_attachment=True,
                filename=file_name,
                content_type=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                )
            )

            return response

        except Exception as e:

            return Response({
                "success": False,
                "message": "Something went wrong",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExpensePagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class ExpenseAPI(APIView):

    # 🔹 GET (List / Single)
    def get(self, request, pk=None):

        # Single record by PK
        if pk:
            try:
                expense = Expense.objects.get(pk=pk)
            except Expense.DoesNotExist:
                return Response(
                    {"success": False, "message": "Expense not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            serializer = ExpenseSerializer(expense)
            return Response({"success": True, "data": serializer.data})

        # 🔹 Filter by employee_id (Query Param)
        employee_id = request.query_params.get('employee_id')

        qs = Expense.objects.all().order_by('-id')

        if employee_id:
            qs = qs.filter(employee_id=employee_id)

           # 🔹 APPLY PAGINATION
        paginator = ExpensePagination()
        paginated_qs = paginator.paginate_queryset(qs, request)

        serializer = ExpenseSerializer(qs, many=True)

        return Response({
            "success": True,
            "message":"Expense fetched successfully",

            "count": qs.count(),
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        })
    # 🔹 POST (Create)
    def post(self, request):
        serializer = ExpenseSerializer(data=request.data)

        # Manual validations


        if not request.data.get("expense_type"):
            return Response({"success": False, "message": "Expense type is required"},
                            status=status.HTTP_400_BAD_REQUEST)

        if not request.data.get("amount"):
            return Response({"success": False, "message": "Amount is required"},
                            status=status.HTTP_400_BAD_REQUEST)

        if Decimal(request.data.get("amount", 0)) <= 0:
            return Response({"success": False, "message": "Amount must be greater than zero"},
                            status=status.HTTP_400_BAD_REQUEST)

        if not request.data.get("date"):
            return Response({"success": False, "message": "Date is required"},
                            status=status.HTTP_400_BAD_REQUEST)

        if serializer.is_valid():
            expense = serializer.save()  # ✅ Correct way

            # ✅ Re-serialize saved instance
            response_serializer = ExpenseSerializer(expense)

            return Response({
                "success": True,
                "message": "Expense created successfully",
                "data": response_serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    # 🔹 PATCH (Partial Update)
    def patch(self, request, pk=None):
        if not pk:
            return Response({"success": False, "message": "ID is required"},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            expense = Expense.objects.get(pk=pk)
        except Expense.DoesNotExist:
            return Response({"success": False, "message": "Expense not found"},
                            status=status.HTTP_404_NOT_FOUND)

        serializer = ExpenseSerializer(expense, data=request.data, partial=True)

        if serializer.is_valid():
            expense = serializer.save()  # ✅ use save()

            response_serializer = ExpenseSerializer(expense)

            return Response({
                "success": True,
                "message": "Expense updated successfully",
                "data": response_serializer.data
            })

        return Response({"success": False, "errors": serializer.errors},
                        status=status.HTTP_400_BAD_REQUEST)

    # 🔹 DELETE
    def delete(self, request, pk=None):
        if not pk:
            return Response({"success": False, "message": "Id is required"},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            expense = Expense.objects.get(pk=pk)
        except Expense.DoesNotExist:
            return Response({"success": False, "message": "Expense not found"},
                            status=status.HTTP_404_NOT_FOUND)

        expense.delete()
        return Response({"success": True, "message": "Expense deleted successfully"},
                        status=status.HTTP_200_OK)
class BlogAPIView(APIView):

    # ---------- GET (All or Single) ----------
    def get(self, request, pk=None):
        if pk:
            blog = get_object_or_404(Blog, pk=pk)
            serializer = BlogSerializer(blog)
            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        blogs = Blog.objects.all()
        serializer = BlogSerializer(blogs, many=True)

        return Response({
            "success": True,
            "message":"Blogs list fetched successfully",
            # "count": len(serializer.data),
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    # ---------- POST ----------
    def post(self, request):
        serializer = BlogSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Blog created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- PATCH (Partial Update) ---------- #
    def patch(self, request, pk=None):
        # ✅ Check ID required
        if not pk:
            return Response(
                {"success": False, "message": "id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        blog = get_object_or_404(Blog, pk=pk)
        serializer = BlogSerializer(blog, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Blog updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- DELETE ----------
    def delete(self, request, pk=None):
        if not pk:
            return Response(
                {"success": False, "message": "id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        blog = get_object_or_404(Blog, pk=pk)
        blog.delete()

        return Response({
            "success": True,
            "message": "Blog deleted successfully"
        }, status=status.HTTP_204_NO_CONTENT)


class VisitAPI(APIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get(self, request, pk=None):
        if pk:
            visit = get_object_or_404(Visit, id=pk)
            serializer = VisitSerializer(
                visit,
                context={'request': request}
            )
            return Response(
                {"success": True, "data": serializer.data},
                status=status.HTTP_200_OK
            )

        # 🔹 Filters
        lead_id = request.query_params.get('lead_id')
        status_filter = request.query_params.get('status')
        followup_type = request.query_params.get('followup_type')
        employee_id = request.query_params.get('employee_id')  # ✅ ADD THIS

        visits = Visit.objects.all().order_by('-id')

        if lead_id:
            visits = visits.filter(lead_id=lead_id)

        if status_filter:
            visits = visits.filter(status=status_filter.upper())

        if followup_type:
            visits = visits.filter(followup_type=followup_type.upper())

        if employee_id:  # ✅ ADD THIS
            visits = visits.filter(employee_id=employee_id)

        serializer = VisitSerializer(
            visits,
            many=True,
            context={'request': request}
        )

        return Response(
            {
                "success": True,
                "message":"Visists list fetched successfully",
                "count": visits.count(),
                "data": serializer.data
            },
            status=status.HTTP_200_OK
        )

    # 🔹 POST create visit
    def post(self, request):
        serializer = VisitSerializer(
            data=request.data,
            context={'request': request}
        )

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Visit created successfully",
                    "data": serializer.data
                },
                status=status.HTTP_201_CREATED
            )

        return Response({

            "status": 'Invalid data',
            "errors": serializer.errors

        }, status=status.HTTP_400_BAD_REQUEST)

    # 🔹 PATCH update visit (Partial Update)
    def patch(self, request, pk=None):
            # ✅ Check ID required
        if not pk:
            return Response(
                {"success": False, "message": "id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        visit = get_object_or_404(Visit, id=pk)

        serializer = VisitSerializer(
            visit,
            data=request.data,
            partial=True,  # ✅ IMPORTANT
            context={'request': request}
        )

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Visit updated successfully",
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )
        return Response(
            {
                "success": False,
                "errors": serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # 🔹 DELETE visit
    def delete(self, request, pk=None):
        if not pk:
            return Response(
                {"success": False, "message": "id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        visit = get_object_or_404(Visit, id=pk)
        visit.delete()

        return Response(
            {
                "success": True,
                "message": "Visit deleted successfully"
            },
            status=status.HTTP_200_OK
        )

class businessLegalInfoAPIView(APIView):

    def get(self,request,pk=None):
        if pk:
            businessLegalInfo = get_object_or_404(BusinessLegalInfo, pk=pk)
            serializer =BusinessLegalInfoSerializer(businessLegalInfo)
            return Response({
                "success": True,
                "message": "BusinessLegalInfo crate successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔹 All attendance list
        businessLegalInfo = BusinessLegalInfo.objects.all().order_by('-created_at')
        serializer = BusinessLegalInfoSerializer(businessLegalInfo, many=True)

        return Response({
            "success": True,
            "message": "BusinessLegalInfo list create successfully",
            "count": businessLegalInfo.count(),
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    def post(self, request):
        """
        POST a new distributor
        """
        serializer = BusinessLegalInfoSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "BusinessLegalInfo created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)
        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk=None):
        businessLegalInfo = get_object_or_404(BusinessLegalInfo, pk=pk)
        serializer = BusinessLegalInfoSerializer(businessLegalInfo, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "BusinessLegalInfo  updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk=None):
        """
        DELETE a distributor by ID
        """
        if not pk:
            return Response({
                "success": False,
                "message": "BusinessLegalInfo id is required for deletion."},
                status=status.HTTP_400_BAD_REQUEST)

        businessLegalInfo = get_object_or_404(BusinessLegalInfo, pk=pk)
        businessLegalInfo.delete()
        return Response({"success": True,
                         "message": "BusinessLegalInfo deleted successfully."},
                        status=status.HTTP_200_OK)

class DistributorInformationPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100
class DistributorInformationAPIView(APIView):

    def get(self, request, pk=None):

        # 🔹 Single distributor
        if pk:
            distributor = get_object_or_404(DistributorInformation, pk=pk)
            serializer = DistributorInformationSerializer(distributor)

            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔹 Queryset (IMPORTANT: add ordering)
        qs = DistributorInformation.objects.all().order_by("-id")

        # 🔹 Filters
        status_filter = request.query_params.get('status', "").strip()
        region_filter = request.query_params.get('sales_region', "").strip()
        state_filter = request.query_params.get('state', "").strip()
        city_filter = request.query_params.get('city', "").strip()

        if status_filter:
            qs = qs.filter(status__iexact=status_filter)

        if region_filter:
            qs = qs.filter(sales_region__icontains=region_filter)

        if state_filter:
            qs = qs.filter(state__icontains=state_filter)

        if city_filter:
            qs = qs.filter(city__icontains=city_filter)

        # 🔹 Pagination
        paginator = DistributorInformationPagination()
        paginated_qs = paginator.paginate_queryset(qs, request)

        serializer = DistributorInformationSerializer(paginated_qs, many=True)

        return Response({
            "success": True,
            "count": qs.count(),
            "message": "Distributor Fetched sucessfully",
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = DistributorInformationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Distributor Information created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk=None):
        if not pk:
            return Response(
                {"success": False, "message": "Distributor ID is required for update."},
                status=status.HTTP_400_BAD_REQUEST
            )

        distributor = get_object_or_404(DistributorInformation, pk=pk)
        serializer = DistributorInformationSerializer(
            distributor,
            data=request.data,
            partial=True
        )

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Distributor updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk=None):
        if not pk:
            return Response(
                {"success": False, "message": "Distributor ID is required for deletion."},
                status=status.HTTP_400_BAD_REQUEST
            )

        distributor = get_object_or_404(DistributorInformation, pk=pk)
        distributor.delete()

        return Response({
            "success": True,
            "message": "Distributor information deleted successfully."
        }, status=status.HTTP_200_OK)

class SerialAPIView(APIView):
    def get(self, request, pk=None):

        # 🔹 Single attendance
        if pk:
            serial = get_object_or_404(Serial, pk=pk)
            serializer = SerialSerializer(serial)
            return Response({
                "success": True,
                "message": "serial fetched successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔹 All attendance list
        serializer = Serial.objects.all().order_by('-created_at')
        serializer = SerialSerializer(serializer, many=True)

        return Response({
            "success": True,
            "message": "serial list fetched successfully",
            # "count": serial.count(),
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = SerialSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Serial Information created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk=None):
        if not pk:
            return Response(
                {"success": False,
                 "message": "Serial ID is required for update."},
                status=status.HTTP_400_BAD_REQUEST)

        serial = get_object_or_404(Serial, pk=pk)
        serializer = SerialSerializer(
            serial,
            data=request.data,
            partial=True
        )

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Serial data updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk=None):
        if not pk:
            return Response(
                {"success": False,
                 "message": "Serial ID is required for deletion."},
                status=status.HTTP_400_BAD_REQUEST
            )

        serial = get_object_or_404(Serial, pk=pk)
        serial.delete()

        return Response({
            "success": True,
            "message": "Serial data deleted successfully."
        }, status=status.HTTP_200_OK)


class DistributorLoginView(APIView):
    def post(self, request):
        serializer = DistributorLoginSerializer(data=request.data)

        if serializer.is_valid():
            distributor = serializer.validated_data

            refresh = RefreshToken.for_user(distributor)

            return Response({
                "message": "Login successful",
                "refresh": str(refresh),
                "access": str(refresh.access_token),
                "distributor_id": distributor.id,
                "status": distributor.status,
                "email_id": distributor.email_id
            }, status=status.HTTP_200_OK)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)



# ---------------- CREATE DISTRIBUTOR ----------------
class DistributorCreateAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = DistributorCreateSerializer(data=request.data)

        if serializer.is_valid():
            distributor = serializer.save()
            return Response({
                "success": True,
                "message": "Distributor created successfully",
                "data": {
                    "id": distributor.id,
                    "email_id": distributor.email_id,
                    "status": distributor.status
                }
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


# ---------------- LOGIN DISTRIBUTOR ----------------
@method_decorator(csrf_exempt, name="dispatch")
class DistributorLoginAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = DistributorLoginSerializer(data=request.data)

        if serializer.is_valid():
            distributor = serializer.validated_data

            refresh = RefreshToken.for_user(distributor)
            refresh["distributor_id"] = distributor.id
            refresh["email_id"] = distributor.email_id
            refresh["is_distributor"] = True

            return Response({
                "message": "Login successful",
                "refresh": str(refresh),
                "access": str(refresh.access_token),
                "distributor_id": distributor.id,
                "status": distributor.status,
                "email_id": distributor.email_id
            }, status=status.HTTP_200_OK)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


# def generate_warranty_bill(warranty):
import os
from django.conf import settings
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
)


def draw_page_border(canv, doc):
    canv.saveState()

    page_width, page_height = A4

    x = 6 * mm
    y = 6 * mm
    w = page_width - 12 * mm
    h = page_height - 12 * mm

    canv.setStrokeColor(colors.HexColor("#D8D8D8"))
    canv.setLineWidth(1)
    canv.roundRect(x, y, w, h, 6 * mm, stroke=1, fill=0)

    canv.restoreState()


def generate_warranty_bill(warranty):
    file_name = f"warranty_bill_{warranty.id}.pdf"
    file_path = os.path.join(settings.MEDIA_ROOT, file_name)

    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception:
            pass

    page_width, page_height = A4
    left = 12 * mm
    right = 12 * mm
    top = 10 * mm
    bottom = 10 * mm
    usable_width = page_width - left - right

    doc = SimpleDocTemplate(
        file_path,
        pagesize=A4,
        leftMargin=left,
        rightMargin=right,
        topMargin=top,
        bottomMargin=bottom,
    )

    styles = getSampleStyleSheet()
    elements = []

    RED = colors.HexColor("#B31224")
    DARK = colors.HexColor("#151515")
    LIGHT_GREY = colors.HexColor("#D9D9D9")
    MID_GREY = colors.HexColor("#777777")
    PALE_YELLOW = colors.HexColor("#FBF3DC")
    WHITE = colors.white
    BLACK = colors.black

    brand_style = ParagraphStyle(
        name="brand_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=24,
        leading=26,
        textColor=RED,
        alignment=TA_LEFT,
        spaceAfter=2,
    )

    sub_brand_style = ParagraphStyle(
        name="sub_brand_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=12,
        textColor=BLACK,
        alignment=TA_LEFT,
        spaceAfter=4,
    )

    address_style = ParagraphStyle(
        name="address_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=10.5,
        textColor=BLACK,
        alignment=TA_LEFT,
        spaceAfter=8,
    )

    contact_left_style = ParagraphStyle(
        name="contact_left_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        textColor=WHITE,
        alignment=TA_LEFT,
    )

    contact_right_style = ParagraphStyle(
        name="contact_right_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        textColor=WHITE,
        alignment=TA_RIGHT,
    )

    title_bar_style = ParagraphStyle(
        name="title_bar_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=17,
        textColor=WHITE,
        alignment=TA_CENTER,
    )

    section_style = ParagraphStyle(
        name="section_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=14,
        textColor=RED,
        alignment=TA_LEFT,
        spaceAfter=3,
        spaceBefore=2,
    )

    bullet_style = ParagraphStyle(
        name="bullet_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9.8,
        leading=13.2,
        textColor=BLACK,
        alignment=TA_LEFT,
        spaceAfter=3,
    )

    claim_style = ParagraphStyle(
        name="claim_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=11.5,
        textColor=BLACK,
        alignment=TA_LEFT,
    )

    detail_label_style = ParagraphStyle(
        name="detail_label_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.7,
        leading=10.5,
        textColor=MID_GREY,
        alignment=TA_LEFT,
    )

    detail_value_style = ParagraphStyle(
        name="detail_value_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9.6,
        leading=11,
        textColor=BLACK,
        alignment=TA_LEFT,
    )

    owner_header_style = ParagraphStyle(
        name="owner_header_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=17,
        textColor=WHITE,
        alignment=TA_CENTER,
    )

    sign_header_style = ParagraphStyle(
        name="sign_header_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.7,
        leading=10.5,
        textColor=WHITE,
        alignment=TA_LEFT,
    )

    sign_text_style = ParagraphStyle(
        name="sign_text_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.7,
        leading=10.5,
        textColor=MID_GREY,
        alignment=TA_LEFT,
    )

    sign_value_style = ParagraphStyle(
        name="sign_value_style",
        parent=styles["Normal"],
        fontName="Helvetica-BoldOblique",
        fontSize=13,
        leading=15,
        textColor=BLACK,
        alignment=TA_CENTER,
    )

    footer_text_style = ParagraphStyle(
        name="footer_text_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=10,
        textColor=BLACK,
        alignment=TA_CENTER,
    )

    footer_bar_style = ParagraphStyle(
        name="footer_bar_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.2,
        leading=10,
        textColor=WHITE,
        alignment=TA_CENTER,
    )

    def safe(value, default=""):
        return str(value).strip() if value not in [None, "None", ""] else default

    def fmt_date(value):
        try:
            return value.strftime("%d / %m / %Y") if value else ""
        except Exception:
            return ""

    def divider():
        t = Table([[""]], colWidths=[usable_width], rowHeights=[0.8])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), LIGHT_GREY),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        return t

    def bullet_list(items):
        return [Paragraph(f"•&nbsp;&nbsp;{item}", bullet_style) for item in items]

    def two_cols(left_items, right_items):
        t = Table(
            [[bullet_list(left_items), bullet_list(right_items)]],
            colWidths=[usable_width * 0.51, usable_width * 0.49],
        )
        t.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        return t

    def red_title_bar(text):
        t = Table(
            [[Paragraph(text, title_bar_style)]],
            colWidths=[usable_width],
            rowHeights=[16 * mm]
        )
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), RED),
            ("BOX", (0, 0), (-1, -1), 1, RED),
            ("ROUNDEDCORNERS", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ]))
        return t

    def owner_vehicle_header():
        left_red = Table([[""]], colWidths=[3 * mm], rowHeights=[13 * mm])
        left_red.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), RED),
            ("BOX", (0, 0), (-1, -1), 0, RED),
            ("ROUNDEDCORNERS", (0, 0), (-1, -1), 2),
        ]))

        right_black = Table(
            [[Paragraph("OWNER & VEHICLE DETAILS", owner_header_style)]],
            colWidths=[usable_width - 4 * mm],
            rowHeights=[13 * mm]
        )
        right_black.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), DARK),
            ("BOX", (0, 0), (-1, -1), 1, DARK),
            ("ROUNDEDCORNERS", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        wrap = Table([[left_red, right_black]], colWidths=[4 * mm, usable_width - 4 * mm])
        wrap.setStyle(TableStyle([
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return wrap

    # =========================
    # DYNAMIC DATA FROM WARRANTY
    # =========================
    owner_name = safe(getattr(warranty, "owner_name", "")) or safe(getattr(warranty, "registered_by", ""))
    owner_mobile = safe(getattr(warranty, "owner_mobile", ""))
    owner_email = safe(getattr(warranty, "owner_email", "")) or safe(getattr(warranty, "email", ""))

    if owner_mobile and owner_email:
        owner_contact = f"{owner_mobile} / {owner_email}"
    else:
        owner_contact = owner_mobile or owner_email or ""

    serial_number = safe(getattr(getattr(warranty, "serial_id", None), "serial_number", ""))
    car_no = safe(getattr(warranty, "license_plate_no", "")) or safe(getattr(warranty, "car_registration_number", ""))
    vehicle_reg = safe(getattr(warranty, "car_registration_number", ""))

    brand = safe(getattr(warranty, "car_brand", ""))
    model = safe(getattr(warranty, "car_model", ""))
    make_model = f"{brand} {model}".strip()

    color_name = safe(getattr(warranty, "color", ""))
    product_name = safe(getattr(getattr(warranty, "product_id", None), "product_name", ""))
    install_date = fmt_date(getattr(warranty, "installation_date", None))

    warranty_period = f"{warranty.warranty_period} Months" if getattr(warranty, "warranty_period", None) else ""
    start_date = fmt_date(getattr(warranty, "warranty_start_date", None))
    end_date = fmt_date(getattr(warranty, "warranty_end_date", None))

    detailer_name = safe(getattr(warranty, "detailer_name", ""))
    detailer_address = safe(getattr(warranty, "address", ""))

    if not detailer_address and getattr(warranty, "distributor_id", None):
        distributor = warranty.distributor_id
        address_parts = [
            safe(getattr(distributor, "shop_name", "")),
            safe(getattr(distributor, "address_line_1", "")),
            safe(getattr(distributor, "address_line_2", "")),
            safe(getattr(distributor, "city", "")),
            safe(getattr(distributor, "state", "")),
            safe(getattr(distributor, "pincode", "")),
        ]
        detailer_address = ", ".join([i for i in address_parts if i])

    # =========================
    # SIGNATURE IMAGE
    # =========================
    signature_image = None
    if getattr(warranty, "signature_img", None):
        try:
            signature_path = warranty.signature_img.path
            if signature_path and os.path.exists(signature_path):
                signature_image = Image(signature_path, width=95, height=45)
                signature_image.hAlign = "RIGHT"
        except Exception:
            signature_image = None

    # =========================
    # HEADER
    # =========================
    elements.append(Paragraph("HOGO AUTO FILM", brand_style))
    elements.append(Paragraph("Premium Automotive Protection", sub_brand_style))
    elements.append(Paragraph("Email: hogonn@hogonnfilm@gmail.com.hk", address_style))
    elements.append(Paragraph("Website: www.hogonnfilm.com.hk", address_style))

    elements.append(red_title_bar("PAINT PROTECTION FILM - WARRANTY CERTIFICATE"))
    elements.append(Spacer(1, 10))

    elements.append(Paragraph("WARRANTY COVERAGE", section_style))
    elements.append(divider())
    elements.append(two_cols(
        [
            "No Bubbling — Film bubbles will not form under film.",
            "Yellowing — No signs of yellowing.",
            "Peeling — Film will not peel off of surface.",
            "Staining — No permanent stains.",
            "Cracking / Blistering — PPF will not separate from or crack.",
            "Clear Coat Failure — No fading of the clear coat under film.",
            "Delamination / Adhesive Failure — Covered under normal usage.",
        ],
        [
            "Accidents, abuse, chips, vandalism, or impact damage.",
            "Improper care/cleaning, or neglect damage.",
            "Intentional removal or misuse.",
            "Chemical spills, stains, or contamination.",
            "Vehicle modifications affecting film performance.",
            "Failure due to acts of nature.",
        ]
    ))
    elements.append(Spacer(1, 7))

    elements.append(Paragraph("WARRANTY DOES NOT COVER", section_style))
    elements.append(divider())
    elements.append(two_cols(
        [
            "Accidents, abuse, chips, vandalism, or impact damage.",
            "Improper care/cleaning, or neglect damage.",
            "Intentional removal caused by misuse.",
            "Chemical spills, stains, or contamination.",
            "Vehicle modifications that disrupt the warranty.",
        ],
        [
            "Do not wash vehicle for 48 hours after installation.",
            "Avoid harsh contact with sharp objects.",
            "Use pH neutral car shampoo only.",
            "Avoid abrasive cleaning compounds.",
            "Regular maintenance is recommended.",
        ]
    ))
    elements.append(Spacer(1, 7))

    elements.append(Paragraph("CARE & MAINTENANCE GUIDELINES", section_style))
    elements.append(divider())
    elements.append(two_cols(
        [
            "Do not wash vehicle for 48 hours after installation.",
            "Avoid physical contact or parking close to rough objects.",
            "Use pH neutral car shampoo only.",
            "Avoid abrasive cleaning compounds.",
        ],
        [
            "Report claim to authorized installer.",
            "Inspection will be conducted at scheduled site.",
            "Final determination depends on manufacturer diagnosis.",
            "Substantiating costs are not covered.",
        ]
    ))
    elements.append(Spacer(1, 7))

    elements.append(Paragraph("WARRANTY CLAIM PROCESS", section_style))
    claim_tbl = Table(
        [[
            Paragraph(
                "PPF warranty is non-transferable to second vehicle. Manufacturer is not liable for indirect or consequential damages.<br/><br/>"
                "<b>Customer Registration:</b> Please register product with an authorized installer. Check warranty through support channel or online system within 7 business days of delivery.",
                claim_style
            )
        ]],
        colWidths=[usable_width]
    )
    claim_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), PALE_YELLOW),
        ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#E5D7A4")),
        ("ROUNDEDCORNERS", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(claim_tbl)
    elements.append(Spacer(1, 8))

    # =========================
    # OWNER & VEHICLE DETAILS ONLY
    # =========================
    elements.append(owner_vehicle_header())
    elements.append(Spacer(1, 4))

    details_data = [
        [
            Paragraph("• Owner's Name", detail_label_style),
            Paragraph(owner_name, detail_value_style),
            Paragraph("• Owner's Contact Information", detail_label_style),
            Paragraph(owner_contact, detail_value_style),
        ],
        [
            Paragraph("• License Plate No. / VIN", detail_label_style),
            Paragraph(f"{car_no} / {serial_number}", detail_value_style),
            Paragraph("• Vehicle Registration Number", detail_label_style),
            Paragraph(vehicle_reg, detail_value_style),
        ],
        [
            Paragraph("• Installed Vehicle Make & Model", detail_label_style),
            Paragraph(make_model, detail_value_style),
            Paragraph("• Vehicle Color", detail_label_style),
            Paragraph(color_name, detail_value_style),
        ],
        [
            Paragraph("• PPF Product Model", detail_label_style),
            Paragraph(product_name, detail_value_style),
            Paragraph("• Serial Number", detail_label_style),
            Paragraph(serial_number, detail_value_style),
        ],
        [
            Paragraph("• Warranty Period", detail_label_style),
            Paragraph(warranty_period, detail_value_style),
            Paragraph("• Installation Date", detail_label_style),
            Paragraph(install_date, detail_value_style),
        ],
        [
            Paragraph("• Warranty Start Date", detail_label_style),
            Paragraph(start_date, detail_value_style),
            Paragraph("• Warranty End Date", detail_label_style),
            Paragraph(end_date, detail_value_style),
        ],
        [
            Paragraph("• Installation Detailer Name", detail_label_style),
            Paragraph(detailer_name, detail_value_style),
            Paragraph("• Detailer Address", detail_label_style),
            Paragraph(detailer_address, detail_value_style),
        ],
    ]

    details_tbl = Table(
        details_data,
        colWidths=[
            usable_width * 0.20,
            usable_width * 0.30,
            usable_width * 0.20,
            usable_width * 0.30,
        ],
    )
    details_tbl.setStyle(TableStyle([
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, LIGHT_GREY),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(details_tbl)
    elements.append(Spacer(1, 6))

    # =========================
    # DISTRIBUTOR / INSTALLER SIGNATURE
    # IMAGE ON RIGHT SIDE
    # =========================
    distributor_sign = Table(
        [
            [Paragraph("DISTRIBUTOR / INSTALLER SIGNATURE", sign_header_style), ""],
            [
                Paragraph("Authorised Dealer Stamp & Sign", sign_text_style),
                signature_image if signature_image else Spacer(1, 45)
            ],
            [
                Paragraph(f"Date:&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;<b>{install_date}</b>", sign_text_style),
                ""
            ],
        ],
        colWidths=[usable_width * 0.58, usable_width * 0.42]
    )

    distributor_sign.setStyle(TableStyle([
        ("SPAN", (0, 0), (1, 0)),
        ("BACKGROUND", (0, 0), (1, 0), RED),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#C99797")),

        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),

        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 1), (1, 1), "RIGHT"),
    ]))

    sign_wrap = Table(
        [[distributor_sign]],
        colWidths=[usable_width]
    )
    sign_wrap.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(sign_wrap)
    elements.append(Spacer(1, 7))

    elements.append(Spacer(1, 3))

    footer_tbl = Table(
        [[Paragraph("HOGON AUTOFILM  |  https://hogoautofilms.co.in/", footer_bar_style)]],
        colWidths=[usable_width],
        rowHeights=[9 * mm]
    )
    footer_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), DARK),
        ("BOX", (0, 0), (-1, -1), 1, DARK),
        ("ROUNDEDCORNERS", (0, 0), (-1, -1), 3),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(footer_tbl)

    doc.build(
        elements,
        onFirstPage=draw_page_border,
        onLaterPages=draw_page_border
    )

    return f"/media/{file_name}"

class DistributorProfileAPIView(APIView):

    # Helper to get distributor from JWT token
    def get_distributor(self, request):
        auth = request.headers.get("Authorization")

        if not auth or not auth.startswith("Bearer "):
            raise AuthenticationFailed("Authorization header missing")

        token_str = auth.split(" ")[1]

        try:
            token = AccessToken(token_str)
        except Exception:
            raise AuthenticationFailed("Invalid or expired token")

        distributor_id = token.get("distributor_id")
        if not distributor_id:
            raise AuthenticationFailed("distributor_id not found in token")

        try:
            return DistributorInformation.objects.get(id=distributor_id)
        except DistributorInformation.DoesNotExist:
            raise AuthenticationFailed("Distributor not found")

    # ---------------- GET ----------------
    def get(self, request):
        distributor = self.get_distributor(request)
        return Response({
            "success": True,
            "message":"Distributor Profile list fetched successfully",
            "data": {
                "id": distributor.id,
                "email_id": distributor.email_id,
                "status": distributor.status,
                "created_at": distributor.created_at
            }
        })

    # ---------------- PATCH ----------------
    def patch(self, request):
        distributor = self.get_distributor(request)
        serializer = DistributorCreateSerializer(distributor, data=request.data, partial=True)

        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data

        # Hash password if being updated
        if "password" in validated_data:
            validated_data["password"] = make_password(validated_data["password"])

        serializer.save()

        return Response({
            "success": True,
            "message": "Profile updated successfully",
            "data": {
                "id": distributor.id,
                "email_id": distributor.email_id,
                "status": distributor.status
            }
        })

    # ---------------- DELETE ----------------
    def delete(self, request):
        distributor = self.get_distributor(request)
        distributor.delete()
        return Response({
            "success": True,
            "message": "Distributor account deleted successfully"
        })



class WarrantyAPI(APIView):

    parser_classes = (JSONParser, MultiPartParser, FormParser)
 
    # helper

    def get_warranty_object(self, value):

        if str(value).isdigit():

            return get_object_or_404(Warranty, pk=value)

        return get_object_or_404(Warranty, serial_id__serial_number=value)
 
    # email helper

    def send_warranty_email(self, to_email, subject, body, bill_url=None, html=False):

        try:

            if not to_email:

                print(" No recipient email found.")

                return False
 
            email = EmailMessage(

                subject=subject,

                body=body,

                from_email=settings.EMAIL_HOST_USER,

                to=[to_email],

            )
 
            if html:

                email.content_subtype = "html"
 
            if bill_url:

                relative_path = str(bill_url).replace("/media/", "").lstrip("/")

                file_path = os.path.join(settings.MEDIA_ROOT, relative_path)
 
                print("DEBUG bill_url:", bill_url)

                print("DEBUG file_path:", file_path)
 
                if os.path.exists(file_path):

                    email.attach_file(file_path)

                else:

                    print(" Attachment file not found:", file_path)
 
            email.send(fail_silently=False)

            print(f" mail sent successfully to {to_email}")

            return True
 
        except Exception:

            print(" mail sending error:")

            print(traceback.format_exc())

            return False
 
    # ---------- GET ----------

    def get(self, request, value=None):
 
        if value:

            warranty = self.get_warranty_object(value)

            search_type = "pk" if str(value).isdigit() else "serial_number"

            serializer = WarrantySerializer(

                warranty,

                context={"request": request}

            )

            bill_url = (

                f"/media/warranty_bill_{warranty.id}.pdf"

                if warranty.generate_bill else None

            )

            return Response({

                "success": True,

                "search_type": search_type,

                "generate_bill_url": bill_url,

                "data": serializer.data

            }, status=status.HTTP_200_OK)

        distributor_id = request.query_params.get("distributor_id")

        serial_id = request.query_params.get("serial_id")

        serial_number = request.query_params.get("serial_number")

        warranty_status = request.query_params.get("warranty_status")

        warranties = Warranty.objects.all().order_by("-id")

        if distributor_id:

            warranties = warranties.filter(

                distributor_id=int(distributor_id)

            )

        if serial_id:

            warranties = warranties.filter(

                serial_id=int(serial_id)

            )

        # Filter by Inventory Serial Number

        if serial_number:

            warranties = warranties.filter(

                serial_id__serial_number__icontains=serial_number

            )

        # Filter by Warranty Status

        if warranty_status:

            warranties = warranties.filter(

                warranty_status__iexact=warranty_status

            )

        serializer = WarrantySerializer(

            warranties,

            many=True,

            context={"request": request}

        )

        return Response({

            "success": True,

            "message": "Warranty list fetched successfully",

            "count": warranties.count(),

            "data": serializer.data

        }, status=status.HTTP_200_OK)
 
    # ---------- POST ----------

    def post(self, request):

        serializer = WarrantySerializer(

            data=request.data,

            context={"request": request}

        )
 
        if serializer.is_valid():

            warranty = serializer.save()
 
            # -------- CAR IMAGES --------

            car_files = request.FILES.getlist("car_images")

            if car_files:

                warranty.car_images = [

                    default_storage.save(f"car_images/{img.name}", img)

                    for img in car_files

                ]
 
            # -------- INSTALLATION IMAGES --------

            install_files = request.FILES.getlist("installation_images")

            if install_files:

                warranty.installation_images = [

                    default_storage.save(f"installation_images/{img.name}", img)

                    for img in install_files

                ]
 
            bill_url = None

            generate_flag = request.data.get("generate_bill")
 
            # -------- GENERATE BILL ONLY --------

            if str(generate_flag).lower() == "true":

                warranty.generate_bill = True

                warranty.save()
 
                # 👉 Only generate PDF (NO EMAIL)

                bill_url = generate_warranty_bill(warranty)
 
            warranty.save()
 
            return Response({

                "success": True,

                "message": "Warranty created successfully",

                "generate_bill_url": bill_url,

                "data": WarrantySerializer(warranty, context={"request": request}).data

            }, status=status.HTTP_201_CREATED)
 
        return Response({

            "status":"Invalid data",

            "errors": serializer.errors

        }, status=status.HTTP_400_BAD_REQUEST)
 
    # ---------- PATCH ----------

    def patch(self, request, value):

        warranty = self.get_warranty_object(value)
 
        old_warranty_status = warranty.warranty_status

        old_product_status = warranty.product_status
 
        serializer = WarrantySerializer(

            warranty,

            data=request.data,

            partial=True,

            context={"request": request}

        )
 
        if serializer.is_valid():

            warranty = serializer.save()
 
            car_files = request.FILES.getlist("car_images")

            if car_files:

                warranty.car_images = [

                    default_storage.save(f"car_images/{img.name}", img)

                    for img in car_files

                ]
 
            install_files = request.FILES.getlist("installation_images")

            if install_files:

                warranty.installation_images = [

                    default_storage.save(f"installation_images/{img.name}", img)

                    for img in install_files

                ]
 
            bill_url = None

            generate_flag = request.data.get("generate_bill")
 
            if str(generate_flag).lower() == "true":

                warranty.generate_bill = True

                bill_url = generate_warranty_bill(warranty)
 
            warranty.save()
 
            # send mail only when newly approved/activated

            is_new_active = (

                warranty.warranty_status == "ACTIVE" and

                old_warranty_status != "ACTIVE"

            )

            is_new_activated = (

                warranty.product_status == "ACTIVATED" and

                old_product_status != "ACTIVATED"

            )
 
            if is_new_active or is_new_activated:

                try:

                    customer_name = warranty.owner_name or warranty.registered_by or "Hogo Customer"

                    product_name = warranty.product_id.product_name if warranty.product_id else "Product"

                    to_email = warranty.owner_email or warranty.email
 
                    print("DEBUG PATCH owner_email:", warranty.owner_email)

                    print("DEBUG PATCH email:", warranty.email)

                    print("DEBUG PATCH final recipient:", to_email)
 
                    if to_email:

                        if not bill_url:

                            bill_url = generate_warranty_bill(warranty)
 
                        subject = "Warranty Activation Confirmation"
 
                        html_body = f"""
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 20px; background-color: #f4f4f4;">
<div style="max-width: 600px; margin: 0 auto; background: #fff; border: 1px solid #ddd; border-radius: 8px; overflow: hidden;">
<div style="background-color: #000; padding: 30px; text-align: center; color: #fff;">
<h1 style="margin: 0;">HOGO AUTOFILM</h1>
<p style="margin: 5px 0 0; font-size: 14px; color: #aaa;">Premium Protection</p>
</div>
 
                                <div style="padding: 40px;">
<h2 style="color: #2563eb; margin-top: 0;">Warranty Activation Confirmed</h2>
 
                                    <p>Dear <strong>{customer_name}</strong>,</p>
 
                                    <p>Your warranty request for <strong>{product_name}</strong> has been successfully approved and is now active.</p>
 
                                    <p>Please find your official Warranty Certificate attached to this email.</p>
 
                                    <p>We recommend keeping this certificate safe for any future service or claims.</p>
 
                                    <p>Warm regards,<br><strong>HOGO AUTOFILM Team</strong></p>
</div>
 
                                <div style="background-color: #1e293b; padding: 20px; text-align: center; font-size: 12px; color: #94a3b8;">

                                    © {datetime.now().year} HOGO AUTOFILM. All rights reserved.
</div>
</div>
</body>
</html>

                        """
 
                        self.send_warranty_email(

                            to_email=to_email,

                            subject=subject,

                            body=html_body,

                            bill_url=bill_url,

                            html=True

                        )

                    else:

                        print("❌ No recipient email found for warranty activation.")
 
                except Exception:

                    print("❌ Error sending activation email:")

                    print(traceback.format_exc())
 
            return Response({

                "success": True,

                "message": "Warranty updated successfully",

                "generate_bill_url": bill_url,

                "data": WarrantySerializer(warranty, context={"request": request}).data

            }, status=status.HTTP_200_OK)
 
        return Response({

            "success": False,

            "errors": serializer.errors

        }, status=status.HTTP_400_BAD_REQUEST)
 
    # ---------- DELETE ----------

    def delete(self, request, value):

        warranty = self.get_warranty_object(value)

        warranty.delete()
 
        return Response({

            "success": True,

            "message": "Warranty deleted successfully"

        }, status=status.HTTP_204_NO_CONTENT)
 





class EmployeeSalaryPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class EmployeeSalaryAPIView(APIView):

    # ---------- GET (All or Single) ----------
    def get(self, request, pk=None, id=None):

        # ✅ Fix: support both pk and id
        record_id = pk if pk is not None else id

        # 🔹 Single
        if record_id:
            employee_salary = get_object_or_404(EmployeeSalary, pk=record_id)
            serializer = EmployeeSalarySerializer(employee_salary)

            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔹 Queryset
        qs = EmployeeSalary.objects.select_related("employee_id").all().order_by("-id")

        # 🔹 Filter by employee_id
        employee_id = request.query_params.get("employee_id", "").strip()
        if employee_id:
            qs = qs.filter(employee_id=employee_id)

        # 🔹 ✅ Filter by status
        status_param = request.query_params.get('status')
        if status_param:
            if status_param.lower() == 'true':
                qs = qs.filter(status=True)
            elif status_param.lower() == 'false':
                qs = qs.filter(status=False)

        # 🔹 Pagination
        paginator = EmployeeSalaryPagination()
        paginated_qs = paginator.paginate_queryset(qs, request)

        serializer = EmployeeSalarySerializer(paginated_qs, many=True)

        return Response({
            "success": True,
            "count": qs.count(),
            "message": "Employee-salary Fatch sucessfully",
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        }, status=status.HTTP_200_OK)
    # ---------- POST (Create) ----------#
    def post(self, request):
        serializer = EmployeeSalarySerializer(data=request.data)
        if serializer.is_valid():
            salary_instance = serializer.save()

            # Automatically deactivate older salary records for the same employee
            if salary_instance.employee_id:
                from .models import EmployeeSalary
                EmployeeSalary.objects.filter(
                    employee_id=salary_instance.employee_id
                ).exclude(id=salary_instance.id).update(status=False)

            return Response({
                "success": True,
                "message": "Employee salary created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- PATCH (Update) ----------#
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        employee_salary = get_object_or_404(EmployeeSalary, pk=pk)
        serializer = EmployeeSalarySerializer(
            employee_salary, data=request.data, partial=True
        )

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Employee salary updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- DELETE ----------#
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "Id is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        employee_salary = get_object_or_404(EmployeeSalary, pk=pk)
        employee_salary.delete()

        return Response({
            "success": True,
            "message": "Employee salary deleted successfully"
        }, status=status.HTTP_200_OK)

class employeepersonaldetailsView(APIView):

    def get(self, request, pk=None):
        if pk:
            employeepersonaldetails = get_object_or_404(Employee_personal_details, pk=pk)
            serializer = EmployeePersonalDetailsSerializer(employeepersonaldetails)
            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔹 Get employee_id from query params
        employee_id = request.GET.get("employee_id")

        employeepersonaldetails = Employee_personal_details.objects.all()

        # 🔹 Apply filter if employee_id is provided
        if employee_id:
            employeepersonaldetails = employeepersonaldetails.filter(employee_id=employee_id)

        serializer = EmployeePersonalDetailsSerializer(employeepersonaldetails, many=True)

        return Response({
            "success": True,
            "message":"Employee Personal Detail list fetched successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)


    # ---------- POST ----------
    def post(self, request):
        serializer = EmployeePersonalDetailsSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(
                {"success": True, "data": serializer.data},
                status=status.HTTP_201_CREATED
            )

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        employee = get_object_or_404(Employee_personal_details, pk=pk)
        serializer = EmployeePersonalDetailsSerializer(
            employee,
            data=request.data,
            partial=True
        )

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Employee updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    def delete(self, request, pk):  # 👈 MUST accept `id`
        instance = get_object_or_404(Employee_personal_details, pk=pk)
        instance.delete()

        return Response({
            "success": True,
            "message": "Employee personal details deleted successfully"
        }, status=status.HTTP_200_OK)

class EmployeeDocumentsAPIView(APIView):

    def get(self, request, pk=None):

        # Single document
        if pk:
            document = get_object_or_404(EmployeeDocuments, pk=pk)
            serializer = EmployeeDocumentsSerializer(document)
            return Response({
                "success": True,
                "message": "Employee document fetched successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔥 Filter by employee_id (NEW)
        employee_id = request.GET.get('employee_id')

        documents = EmployeeDocuments.objects.all().order_by('-uploaded_at')

        if employee_id:
            documents = documents.filter(employee_id=employee_id)

        serializer = EmployeeDocumentsSerializer(documents, many=True)

        return Response({
            "success": True,
            "message": "Employee documents list fetched successfully",
            "count": documents.count(),
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = EmployeeDocumentsSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Employee documents uploaded successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


    def patch(self, request, pk=None):
        document = get_object_or_404(EmployeeDocuments, pk=pk)
        serializer = EmployeeDocumentsSerializer(document, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Employee document updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # DELETE
    def delete(self, request, pk=None):
        document = get_object_or_404(EmployeeDocuments, pk=pk)
        document.delete()
        return Response({
            "success": True,
            "message": "Employee document deleted successfully"
        }, status=status.HTTP_200_OK)



class UserAPI(APIView):

    # 🔹 GET (all users or single user)
    def get(self, request, pk=None):
        if pk:
            user = User.objects.filter(id=pk).first()
            if not user:
                return Response(
                    {"success": False, "message": "User not found"},
                    status=status.HTTP_404_NOT_FOUND
                )
            serializer = UserSerializer(user)
            return Response(
                {"success": True, "data": serializer.data},
                status=status.HTTP_200_OK
            )

        # 🔹 Get filter param
        employee_id = request.GET.get("employee_id")

        users = User.objects.all().order_by('-id')

        # 🔹 Apply filter
        if employee_id:
            users = users.filter(employee_id=employee_id)

        serializer = UserSerializer(users, many=True)

        return Response(
            {
                "success": True,
                "message":"User list fetched successfully",
                "count": users.count(),
                "data": serializer.data
            },
            status=status.HTTP_200_OK
        )

    # 🔹 POST (create user)
    def post(self, request):
        serializer = UserSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "User created successfully",
                    "data": serializer.data
                },
                status=status.HTTP_201_CREATED
            )

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # 🔹 PATCH (partial update)
    def patch(self, request, pk=None):
        # ✅ Check ID required
        if not pk:
            return Response(
                {"success": False, "message": "id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        user = User.objects.filter(id=pk).first()
        if not user:
            return Response(
                {"success": False, "message": "User not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = UserSerializer(user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "User updated successfully",
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )

        return Response(
            {"success": False, "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )

    # 🔹 DELETE
    def delete(self, request, pk=None):
        if not pk:
            return Response(
                {"success": False, "message": "id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        user = User.objects.filter(id=pk).first()
        if not user:
            return Response(
                {"success": False, "message": "User not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        user.delete()
        return Response(
            {"success": True, "message": "User deleted successfully"},
            status=status.HTTP_200_OK
        )



class InventorySerialPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


class InventorySerialAPI(APIView):

    def get(self, request, serial_id=None):


    # 🔹 Get single record
        if serial_id:
            try:
                serial = InventorySerial.objects.get(id=serial_id)
                serializer = InventorySerialSerializer(serial)
                return Response({
                    "success": True,
                    "message": "Inventory serial fetched successfully",
                    "data": serializer.data
                })
            except InventorySerial.DoesNotExist:
                return Response({
                    "success": False,
                    "message": "Inventory serial not found"
                }, status=status.HTTP_404_NOT_FOUND)

        # 🔹 Get all records
        queryset = InventorySerial.objects.all().order_by("-id")

        # 🔹 Query params
        product_id = request.GET.get("product_id")
        batch_id = request.GET.get("batch_id")
        serial_number = request.GET.get("serial_number")
        status_param = request.GET.get("status")
        sku = request.GET.get("sku")
        distributor_id = request.GET.get("distributor_id")  # ✅ NEW FILTER

        # 🔹 Filters
        if product_id:
            queryset = queryset.filter(product_id=product_id)

        if batch_id:
            queryset = queryset.filter(batch_id__icontains=batch_id)

        if serial_number:
            queryset = queryset.filter(serial_number__icontains=serial_number)

        if status_param:
            queryset = queryset.filter(status__iexact=status_param)

        if sku:
            queryset = queryset.filter(product_id__sku__icontains=sku)

        # ✅ Filter by distributor_id through DistributorOrders relation
        if distributor_id:
            queryset = queryset.filter(
                distributororders__distributor_id=distributor_id
            ).distinct()

        # 🔹 Pagination
        paginator = InventorySerialPagination()
        paginated_queryset = paginator.paginate_queryset(queryset, request)

        # 🔹 Serialize paginated data
        serializer = InventorySerialSerializer(paginated_queryset, many=True)

        return Response({
            "success": True,
            "message": "Inventory serial list fetched successfully",
            "count": queryset.count(),
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        })


    def post(self, request):
        """
        Create inventory serials according to Shipment_product quantity
        """
        serializer = InventorySerialSerializer(data=request.data)
        if serializer.is_valid():
            try:
                created_serials = serializer.save()
            except Exception as e:
                return Response({
                    "success": False,
                    "message": str(e)
                }, status=status.HTTP_400_BAD_REQUEST)

            # ✅ Generate barcodes for each created serial
            for s in created_serials:
                try:
                    generate_inventory_barcode(s, request=request)
                except Exception as bar_err:
                    print(f"Barcode generation failed for {s.serial_number}: {bar_err}")

            resp_serializer = InventorySerialSerializer(created_serials, many=True)
            return Response({
                "success": True,
                "message": f"{len(created_serials)} inventory serials created successfully",
                "data": resp_serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, serial_id):
        """
        Update a serial by ID
        """
        try:
            serial = InventorySerial.objects.get(id=serial_id)
        except InventorySerial.DoesNotExist:
            return Response({
                "success": False,
                "message": "Inventory serial not found"
            }, status=status.HTTP_404_NOT_FOUND)

        serializer = InventorySerialSerializer(serial, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Inventory serial updated successfully",
                "data": serializer.data
            })

        return Response({
            "success": False,
            "message": "Validation error",
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, serial_id):
        """
        Delete a serial by ID
        """
        try:
            serial = InventorySerial.objects.get(id=serial_id)
            serial.delete()
            return Response({
                "success": True,
                "message": "Inventory serial deleted successfully"
            })
        except InventorySerial.DoesNotExist:
            return Response({
                "success": False,
                "message": "Inventory serial not found"
            }, status=status.HTTP_404_NOT_FOUND)


class InventorySerialHistoryAPI(APIView):

    def get(self, request):
        """
        Filters:
        ?month=3&year=2026&product_id=1&status=IN_STOCK&distributor_id=2
        """

        queryset = InventorySerialHistory.objects.all().order_by("-id")

        month = request.GET.get("month")
        year = request.GET.get("year")
        product_id = request.GET.get("product_id")
        status_param = request.GET.get("status")

        distributor_id = request.GET.get("distributor_id")
        distributor_name = request.GET.get("distributor_name")

        download = request.GET.get("download")

        # 🔹 Filters
        if month:
            queryset = queryset.filter(month=month)

        if year:
            queryset = queryset.filter(year=year)

        if product_id:
            queryset = queryset.filter(product_id=product_id)

        if status_param:
            queryset = queryset.filter(status__iexact=status_param)

        # ✅ Distributor filter (IMPORTANT)
        if distributor_id:
            queryset = queryset.filter(distributor_id=distributor_id)

        if distributor_name:
            queryset = queryset.filter(distributor__name__icontains=distributor_name)

        # 🔽 Excel Download
        if download == "true":
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventory History"

            headers = [
                "ID", "Product", "Serial Number", "Status",
                "Batch ID", "Distributor", "Location",
                "Warehouse", "Snapshot Date"
            ]

            for col_num, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_num, value=header).font = Font(bold=True)

            for row_num, obj in enumerate(queryset, start=2):
                ws.cell(row=row_num, column=1, value=obj.id)
                ws.cell(row=row_num, column=2, value=obj.product.product_name if obj.product else "-")
                ws.cell(row=row_num, column=3, value=obj.serial_number)
                ws.cell(row=row_num, column=4, value=obj.status)
                ws.cell(row=row_num, column=5, value=obj.batch_id)

                # ✅ distributor added
                ws.cell(row=row_num, column=6, value=obj.distributor.name if obj.distributor else "-")

                ws.cell(row=row_num, column=7, value=obj.location.name if obj.location else "-")
                ws.cell(row=row_num, column=8, value=obj.warehouse.name if obj.warehouse else "-")
                ws.cell(row=row_num, column=9, value=str(obj.snapshot_date))

            report_dir = os.path.join(settings.MEDIA_ROOT, "reports")
            os.makedirs(report_dir, exist_ok=True)

            filename = f"inventory_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(report_dir, filename)
            wb.save(filepath)

            file_url = request.build_absolute_uri(f"{settings.MEDIA_URL}reports/{filename}")

            return Response({
                "success": True,
                "message": "Excel report generated",
                "file_url": file_url
            })

        # 🔹 Normal Response
        serializer = InventorySerialHistorySerializer(queryset, many=True)

        return Response({
            "success": True,
            "message": "Inventory serial history fetched successfully",
            "count": queryset.count(),
            "data": serializer.data
        })


    def post(self, request):
        """
        Manually trigger a snapshot of current inventory serials
        """
        today = timezone.now().date()
        month = request.data.get("month", today.month)
        year = request.data.get("year", today.year)
        snapshot_date = request.data.get("snapshot_date", today)

        # Optional: Check if already exists for this month/year
        if InventorySerialHistory.objects.filter(month=month, year=year).exists():
             # We can either delete or just return error. Let's allow update by deleting old one.
             InventorySerialHistory.objects.filter(month=month, year=year).delete()

        serials = InventorySerial.objects.all()
        history_records = []
        for s in serials:
            history_records.append(InventorySerialHistory(
                product=s.product_id,
                serial_number=s.serial_number,
                status=s.status,
                batch_id=s.batch_id,
                location=s.location,
                warehouse=s.warehouse,
                snapshot_date=snapshot_date,
                month=month,
                year=year
            ))

        InventorySerialHistory.objects.bulk_create(history_records)

        return Response({
            "success": True,
            "message": f"Snapshot created for {month}/{year}. {len(history_records)} records stored."
        }, status=status.HTTP_201_CREATED)


class InventorySerialScanDetailAPI(APIView):
    """
    Detail View for scanning a barcode.
    Provides comprehensive information: Product, Serial Info, History, and Purchase Order links.
    """
    def get(self, request, serial_number):
        serial = get_object_or_404(InventorySerial, serial_number=serial_number)
        product = serial.product_id

        # 1. Basic Serial & Product Info
        data = {
            "serial_number": serial.serial_number,
            "batch_id": serial.batch_id,
            "status": serial.status,
            "product": {
                "id": product.id,
                "product_name": product.product_name,
                "sku": product.sku,
                "mrp": float(product.mrp or 0),
            },
            "location": {"id": serial.location.id, "name": serial.location.name} if serial.location else None,
            "warehouse": {"id": serial.warehouse.id, "name": serial.warehouse.name} if serial.warehouse else None,
            "barcode_image": request.build_absolute_uri(serial.barcode_image.url) if serial.barcode_image else None,
            "created_at": serial.created_at,
        }

        # 2. Status History (from InventorySerialHistory)
        history_qs = InventorySerialHistory.objects.filter(
            serial_number=serial_number,
            product=product
        ).order_by("-id")

        data["history"] = [
            {
                "status": h.status,
                "date": h.snapshot_date,
                "location": h.location.name if h.location else None,
                "warehouse": h.warehouse.name if h.warehouse else None,
                "timestamp": h.created_at
            } for h in history_qs
        ]

        # 3. Purchase Order Link (via DistributerOrders)
        dist_order = DistributerOrders.objects.filter(inventory_serial_id=serial).first()
        if dist_order:
            po = dist_order.purchase_order_id
            data["purchase_order"] = {
                "id": po.id,
                "po_number": po.po_number,
                "status": po.status,
                "date": po.po_date,
                "distributor": po.distributor_id.distributor_name if po.distributor_id else None
            }
        else:
            data["purchase_order"] = None

        return Response({
            "success": True,
            "message": "Scale detail fetched successfully",
            "data": data
        }, status=status.HTTP_200_OK)



class ProductScanDetailView(APIView):
    """
    Production-level detail view for scanning barcodes.
    Supports clean URLs: /product/?barcode=XXX
    Returns HTML for browser scans and JSON for API calls.
    """

    def get(self, request, serial_number=None):
        # 1. Get serial number
        sn = serial_number or request.GET.get('barcode')

        if not sn:
            return Response({
                "success": False,
                "message": "Serial number or barcode parameter is required."
            }, status=status.HTTP_400_BAD_REQUEST)

        # 2. Fetch object
        serial = get_object_or_404(InventorySerial, serial_number=sn)

        # ✅ ✅ FIX: Generate barcode if not exists
        try:
            if not serial.barcode_image:
                generate_inventory_barcode(serial, request)
        except Exception as e:
            print("Barcode generation error:", e)   # debug safe

        product = serial.product_id

        # 3. Prepare Data
        data = {
            "serial_number": serial.serial_number,
            "batch_id": serial.batch_id,
            "status": serial.status,
            "barcode_image": request.build_absolute_uri(serial.barcode_image.url) if serial.barcode_image else None,
            "product": {
                "id": product.id,
                "product_name": product.product_name,
                "sku": product.sku,
                "mrp": float(product.mrp or 0),
            },
            "location": {"id": serial.location.id, "name": serial.location.name} if serial.location else None,
            "warehouse": {"id": serial.warehouse.id, "name": serial.warehouse.name} if serial.warehouse else None,
            "created_at": serial.created_at,
        }

        # 4. History Timeline
        history_qs = InventorySerialHistory.objects.filter(
            serial_number=sn,
            product=product
        ).order_by("-id")

        data["history"] = history_qs

        # 5. Purchase Order
        dist_order = DistributerOrders.objects.filter(inventory_serial_id=serial).first()
        if dist_order:
            po = dist_order.purchase_order_id
            data["purchase_order"] = {
                "id": po.id,
                "po_number": po.po_number,
                "status": po.status,
                "date": po.po_date,
                "distributor": po.distributor_id.distributor_name if po.distributor_id else None
            }
        else:
            data["purchase_order"] = None

        # 6. Warranty
        warranty = serial.warranties.filter(product_status='ACTIVATED').first() or serial.warranties.first()
        if warranty:
            data["warranty"] = {
                "start_date": warranty.warranty_start_date,
                "end_date": warranty.warranty_end_date,
                "duration": warranty.warranty_period,
                "status": warranty.warranty_status
            }
        else:
            data["warranty"] = None

        # 7. Response type
        accept_header = request.headers.get('Accept', '')

        if 'text/html' in accept_header:
            return render(request, 'scanner.html', {'data': data})

        return Response({
            "success": True,
            "message":"ProductScanDetail list fetched successfully",
            "data": data
        }, status=status.HTTP_200_OK)





from django.db.models.functions import Coalesce
from dateutil.relativedelta import relativedelta   # ✅ ADD THIS IMPORT

class EmployeeAttendenceAPI(APIView):

    def get(self, request, pk=None):

        employee_id = request.GET.get("employee_id")
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        date = request.GET.get("date")
        month = request.GET.get("month")
        year = request.GET.get("year")

        prev_month = request.GET.get("prev_month")
        next_month = request.GET.get("next_month")


        # ✅ Support both names
        half_leave = request.GET.get("half_leave") or request.GET.get("half_day")
        full_leave = request.GET.get("full_leave") or request.GET.get("full_day")

        # 🔥 total hours filters
        min_hours = request.GET.get("min_hours")
        max_hours = request.GET.get("max_hours")

        # 🔥 summary flag
        total_flag = request.GET.get("total")

        # =========================
        # 🔹 SINGLE RECORD
        # =========================
        if pk:
            try:
                obj = Expenses_login.objects.select_related('employee').get(id=pk)
                serializer = ExpensesLoginSerializer(obj)
                return Response({
                    "success": True,
                    "message": "Record fetched",
                    "count": 1,
                    "data": serializer.data
                })
            except Expenses_login.DoesNotExist:
                return Response({
                    "success": False,
                    "message": "Record not found",
                    "count": 0,
                    "data": []
                }, status=status.HTTP_404_NOT_FOUND)

        # =========================
        # 🔹 MONTH FORMAT SUPPORT
        # =========================
        selected_month = None
        selected_year = None

        # ✅ ADDITION: PREVIOUS / NEXT MONTH LOGIC
        if prev_month == "true" or next_month == "true":
            today = timezone.now().date()

            if prev_month == "true":
                target_date = today - relativedelta(months=1)
            else:
                target_date = today + relativedelta(months=1)

            selected_month = target_date.month
            selected_year = target_date.year

        # 🔹 YOUR EXISTING LOGIC (UNCHANGED)
        if month:
            try:
                if "-" in month:
                    parts = month.split("-")
                    if len(parts) == 2:
                        selected_year = int(parts[0])
                        selected_month = int(parts[1])
                    else:
                        return Response({
                            "success": False,
                            "message": "Invalid month format. Use YYYY-MM or month with separate year."
                        }, status=status.HTTP_400_BAD_REQUEST)
                else:
                    selected_month = int(month)
                    selected_year = int(year) if year else timezone.now().year

                if selected_month < 1 or selected_month > 12:
                    return Response({
                        "success": False,
                        "message": "Month must be between 1 and 12"
                    }, status=status.HTTP_400_BAD_REQUEST)

            except ValueError:
                return Response({
                    "success": False,
                    "message": "Invalid month/year format"
                }, status=status.HTTP_400_BAD_REQUEST)
        elif year:
            try:
                selected_year = int(year)
            except ValueError:
                return Response({
                    "success": False,
                    "message": "Year must be a valid number"
                }, status=status.HTTP_400_BAD_REQUEST)

        # =========================
        # 🔥 MONTHLY TOTAL HOURS (ONLY WHEN total=true)
        # =========================
        if employee_id and (month or prev_month == "true" or next_month == "true") and total_flag == "true":
            total = Expenses_login.objects.filter(
                employee_id=employee_id,
                date__month=selected_month,
                date__year=selected_year
            ).aggregate(
                total_hours=Coalesce(Sum('total_hours'), Decimal('0.00'))
            )

            total_hours = Decimal(str(total["total_hours"] or 0)).quantize(Decimal("0.01"))

            # 🔥 Apply min/max on total
            if min_hours and total_hours < Decimal(str(min_hours)):
                return Response({
                    "success": False,
                    "message": "Total hours less than minimum filter",
                    "employee_id": int(employee_id),
                    "month": f"{selected_year}-{str(selected_month).zfill(2)}",
                    "total_hours": float(total_hours)
                })

            if max_hours and total_hours > Decimal(str(max_hours)):
                return Response({
                    "success": False,
                    "message": "Total hours greater than maximum filter",
                    "employee_id": int(employee_id),
                    "month": f"{selected_year}-{str(selected_month).zfill(2)}",
                    "total_hours": float(total_hours)
                })

            return Response({
                "success": True,
                "message": "Monthly total hours",
                "employee_id": int(employee_id),
                "month": f"{selected_year}-{str(selected_month).zfill(2)}",
                "total_hours": float(total_hours)
            }, status=status.HTTP_200_OK)

        # =========================
        # 🔹 BASE QUERYSET
        # =========================
        queryset = Expenses_login.objects.select_related('employee').all().order_by("-id")

        # =========================
        # 🔹 EMPLOYEE FILTER
        # =========================
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)

        # =========================
        # 🔹 BOOLEAN FILTER
        # =========================
        def apply_boolean_filter(queryset, field, value):
            if value is not None:
                if value.lower() == "true":
                    return queryset.filter(**{field: True})
                elif value.lower() == "false":
                    return queryset.filter(**{field: False})
            return queryset

        queryset = apply_boolean_filter(queryset, 'half_leave', half_leave)
        queryset = apply_boolean_filter(queryset, 'full_leave', full_leave)

        # =========================
        # 🔹 TOTAL HOURS FILTER
        # =========================
        if min_hours:
            queryset = queryset.filter(total_hours__gte=float(min_hours))

        if max_hours:
            queryset = queryset.filter(total_hours__lte=float(max_hours))

        # =========================
        # 🔥 DATE FILTERING
        # =========================
        if date:
            queryset = queryset.filter(date=date)

        elif start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])

        # ✅ ADDITION SUPPORT
        elif month or prev_month == "true" or next_month == "true":
            queryset = queryset.filter(
                date__month=selected_month,
                date__year=selected_year
            )

        elif year:
            queryset = queryset.filter(date__year=selected_year)

        # =========================
        # 🔹 RESPONSE
        # =========================
        serializer = ExpensesLoginSerializer(queryset, many=True)

        response_data = {
            "success": True,
            "message": "Records fetched successfully",
            "count": queryset.count(),
            "data": serializer.data
        }

        # ✅ ADDITION SUPPORT
        if employee_id and (month or prev_month == "true" or next_month == "true"):
            monthly_total = queryset.aggregate(
                total_hours=Coalesce(Sum('total_hours'), Decimal('0.00'))
            )
            response_data["total_month_hours"] = float(
                Decimal(str(monthly_total["total_hours"] or 0)).quantize(Decimal("0.01"))
            )
            response_data["employee_id"] = int(employee_id)
            response_data["month"] = f"{selected_year}-{str(selected_month).zfill(2)}"

        return Response(response_data, status=status.HTTP_200_OK)


    def post(self, request):
        serializer = ExpensesLoginSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Check-in successful. Have a productive day!",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # 🔹 PATCH
    def patch(self, request, pk=None):
        if not id:
            return Response({'status': 'error', 'message': 'ID is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            obj = Expenses_login.objects.get(id=pk)
        except Expenses_login.DoesNotExist:
            return Response({
                "success": False,
                "message": "Id is required"
            }, status=status.HTTP_404_NOT_FOUND)

        serializer = ExpensesLoginSerializer(obj, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Check-out successful. See you tomorrow!",
                "data": serializer.data
            })

        return Response({
            "success": False,
            "message": "Update failed",
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # 🔹 DELETE
    def delete(self, request, pk=None):
        if not id:
            return Response({'status': 'error', 'message': 'ID is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            obj = Expenses_login.objects.get(id=pk)
            obj.delete()
            return Response({
                "success": True,
                "message": "Record deleted successfully"
            })
        except Expenses_login.DoesNotExist:
            return Response({
                "success": False,
                "message": "Record not found"
            }, status=status.HTTP_404_NOT_FOUND)





# import os
# from datetime import datetime

# from django.conf import settings
# from django.utils import timezone
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status

# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# from openpyxl.utils import get_column_letter

# from .models import Expenses_login, Employee


# class AttendanceReportAPI(APIView):

#     def get(self, request):
#         try:
#             employee_id = request.GET.get("employee_id")
#             date = request.GET.get("date")
#             start_date = request.GET.get("start_date")
#             end_date = request.GET.get("end_date")
#             month = request.GET.get("month")
#             year = request.GET.get("year")
#             half_leave = request.GET.get("half_leave")
#             full_leave = request.GET.get("full_leave")
#             min_hours = request.GET.get("min_hours")
#             max_hours = request.GET.get("max_hours")

#             queryset = Expenses_login.objects.select_related("employee").all().order_by("-id")

#             employee_obj = None
#             effective_year = year

#             # Employee filter
#             if employee_id:
#                 employee_obj = Employee.objects.filter(id=employee_id).first()
#                 if not employee_obj:
#                     return Response(
#                         {
#                             "success": False,
#                             "message": "Employee not found."
#                         },
#                         status=status.HTTP_404_NOT_FOUND
#                     )
#                 queryset = queryset.filter(employee_id=employee_id)

#             # Date filters
#             if date:
#                 queryset = queryset.filter(date=date)

#             if start_date:
#                 queryset = queryset.filter(date__gte=start_date)

#             if end_date:
#                 queryset = queryset.filter(date__lte=end_date)

#             if month:
#                 try:
#                     queryset = queryset.filter(date__month=int(month))
#                 except ValueError:
#                     return Response(
#                         {
#                             "success": False,
#                             "message": "Month must be a valid number."
#                         },
#                         status=status.HTTP_400_BAD_REQUEST
#                     )

#             if year:
#                 try:
#                     queryset = queryset.filter(date__year=int(year))
#                     effective_year = str(year)
#                 except ValueError:
#                     return Response(
#                         {
#                             "success": False,
#                             "message": "Year must be a valid number."
#                         },
#                         status=status.HTTP_400_BAD_REQUEST
#                     )

#             if month and not year:
#                 effective_year = str(timezone.now().year)
#                 queryset = queryset.filter(date__year=timezone.now().year)

#             # Half leave filter
#             if half_leave is not None:
#                 value = str(half_leave).strip().lower()
#                 if value == "true":
#                     queryset = queryset.filter(half_leave=True)
#                 elif value == "false":
#                     queryset = queryset.filter(half_leave=False)

#             # Full leave filter
#             if full_leave is not None:
#                 value = str(full_leave).strip().lower()
#                 if value == "true":
#                     queryset = queryset.filter(full_leave=True)
#                 elif value == "false":
#                     queryset = queryset.filter(full_leave=False)

#             # Hours filter
#             if min_hours:
#                 try:
#                     queryset = queryset.filter(total_hours__gte=float(min_hours))
#                 except ValueError:
#                     return Response(
#                         {
#                             "success": False,
#                             "message": "min_hours must be a valid number."
#                         },
#                         status=status.HTTP_400_BAD_REQUEST
#                     )

#             if max_hours:
#                 try:
#                     queryset = queryset.filter(total_hours__lte=float(max_hours))
#                 except ValueError:
#                     return Response(
#                         {
#                             "success": False,
#                             "message": "max_hours must be a valid number."
#                         },
#                         status=status.HTTP_400_BAD_REQUEST
#                     )

#             if not queryset.exists():
#                 return Response(
#                     {
#                         "success": False,
#                         "message": "No attendance data found for the given filters."
#                     },
#                     status=status.HTTP_404_NOT_FOUND
#                 )

#             # Employee details for upper section
#             if employee_obj:
#                 employee_name = (
#                     getattr(employee_obj, "employee_name", None)
#                     or getattr(employee_obj, "full_name", None)
#                     or getattr(employee_obj, "name", None)
#                     or str(employee_obj)
#                 )
#                 employee_email = getattr(employee_obj, "email", "-")
#                 employee_contact = (
#                     getattr(employee_obj, "contact", None)
#                     or getattr(employee_obj, "phone", None)
#                     or "-"
#                 )
#             else:
#                 employee_name = "All Employees"
#                 employee_email = "-"
#                 employee_contact = "-"

#             # Excel create
#             wb = Workbook()
#             ws = wb.active
#             ws.title = "Attendance Report"

#             title_font = Font(size=14, bold=True, color="FFFFFF")
#             header_font = Font(size=11, bold=True, color="FFFFFF")
#             bold_font = Font(bold=True)

#             title_fill = PatternFill("solid", fgColor="1F4E78")
#             header_fill = PatternFill("solid", fgColor="4F81BD")

#             center_align = Alignment(horizontal="center", vertical="center")
#             left_align = Alignment(horizontal="left", vertical="center")

#             thin_border = Border(
#                 left=Side(style="thin", color="000000"),
#                 right=Side(style="thin", color="000000"),
#                 top=Side(style="thin", color="000000"),
#                 bottom=Side(style="thin", color="000000"),
#             )

#             row_num = 1

#             # Title
#             ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=12)
#             title_cell = ws.cell(row=row_num, column=1, value="Employee Attendance Report")
#             title_cell.font = title_font
#             title_cell.fill = title_fill
#             title_cell.alignment = center_align
#             title_cell.border = thin_border
#             row_num += 2

#             # Upper filter section
#             filter_rows = [
#                 ["Employee ID", employee_id or "All"],
#                 ["Employee Name", employee_name],
#                 ["Employee Email", employee_email],
#                 ["Employee Contact", employee_contact],
#                 # ["Date", date if date else "All"],
#                 # ["Start Date", start_date or "start_date"],
#                 # ["End Date", end_date or "end_date"],
#                 ["Month", month or "month"],
#                 ["Year", effective_year or "year"],

#             ]
#             for label, value in filter_rows:
#                 c1 = ws.cell(row=row_num, column=1, value=label)
#                 c2 = ws.cell(row=row_num, column=2, value=str(value))

#                 c1.font = bold_font
#                 c1.alignment = left_align
#                 c2.alignment = left_align
#                 c1.border = thin_border
#                 c2.border = thin_border

#                 row_num += 1

#             row_num += 2

#             # Table headers
#             headers = [
#                 "ID",
#                 "Employee ID",
#                 "Employee Name",
#                 "Employee Email",
#                 "Employee Contact",
#                 "Date",
#                 "Start Time",
#                 "End Time",
#                 "Status",
#                 "Full Leave",
#                 "Half Leave",
#                 "Total Hours",
#             ]

#             for col_num, header in enumerate(headers, start=1):
#                 cell = ws.cell(row=row_num, column=col_num, value=header)
#                 cell.font = header_font
#                 cell.fill = header_fill
#                 cell.alignment = center_align
#                 cell.border = thin_border

#             row_num += 1

#             # Data rows
#             for obj in queryset:
#                 row_employee = getattr(obj, "employee", None)

#                 row_employee_name = "-"
#                 row_employee_email = "-"
#                 row_employee_contact = "-"

#                 if row_employee:
#                     row_employee_name = (
#                         getattr(row_employee, "employee_name", None)
#                         or getattr(row_employee, "full_name", None)
#                         or getattr(row_employee, "name", None)
#                         or str(row_employee)
#                     )
#                     row_employee_email = getattr(row_employee, "email", "-")
#                     row_employee_contact = (
#                         getattr(row_employee, "contact", None)
#                         or getattr(row_employee, "phone", None)
#                         or "-"
#                     )

#                 row_data = [
#                     obj.id,
#                     obj.employee_id,
#                     row_employee_name,
#                     row_employee_email,
#                     row_employee_contact,
#                     obj.date.strftime("%Y-%m-%d") if obj.date else "-",
#                     obj.start_time.strftime("%H:%M:%S") if obj.start_time else "-",
#                     obj.end_time.strftime("%H:%M:%S") if obj.end_time else "-",
#                     "Present" if obj.status else "Absent",
#                     "Yes" if obj.full_leave else "No",
#                     "Yes" if obj.half_leave else "No",
#                     float(obj.total_hours or 0),
#                 ]

#                 for col_num, value in enumerate(row_data, start=1):
#                     cell = ws.cell(row=row_num, column=col_num, value=value)
#                     cell.border = thin_border
#                     cell.alignment = left_align if col_num in [3, 4, 5] else center_align

#                 row_num += 1

#             # Column widths
#             column_widths = {
#                 1: 10,
#                 2: 15,
#                 3: 25,
#                 4: 30,
#                 5: 18,
#                 6: 15,
#                 7: 15,
#                 8: 15,
#                 9: 15,
#                 10: 15,
#                 11: 15,
#                 12: 15,
#             }

#             for col_num, width in column_widths.items():
#                 ws.column_dimensions[get_column_letter(col_num)].width = width

#             # Save report
#             report_dir = os.path.join(settings.MEDIA_ROOT, "reports")
#             os.makedirs(report_dir, exist_ok=True)

#             timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#             filename = f"attendance_report_{timestamp}.xlsx"
#             filepath = os.path.join(report_dir, filename)

#             wb.save(filepath)

#             file_url = request.build_absolute_uri(f"{settings.MEDIA_URL}reports/{filename}")

#             return Response(
#                 {
#                     "success": True,
#                     "message": "Attendance report generated successfully.",
#                     "file_url": file_url
#                 },
#                 status=status.HTTP_200_OK
#             )

#         except Exception as e:
#             return Response(
#                 {
#                     "success": False,
#                     "message": "Something went wrong while generating attendance report.",
#                     "error": str(e)
#                 },
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )



from reportlab.lib.pagesizes import A4, landscape
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from django.utils import timezone

def generate_picklist_pdf(po):

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    elements = []
    styles = getSampleStyleSheet()

    # ================= HEADER =================
    elements.append(Paragraph("<b>Company Name:</b> HogoAutoFilms", styles["Normal"]))
    elements.append(Paragraph("<b>Warehouse / Location:</b> Main Warehouse", styles["Normal"]))
    elements.append(Paragraph(f"<b>Pick List No:</b> PL-{po.id}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Date:</b> {timezone.now().date()}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Customer / Distributor:</b> {po.distributor_id}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Sales Order / Invoice No:</b> SO-{po.id}", styles["Normal"]))

    elements.append(Spacer(1, 20))

    # ================= TABLE HEADER =================
    data = [[
        "Sr No",
        "Item Code",
        "Item Description",
        "Batch/Lot",
        "Warehouse",
        "Location",
        "Qty Ordered",
        "UOM",
        "Qty Picked",
        "Remarks",
        "Short Pick"
    ]]

    product_items = po.product_items or []

    for index, item in enumerate(product_items, start=1):

        product_id = item.get("product_id")
        product_name = item.get("product_name")
        quantity = item.get("quantity")

        # Get SKU
        try:
            product = Product.objects.get(id=product_id)
            sku = product.sku
        except Product.DoesNotExist:
            sku = ""

        batch = ""
        warehouse_name = ""
        location_name = ""

        # Always fetch from InventorySerial by product_id (no status filter)
        inv_serial = InventorySerial.objects.filter(
            product_id=product_id
        ).select_related("warehouse", "location").first()

        if inv_serial:
            batch = inv_serial.batch_id or ""
            warehouse_name = inv_serial.warehouse.name if inv_serial.warehouse else ""
            location_name = inv_serial.location.name if inv_serial.location else ""

            # Fallback: if location/warehouse still null (old data before migration 0021),
            # look them up from the matching Shipment_product record
            if not warehouse_name or not location_name:
                sp = Shipment_product.objects.filter(
                    product_id=product_id,
                    batch_data=inv_serial.batch_id
                ).select_related("location_id", "warehouse_id").first()
                if sp:
                    if not warehouse_name and sp.warehouse_id:
                        warehouse_name = sp.warehouse_id.name
                    if not location_name and sp.location_id:
                        location_name = sp.location_id.name


        qty_picked_raw = item.get("qty_picked", "")
        remarks = item.get("remarks", "")

        qty_picked = ""
        short_pick = ""

        if str(qty_picked_raw).isdigit() and int(qty_picked_raw) > 0:
            qty_picked = str(qty_picked_raw)
            if str(quantity).isdigit():
                s_val = int(quantity) - int(qty_picked_raw)
                if s_val > 0:
                    short_pick = str(s_val)

        data.append([
            str(index),
            str(sku),
            str(product_name),
            str(batch),
            str(warehouse_name),
            str(location_name),
            str(quantity),
            "PCS",
            str(qty_picked),
            str(remarks),
            str(short_pick)
        ])

    if len(data) == 1:
        data.append(["-", "-", "No Products", "-", "-", "-", "-", "-", "-", "-", "-"])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return buffer


from django.db import transaction


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.http import HttpResponse




def generate_packing_pdf(po):
    from io import BytesIO
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from django.utils import timezone
    from .models import InventorySerial, Product, Shipment_product

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20
    )
    elements = []
    styles = getSampleStyleSheet()

    header_warehouse = "Main Warehouse"
    header_location = ""
    product_items = po.product_items or []

    if product_items:
        first_pid = product_items[0].get("product_id")
        inv_s = InventorySerial.objects.filter(product_id=first_pid).select_related("warehouse", "location").first()
        if inv_s:
            if inv_s.warehouse: header_warehouse = inv_s.warehouse.name
            if inv_s.location: header_location = inv_s.location.name
            if not header_warehouse or not header_location:
                sp = Shipment_product.objects.filter(product_id=first_pid, batch_data=inv_s.batch_id).select_related("location_id", "warehouse_id").first()
                if sp:
                    if not header_warehouse and sp.warehouse_id: header_warehouse = sp.warehouse_id.name
                    if not header_location and sp.location_id: header_location = sp.location_id.name

    warehouse_location = f"{header_warehouse}" + (f" / {header_location}" if header_location else "")
    distributor = po.distributor_id
    dispatch_address = getattr(distributor, "address", "") or ""

    elements.append(Paragraph("<b>Company Name:</b> HogoAutoFilms", styles["Normal"]))
    elements.append(Paragraph(f"<b>Warehouse Location:</b> {warehouse_location}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Packing List No.:</b> PL-{po.id}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Date:</b> {timezone.now().date()}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Customer / Distributor Name:</b> {distributor}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Dispatch Address:</b> {dispatch_address}", styles["Normal"]))
    elements.append(Spacer(1, 16))

    data = [[
        "Sr No", "Product Code", "Product Description", "HSN Code",
        "Batch / Roll No.", "Qty", "Unit", "No. of Cartons /\nBoxes",
        "Gross Weight\n(Kg)", "Net Weight\n(Kg)", "Remarks"
    ]]

    total_qty, total_cartons, total_gross, total_net = 0, 0, 0.0, 0.0

    for index, item in enumerate(product_items, start=1):
        product_id = item.get("product_id")
        product_name = item.get("product_name", "")
        quantity = item.get("quantity", 0)

        try:
            product_obj = Product.objects.get(id=product_id)
            sku = product_obj.sku
            hsn_code = str(product_obj.hsn_code or "")
        except Product.DoesNotExist:
            sku = ""
            hsn_code = ""

        batch = ""
        inv_serial = InventorySerial.objects.filter(product_id=product_id).first()
        if inv_serial:
            batch = inv_serial.batch_id or ""
            if not batch:
                sp = Shipment_product.objects.filter(product_id=product_id).first()
                if sp: batch = sp.batch_data or ""

        cartons_raw = item.get("cartons", "")
        gross_raw = item.get("gross_weight", "")
        net_raw = item.get("net_weight", "")
        remarks = str(item.get("remarks", "") or "")

        cartons_str = str(cartons_raw) if cartons_raw not in (None, "", 0, "0") else ""
        gross_str = str(gross_raw) if gross_raw not in (None, "", 0, "0") else ""
        net_str = str(net_raw) if net_raw not in (None, "", 0, "0") else ""

        total_qty += int(quantity) if str(quantity).isdigit() else 0
        try: total_cartons += int(cartons_raw)
        except: pass
        try: total_gross += float(gross_raw)
        except: pass
        try: total_net += float(net_raw)
        except: pass

        data.append([
            str(index), str(sku), str(product_name), hsn_code,
            str(batch), str(quantity), "PCS", cartons_str,
            gross_str, net_str, remarks
        ])

    if len(data) == 1:
        data.append(["-"] * 11)

    table = Table(data, colWidths=[35, 70, 130, 60, 70, 35, 35, 65, 65, 65, 65], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 16))

    elements.append(Paragraph(f"<b>Total Quantity:</b> {total_qty}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Total Cartons / Boxes:</b> {total_cartons if total_cartons else ''}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Total Gross Weight (Kg):</b> {total_gross:.2f}" if total_gross else "<b>Total Gross Weight (Kg):</b> ", styles["Normal"]))
    elements.append(Paragraph(f"<b>Total Net Weight (Kg):</b> {total_net:.2f}" if total_net else "<b>Total Net Weight (Kg):</b> ", styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)
    return buffer


from django.urls import reverse

# class PurchaseOrderAPIView(APIView):

#     # ================= SHORT PICK PO CREATION =================
#     def create_short_pick_po(self, po):
#         short_items = []
#         for item in po.product_items or []:
#             qty = int(item.get("quantity", 0) or 0)
#             picked = int(item.get("qty_picked", 0) or 0)
#             if picked > qty:
#                 picked = qty
#             short_qty = qty - picked
#             if short_qty > 0:
#                 short_item = item.copy()
#                 short_item["quantity"] = short_qty
#                 short_item["qty_picked"] = 0
#                 short_items.append(short_item)

#         if not short_items:
#             return None

#         base_po_number = po.po_number
#         suffix = "-SP"
#         candidate_po_number = f"{base_po_number}{suffix}"
#         counter = 1
#         while PurchaseOrder.objects.filter(po_number=candidate_po_number).exists():
#             counter += 1
#             candidate_po_number = f"{base_po_number}{suffix}{counter}"

#         new_po = PurchaseOrder.objects.create(
#             po_number=candidate_po_number,
#             distributor_id=po.distributor_id,
#             product_items=short_items,
#             total_items=len(short_items),
#             total_quantity=sum(int(i.get("quantity", 0) or 0) for i in short_items),
#             status="SUBMITTED",
#             remarks=f"Short picked from PO {po.po_number}",
#             created_by=po.created_by,
#         )

#         import re
#         from datetime import datetime

#         # 🔄 Transfer remaining RESERVED serials to the new PO
#         for item in short_items:
#             prod_id = item.get("product_id")
#             short_q = int(item.get("quantity", 0))

#             # Since these are short items that weren't available in IN_STOCK originally,
#             # we need to generate new reserved inventory serials for them in the background.
#             try:
#                 product = Product.objects.get(id=prod_id)
#                 sku = product.sku if hasattr(product, 'sku') and product.sku else "SKU"
#                 batch_data = "PO_AUTO"
#                 timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

#                 last_serial = (
#                     InventorySerial.objects
#                     .filter(product_id=product)
#                     .order_by("-id")
#                     .first()
#                 )

#                 start = 1
#                 if last_serial and last_serial.serial_number:
#                     match = re.search(r"-(\d+)$", str(last_serial.serial_number))
#                     if match:
#                         start = int(match.group(1)) + 1

#                 for i in range(start, start + short_q):
#                     serial_number = f"{sku}-{batch_data}-{timestamp}-{i}"
#                     new_serial = InventorySerial.objects.create(
#                         product_id=product,
#                         batch_id=batch_data,
#                         serial_number=serial_number,
#                         status="RESERVED"
#                     )
#                     DistributerOrders.objects.create(
#                         distributor_id=new_po.distributor_id,
#                         inventory_serial_id=new_serial,
#                         purchase_order_id=new_po
#                     )
#             except Product.DoesNotExist:
#                 continue

#         return new_po


#     # ================= GET =================
#     def get(self, request, pk=None):

#         if pk:
#             po = get_object_or_404(PurchaseOrder, pk=pk)

#             return Response(
#                 {"success": True, "data": PurchaseOrderSerializer(po).data},
#                 status=status.HTTP_200_OK
#             )

#         orders = PurchaseOrder.objects.all().order_by("-id")

#         status_filter = request.query_params.get("status")
#         if status_filter:
#             orders = orders.filter(status=status_filter)

#         po_date = request.query_params.get("po_date")
#         if po_date:
#             orders = orders.filter(po_date=po_date)

#         distributor_id = request.query_params.get("distributor_id")
#         if distributor_id:
#             orders = orders.filter(distributor_id=distributor_id)

#         return Response(
#             {"success": True, "data": PurchaseOrderSerializer(orders, many=True).data},
#             status=status.HTTP_200_OK
#         )


#     # ================= POST =================
#     def post(self, request):

#         serializer = PurchaseOrderSerializer(data=request.data)
#         serializer.is_valid(raise_exception=True)

#         po = serializer.save()

#         response_data = {
#             "success": True,
#             "message": "Purchase order created successfully",
#             "data": PurchaseOrderSerializer(po).data
#         }

#         return Response(response_data, status=status.HTTP_201_CREATED)


#     # ================= PATCH =================
#     @transaction.atomic
#     def patch(self, request, pk=None):

#         po = get_object_or_404(PurchaseOrder, pk=pk)

#         old_status = po.status

#         serializer = PurchaseOrderSerializer(po, data=request.data, partial=True)
#         serializer.is_valid(raise_exception=True)

#         po = serializer.save()

#         new_status = po.status

#         new_po = None

#         # ===== APPROVED =====
#         if new_status == "APPROVED":
#             if old_status != "APPROVED":
#                 reserved_serials = InventorySerial.objects.filter(
#                     orders__purchase_order_id=po,
#                     status="RESERVED"
#                 )
#                 for serial in reserved_serials:
#                     # ✅ Track History when RESERVED/APPROVED
#                     InventorySerialHistory.objects.get_or_create(
#                         product=serial.product_id,
#                         serial_number=serial.serial_number,
#                         status="RESERVED",
#                         purchase_order=po,
#                         defaults={
#                             "batch_id": serial.batch_id,
#                             "location": serial.location,
#                             "warehouse": serial.warehouse,
#                             "distributor": po.distributor_id,
#                             "snapshot_date": timezone.now().date(),
#                             "month": timezone.now().month,
#                             "year": timezone.now().year
#                         }
#                     )

#         # ===== PICKED =====
#         if new_status == "PICKED":
#             if old_status != "PICKED":
#                 # Create Pic_product marker
#                 Pic_product.objects.get_or_create(purchase_order_id=po)

#                 # For each product, mark the Picked quantity of serials as PICKED
#                 for item in po.product_items or []:
#                     prod_id = item.get("product_id")
#                     picked_q = int(item.get("qty_picked", 0) or 0)

#                     if picked_q > 0:
#                         orders_to_pick = DistributerOrders.objects.filter(
#                             purchase_order_id=po,
#                             inventory_serial_id__product_id_id=prod_id,
#                             inventory_serial_id__status="RESERVED"
#                         )[:picked_q]

#                         for order in orders_to_pick:
#                             serial = order.inventory_serial_id
#                             serial.status = "PICKED"
#                             serial.save()

#                             # ✅ Track History only when PICKED
#                             InventorySerialHistory.objects.create(
#                                 product=serial.product_id,
#                                 serial_number=serial.serial_number,
#                                 status="PICKED",
#                                 batch_id=serial.batch_id,
#                                 location=serial.location,
#                                 warehouse=serial.warehouse,
#                                 purchase_order=po,
#                                 distributor=po.distributor_id,
#                                 snapshot_date=timezone.now().date(),
#                                 month=timezone.now().month,
#                                 year=timezone.now().year
#                             )

#                             # ✅ AUTOMATED WARRANTY ACTIVATION
#                             try:
#                                 # Parse warranty months from product.warranty string (e.g. "12 Months")
#                                 warranty_str = str(serial.product_id.warranty or "12")
#                                 import re
#                                 match = re.search(r'(\d+)', warranty_str)
#                                 months = int(match.group(1)) if match else 12

#                                 start_date = timezone.now().date()
#                                 end_date = start_date + timedelta(days=months * 30) # Approximate

#                                 Warranty.objects.get_or_create(
#                                     serial_id=serial,
#                                     product_id=serial.product_id,
#                                     defaults={
#                                         "distributor_id": po.distributor_id,
#                                         "warranty_period": months,
#                                         "warranty_start_date": start_date,
#                                         "warranty_end_date": end_date,
#                                         "warranty_status": "ACTIVATED",
#                                         "product_status": "ACTIVATED"
#                                     }
#                                 )
#                             except Exception as e:
#                                 print(f"Warranty activation error for serial {serial.serial_number}: {e}")

#                 # CREATE SHORT PICK PO HERE (Remaining RESERVED will be transferred)
#                 new_po = None

#             if old_status != "PICKED":

#                 pdf_buffer = generate_picklist_pdf(po)

#                 response = HttpResponse(
#                     pdf_buffer,
#                     content_type="application/pdf",
#                     headers={
#                         "Content-Disposition": f'attachment; filename="PickList_PO_{po.id}.pdf"'
#                     }
#                 )
#                 if new_po:
#                     response["X-Short-Pick-PO-ID"] = str(new_po.id)
#                     response["X-Short-Pick-PO-Number"] = str(new_po.po_number)

#                 # ✅ Add Warranty Card Link for the frontend/admin
#                 warranty_link = request.build_absolute_uri(
#                     reverse("warranty-card-bulk-pdf") + f"?po_id={po.id}"
#                 )
#                 response["X-Warranty-Cards-URL"] = warranty_link

#                 return response

#         # ===== PACKED =====
#         if new_status == "PACKED":

#             Pack_product.objects.get_or_create(purchase_order_id=po)

#             target_serials = InventorySerial.objects.filter(
#                 orders__purchase_order_id=po
#             ).exclude(status="PACKED")

#             for serial in target_serials:
#                 serial.status = "PACKED"
#                 serial.save()

#             incoming_items = request.data.get("product_items", [])

#             if incoming_items:

#                 current_items = po.product_items or []

#                 item_map = {str(it.get("product_id")): it for it in incoming_items}

#                 changed = False

#                 for idx, ci in enumerate(current_items):

#                     pid_key = str(ci.get("product_id"))

#                     if pid_key in item_map:

#                         incoming = item_map[pid_key]

#                         for field in ("cartons", "gross_weight", "net_weight"):

#                             if field in incoming:
#                                 current_items[idx][field] = incoming[field]
#                                 changed = True

#                 if changed:
#                     po.product_items = current_items
#                     po.save(update_fields=["product_items"])

#             if old_status != "PACKED":

#                 pdf_buffer = generate_packing_pdf(po)

#                 return HttpResponse(
#                     pdf_buffer,
#                     content_type="application/pdf",
#                     headers={
#                         "Content-Disposition": f'attachment; filename="PackingList_PO_{po.id}.pdf"'
#                     }
#                 )

#         # ===== CANCELLED / REJECTED =====
#         if new_status in ["CANCELLED", "REJECTED"]:

#             distributor_orders = DistributerOrders.objects.filter(
#                 purchase_order_id=po
#             )

#             for order in distributor_orders:

#                 serial = order.inventory_serial_id
#                 serial.status = "IN_STOCK"
#                 serial.save()

#             distributor_orders.delete()

#         response_data = {
#             "success": True,
#             "message": "Purchase order updated successfully",
#             "data": PurchaseOrderSerializer(po).data
#         }

#         if new_po:
#             response_data["short_pick_po"] = PurchaseOrderSerializer(new_po).data
#             response_data["message"] += f". Short pick PO created with ID {new_po.id}"

#         return Response(response_data, status=status.HTTP_200_OK)


#     # ================= DELETE =================
#     @transaction.atomic
#     def delete(self, request, pk=None):

#         po = get_object_or_404(PurchaseOrder, pk=pk)

#         distributor_orders = DistributerOrders.objects.filter(
#             purchase_order_id=po
#         )

#         for order in distributor_orders:

#             serial = order.inventory_serial_id
#             serial.status = "IN_STOCK"
#             serial.save()

#         distributor_orders.delete()
#         po.delete()

#         return Response(
#             {
#                 "success": True,
#                 "message": "Purchase order deleted successfully"
#             },
#             status=status.HTTP_200_OK
#         )





class PurchaseOrderAPIView(APIView):

    # ================= SHORT PICK PO CREATION =================
    def create_short_pick_po(self, po):
        short_items = []
        for item in po.product_items or []:
            qty = int(item.get("quantity", 0) or 0)
            picked = int(item.get("qty_picked", 0) or 0)
            if picked > qty:
                picked = qty
            short_qty = qty - picked
            if short_qty > 0:
                short_item = item.copy()
                short_item["quantity"] = short_qty
                short_item["qty_picked"] = 0
                short_items.append(short_item)

        if not short_items:
            return None

        base_po_number = po.po_number
        suffix = "-SP"
        candidate_po_number = f"{base_po_number}{suffix}"
        counter = 1
        while PurchaseOrder.objects.filter(po_number=candidate_po_number).exists():
            counter += 1
            candidate_po_number = f"{base_po_number}{suffix}{counter}"

        new_po = PurchaseOrder.objects.create(
            po_number=candidate_po_number,
            distributor_id=po.distributor_id,
            product_items=short_items,
            total_items=len(short_items),
            total_quantity=sum(int(i.get("quantity", 0) or 0) for i in short_items),
            status="SUBMITTED",
            remarks=f"Short picked from PO {po.po_number}",
            created_by=po.created_by,
        )

        import re
        from datetime import datetime

        # 🔄 Transfer remaining RESERVED serials to the new PO
        for item in short_items:
            prod_id = item.get("product_id")
            short_q = int(item.get("quantity", 0))

            # Since these are short items that weren't available in IN_STOCK originally,
            # we need to generate new reserved inventory serials for them in the background.
            try:
                product = Product.objects.get(id=prod_id)
                sku = product.sku if hasattr(product, 'sku') and product.sku else "SKU"
                batch_data = "PO_AUTO"
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

                last_serial = (
                    InventorySerial.objects
                    .filter(product_id=product)
                    .order_by("-id")
                    .first()
                )

                start = 1
                if last_serial and last_serial.serial_number:
                    match = re.search(r"-(\d+)$", str(last_serial.serial_number))
                    if match:
                        start = int(match.group(1)) + 1

                for i in range(start, start + short_q):
                    serial_number = f"{sku}-{batch_data}-{timestamp}-{i}"
                    new_serial = InventorySerial.objects.create(
                        product_id=product,
                        batch_id=batch_data,
                        serial_number=serial_number,
                        status="RESERVED"
                    )
                    DistributerOrders.objects.create(
                        distributor_id=new_po.distributor_id,
                        inventory_serial_id=new_serial,
                        purchase_order_id=new_po
                    )
            except Product.DoesNotExist:
                continue

        return new_po


    # ================= GET =================
    def get(self, request, pk=None):

        if pk:
            po = get_object_or_404(PurchaseOrder, pk=pk)

            return Response(
                {"success": True, "data": PurchaseOrderSerializer(po).data},
                status=status.HTTP_200_OK
            )

        orders = PurchaseOrder.objects.all().order_by("-id")

        status_filter = request.query_params.get("status")
        if status_filter:
            orders = orders.filter(status=status_filter)

        po_date = request.query_params.get("po_date")
        if po_date:
            orders = orders.filter(po_date__date=po_date)

        distributor_id = request.query_params.get("distributor_id")
        if distributor_id:
            orders = orders.filter(distributor_id_id=distributor_id)
            # agar model me field name distributor_id hi ForeignKey hai,
            # Django DB column distributor_id_id ban sakta hai

        return Response(
            {
                "success": True,
                "count": orders.count(),
                "data": PurchaseOrderSerializer(orders, many=True).data
            },
            status=status.HTTP_200_OK
        )

    # ================= POST =================
    def post(self, request):

        serializer = PurchaseOrderSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        po = serializer.save()

        response_data = {
            "success": True,
            "message": "Purchase order created successfully",
            "data": PurchaseOrderSerializer(po).data
        }

        return Response(response_data, status=status.HTTP_201_CREATED)


    # ================= PATCH =================
    @transaction.atomic
    def patch(self, request, pk=None):

        po = get_object_or_404(PurchaseOrder, pk=pk)

        old_status = po.status

        serializer = PurchaseOrderSerializer(po, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)

        po = serializer.save()

        new_status = po.status

        new_po = None

        # ===== APPROVED =====
        if new_status == "APPROVED":
            if old_status != "APPROVED":
                reserved_serials = InventorySerial.objects.filter(
                    orders__purchase_order_id=po,
                    status="RESERVED"
                )

                for serial in reserved_serials:
                    InventorySerialHistory.objects.get_or_create(
                        product=serial.product_id,
                        serial_number=serial.serial_number,
                        status="RESERVED",
                        purchase_order_id=po.id,   # ✅ FIXED
                        defaults={
                            "batch_id": serial.batch_id,
                            "location": serial.location,
                            "warehouse": serial.warehouse,
                            "snapshot_date": timezone.now().date(),
                            "month": timezone.now().month,
                            "year": timezone.now().year
                        }
                    )

        # ===== PICKED =====
        if new_status == "PICKED":

            if old_status != "PICKED":
                Pic_product.objects.get_or_create(purchase_order_id=po)

                for item in po.product_items or []:
                    prod_id = item.get("product_id")
                    picked_q = int(item.get("qty_picked", 0) or 0)

                    if picked_q > 0:
                        orders_to_pick = DistributerOrders.objects.filter(
                            purchase_order_id=po,
                            inventory_serial_id__product_id_id=prod_id,
                            inventory_serial_id__status="RESERVED"
                        )[:picked_q]

                        for order in orders_to_pick:
                            serial = order.inventory_serial_id
                            serial.status = "PICKED"
                            serial.save(update_fields=["status"])

                            InventorySerialHistory.objects.create(
                                product=serial.product_id,
                                serial_number=serial.serial_number,
                                status="PICKED",
                                batch_id=serial.batch_id,
                                location=serial.location,
                                warehouse=serial.warehouse,
                                purchase_order_id=po,
                                snapshot_date=timezone.now().date(),
                                month=timezone.now().month,
                                year=timezone.now().year
                            )

                            try:
                                warranty_str = str(serial.product_id.warranty or "12")

                                import re
                                match = re.search(r'(\d+)', warranty_str)
                                months = int(match.group(1)) if match else 12

                                start_date = timezone.now().date()
                                end_date = start_date + timedelta(days=months * 30)

                                Warranty.objects.get_or_create(
                                    serial_id=serial,
                                    product_id=serial.product_id,
                                    defaults={
                                        "distributor_id": po.distributor_id,
                                        "warranty_period": months,
                                        "warranty_start_date": start_date,
                                        "warranty_end_date": end_date,
                                        "warranty_status": "ACTIVATED",
                                        "product_status": "ACTIVATED"
                                    }
                                )

                            except Exception as e:
                                print(f"Warranty activation error for serial {serial.serial_number}: {e}")

                new_po = self.create_short_pick_po(po)

            # ✅ Refresh to ensure we have latest JSON product_items for PDF
            po.refresh_from_db()

            # ✅ Always generate PDF when final status is PICKED
            try:
                pdf_buffer = generate_picklist_pdf(po)

                response = HttpResponse(
                    pdf_buffer,
                    content_type="application/pdf",
                    headers={
                        "Content-Disposition": f'attachment; filename="PickList_PO_{po.id}.pdf"'
                    }
                )

                if new_po:
                    response["X-Short-Pick-PO-ID"] = str(new_po.id)
                    response["X-Short-Pick-PO-Number"] = str(new_po.po_number)

                warranty_link = request.build_absolute_uri(
                    reverse("warranty-card-bulk-pdf") + f"?po_id={po.id}"
                )
                response["X-Warranty-Cards-URL"] = warranty_link

                return response

            except Exception as e:
                print(f"Picklist PDF generation failed: {e}")

        # ===== PACKED =====
        if new_status == "PACKED":

            Pack_product.objects.get_or_create(purchase_order_id=po)

            # ✅ Update Serial statuses to PACKED
            target_serials = InventorySerial.objects.filter(
                orders__purchase_order_id=po
            ).exclude(status="PACKED")

            for serial in target_serials:
                serial.status = "PACKED"
                serial.save(update_fields=["status"])

            incoming_items = request.data.get("product_items", [])

            if incoming_items:

                current_items = po.product_items or []

                item_map = {str(it.get("product_id")): it for it in incoming_items}

                changed = False

                for idx, ci in enumerate(current_items):

                    pid_key = str(ci.get("product_id"))

                    if pid_key in item_map:

                        incoming = item_map[pid_key]

                        for field in ("cartons", "gross_weight", "net_weight"):

                            if field in incoming:
                                current_items[idx][field] = incoming[field]
                                changed = True

                if changed:
                    po.product_items = current_items
                    po.save(update_fields=["product_items"])

            if old_status != "PACKED":
                po.refresh_from_db()
                pdf_buffer = generate_packing_pdf(po)

                return HttpResponse(
                    pdf_buffer,
                    content_type="application/pdf",
                    headers={
                        "Content-Disposition": f'attachment; filename="PackingList_PO_{po.id}.pdf"'
                    }
                )

        # ===== CANCELLED / REJECTED =====
        if new_status in ["CANCELLED", "REJECTED"]:
            pass

        response_data = {
            "success": True,
            "message": "Purchase order updated successfully",
            "data": PurchaseOrderSerializer(po).data
        }

        if new_po:
            response_data["short_pick_po"] = PurchaseOrderSerializer(new_po).data
            response_data["message"] += f". Short pick PO created with ID {new_po.id}"

        return Response(response_data, status=status.HTTP_200_OK)


    # ================= DELETE =================
    @transaction.atomic
    def delete(self, request, pk=None):

        po = get_object_or_404(PurchaseOrder, pk=pk)

        distributor_orders = DistributerOrders.objects.filter(
            purchase_order_id=po
        )

        for order in distributor_orders:

            serial = order.inventory_serial_id
            serial.status = "IN_STOCK"
            serial.save()

        distributor_orders.delete()
        po.delete()

        return Response(
            {
                "success": True,
                "message": "Purchase order deleted successfully"
            },
            status=status.HTTP_200_OK
        )





class PurchaseOrderPDFView(APIView):
    """Re-download the Pick List PDF for any approved/processed order."""
    def get(self, request, pk=None):
        po = get_object_or_404(PurchaseOrder, pk=pk)

        # Ensure it has been at least conditionally picked/reserved
        if po.status not in ["APPROVED", "PARTIALLY_APPROVED", "PICKED", "PACKED", "DELIVERED"]:
            return Response(
                {"success": False, "message": "PDF can only be generated for approved or processed orders."},
                status=400
            )

        pdf_buffer = generate_picklist_pdf(po)

        return HttpResponse(
            pdf_buffer,
            content_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="PickList_PO_{po.id}.pdf"'
            }
        )

class PurchaseOrderPackingPDFView(APIView):
    """Re-download the Packing List PDF for a PACKED or DELIVERED order."""
    def get(self, request, pk=None):
        po = get_object_or_404(PurchaseOrder, pk=pk)

        if po.status not in ["PICKED", "PACKED", "DELIVERED"]:
            return Response(
                {
                    "success": False,
                    "message": "Packing PDF is only available for PICKED, PACKED or DELIVERED orders."
                },
                status=400
            )

        pdf_buffer = generate_packing_pdf(po)

        return HttpResponse(
            pdf_buffer,
            content_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="PackingList_PO_{po.id}.pdf"'
            }
        )


from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone

from rest_framework.views import APIView
from rest_framework.response import Response

from io import BytesIO
import os

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from .models import PurchaseOrder, Product, CompanyProfile


class PurchaseOrderInvoicePDFView(APIView):
    """Generate professional Tax Invoice PDF for a Purchase Order."""

    def get(self, request, pk=None):
        po = get_object_or_404(PurchaseOrder, pk=pk)

        if po.status not in ["APPROVED", "PARTIALLY_APPROVED", "PICKED", "PACKED", "DELIVERED"]:
            return Response(
                {
                    "success": False,
                    "message": "Invoice PDF can only be generated for approved or processed orders."
                },
                status=400
            )

        pdf_buffer = self.generate_invoice_pdf(po)

        filename = f"Invoice_{po.po_number}.pdf"
        return HttpResponse(
            pdf_buffer,
            content_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )

    def num_to_words(self, n):
        if n == 0:
            return "Zero"

        ones = [
            "", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
            "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
            "Seventeen", "Eighteen", "Nineteen"
        ]
        tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

        def _helper(num):
            if num < 20:
                return ones[num]
            if num < 100:
                return tens[num // 10] + (" " + ones[num % 10] if num % 10 != 0 else "")
            if num < 1000:
                return ones[num // 100] + " Hundred" + (" " + _helper(num % 100) if num % 100 != 0 else "")
            if num < 100000:
                return _helper(num // 1000) + " Thousand" + (" " + _helper(num % 1000) if num % 1000 != 0 else "")
            if num < 10000000:
                return _helper(num // 100000) + " Lakh" + (" " + _helper(num % 100000) if num % 100000 != 0 else "")
            return _helper(num // 10000000) + " Crore" + (" " + _helper(num % 10000000) if num % 10000000 != 0 else "")

        integer_part = int(n)
        decimal_part = int(round((n - integer_part) * 100))

        res = _helper(integer_part) + " Only"
        if decimal_part > 0:
            res = _helper(integer_part) + " and " + _helper(decimal_part) + " Paise Only"
        return res

    def indian_format(self, n):
        try:
            n = float(n or 0)
        except Exception:
            n = 0.0

        s = f"{n:,.2f}"
        parts = s.split('.')
        integer_part = parts[0].replace(',', '')
        decimal_part = parts[1] if len(parts) > 1 else '00'

        if len(integer_part) <= 3:
            return integer_part + '.' + decimal_part

        result = integer_part[-3:]
        integer_part = integer_part[:-3]

        while integer_part:
            result = integer_part[-2:] + ',' + result
            integer_part = integer_part[:-2]

        return result + '.' + decimal_part

    def _get_company_value(self, obj, *field_names, default=""):
        if not obj:
            return default

        for field in field_names:
            if hasattr(obj, field):
                value = getattr(obj, field)
                if value not in [None, ""]:
                    return value
        return default

    def _get_company_file_path(self, obj, *field_names):
        if not obj:
            return None

        for field in field_names:
            if hasattr(obj, field):
                file_obj = getattr(obj, field)
                if file_obj:
                    try:
                        if hasattr(file_obj, "path") and os.path.exists(file_obj.path):
                            return file_obj.path
                    except Exception:
                        pass
        return None

    def _build_safe_image(self, image_path, width, height):
        try:
            if not image_path or not os.path.exists(image_path):
                return ""

            # ✅ Handle SVG separately
            if image_path.lower().endswith(".svg"):
                try:
                    import cairosvg
                    from io import BytesIO

                    png_io = BytesIO()
                    cairosvg.svg2png(url=image_path, write_to=png_io)
                    png_io.seek(0)

                    return Image(png_io, width=width, height=height)
                except Exception as e:
                    print("SVG conversion failed:", e)
                    return ""

            # ✅ Normal images (PNG/JPG)
            return Image(image_path, width=width, height=height)

        except Exception as e:
            print("Image load failed:", e)
            return ""


    def generate_invoice_pdf(self, po):
        PAGE_W = A4[0]
        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=25,
            leftMargin=25,
            topMargin=25,
            bottomMargin=25
        )

        usable = PAGE_W - 50
        LEFT_COL = usable * 0.52
        RIGHT_COL = usable * 0.48

        elements = []
        styles = getSampleStyleSheet()

        s7 = ParagraphStyle('s7', parent=styles['Normal'], fontSize=7, leading=9)
        s7b = ParagraphStyle('s7b', parent=styles['Normal'], fontSize=7, leading=9, fontName='Helvetica-Bold')
        s8 = ParagraphStyle('s8', parent=styles['Normal'], fontSize=8, leading=10)
        s8b = ParagraphStyle('s8b', parent=styles['Normal'], fontSize=8, leading=10, fontName='Helvetica-Bold')
        s8bi = ParagraphStyle('s8bi', parent=styles['Normal'], fontSize=8, leading=10, fontName='Helvetica-BoldOblique')
        s8r = ParagraphStyle('s8r', parent=styles['Normal'], fontSize=8, leading=10, alignment=TA_RIGHT)
        s8br = ParagraphStyle('s8br', parent=styles['Normal'], fontSize=8, leading=10, fontName='Helvetica-Bold', alignment=TA_RIGHT)
        s8c = ParagraphStyle('s8c', parent=styles['Normal'], fontSize=8, leading=10, alignment=TA_CENTER)
        s8bc = ParagraphStyle('s8bc', parent=styles['Normal'], fontSize=8, leading=10, fontName='Helvetica-Bold', alignment=TA_CENTER)
        s7bc = ParagraphStyle('s7bc', parent=styles['Normal'], fontSize=7, leading=9, fontName='Helvetica-Bold', alignment=TA_CENTER)

        # ✅ SELECTED COMPANY FROM PURCHASE ORDER
        company = po.company_id if po.company_id else CompanyProfile.objects.first()

        company_name = self._get_company_value(company, "company_name", "name", "company_profile_name", default="")
        company_address_1 = self._get_company_value(company, "address_line_1", "address1", "address", default="")
        company_address_2 = self._get_company_value(company, "address_line_2", "address2", default="")
        company_city = self._get_company_value(company, "city", default="")
        company_state = self._get_company_value(company, "state", "state_name", default="")
        company_pincode = self._get_company_value(company, "pincode", "pin_code", "postal_code", "zip_code", default="")
        company_gstin = self._get_company_value(company, "gstin", "gst_number", "gstin_number", default="")
        company_email = self._get_company_value(company, "email", "company_email", default="")
        company_pan = self._get_company_value(company, "pan_number", "pan_no", "pan", default="")
        company_bank_name = self._get_company_value(company, "bank_name", default="")
        company_account_no = self._get_company_value(company, "account_number", "account_no", default="")
        company_branch = self._get_company_value(company, "branch_name", "branch", default="")
        company_ifsc = self._get_company_value(company, "ifsc_code", "ifsc", default="")
        company_state_code = self._get_company_value(company, "state_code", default="")
        company_phone = self._get_company_value(company, "phone", "phone_number", "mobile", "mobile_number", "contact", default="")
        company_website = self._get_company_value(company, "website", "site_url", "web", default="")
        company_authorized_person = self._get_company_value(
            company, "authorized_person_name", "authorised_person_name",
            "authorized_signatory_name", "signatory_name", default=""
        )
        company_designation = self._get_company_value(
            company, "designation", "authorized_person_designation",
            "authorised_person_designation", default=""
        )

        company_logo_path = self._get_company_file_path(
            company, "logo", "company_logo", "image", "company_image"
        )
        company_signature_path = self._get_company_file_path(
            company, "signature", "sign", "signature_image", "authorized_signature"
        )


        irn_text = getattr(po, 'irn', 'f5207314566a607fe8fcea74054f4077cbc12ad8bc9b-46a44a66047a5dcd1f1d')
        ack_no = getattr(po, 'ack_no', '132524442763340')
        ack_date = po.po_date.strftime('%d-%b-%y') if po.po_date else timezone.now().strftime('%d-%b-%y')

        logo_img = self._build_safe_image(company_logo_path, width=80, height=45)

        irn_left_table = Table(
            [
                [Paragraph("IRN", s7b), Paragraph(f": <b>{irn_text}</b>", s7)],
                [Paragraph("Ack No.", s7), Paragraph(f": <b>{ack_no}</b>", s7)],
                [Paragraph("Ack Date", s7), Paragraph(f": <b>{ack_date}</b>", s7)],
            ],
            colWidths=[50, (usable * 0.78) - 50]
        )
        irn_left_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))

        irn_table = Table([[irn_left_table, logo_img]], colWidths=[usable * 0.78, usable * 0.22])
        irn_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (0, 0), 'TOP'),
            ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 4),   # logo ko slightly niche lane ke liye
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ]))
        elements.append(irn_table)
        elements.append(Spacer(1, 4))

        distributor = po.distributor_id

        seller_parts = []

        if company_name:
            seller_parts.append(f"<b>{company_name}</b>")

        if company_address_1:
            seller_parts.append(company_address_1)

        if company_address_2:
            seller_parts.append(company_address_2)

        city_state_pin_parts = []
        if company_city:
            city_state_pin_parts.append(str(company_city))
        if company_state:
            city_state_pin_parts.append(str(company_state))

        city_state_pin = ", ".join(city_state_pin_parts)
        if company_pincode:
            if city_state_pin:
                city_state_pin = f"{city_state_pin} - {company_pincode}"
            else:
                city_state_pin = str(company_pincode)

        if city_state_pin:
            seller_parts.append(city_state_pin)

        if company_phone:
            seller_parts.append(f"Phone : {company_phone}")

        if company_email:
            seller_parts.append(f"E-Mail : {company_email}")

        if company_website:
            seller_parts.append(f"Website : {company_website}")

        if company_gstin:
            seller_parts.append(f"GSTIN/UIN : {company_gstin}")

        if company_pan:
            seller_parts.append(f"PAN : {company_pan}")

        if company_state:
            if company_state_code:
                seller_parts.append(f"State Name : {company_state}, Code : {company_state_code}")
            else:
                seller_parts.append(f"State Name : {company_state}")

        seller_para = Paragraph("<br/>".join(seller_parts), s8)

        d = distributor
        d_name = getattr(d, 'distributor_name', '') or ''
        d_addr1 = getattr(d, 'address_line_1', '') or ''
        d_addr2 = getattr(d, 'address_line_2', '') or ''
        d_city = getattr(d, 'city', '') or ''
        d_state = getattr(d, 'state', '') or ''
        d_pin = getattr(d, 'pincode', '') or ''
        d_gst = getattr(d, 'gst_number', '') or 'N/A'
        d_pan = getattr(d, 'pan_number', '') or 'N/A'

        addr_parts = [p for p in [d_addr1, d_addr2] if p]
        full_addr = ', '.join(addr_parts)
        city_line = f"{d_city}, {d_state} - {d_pin}".strip(', ')

        consignee_para = Paragraph(
            f"Consignee (Ship to)<br/>"
            f"<b>{d_name}</b><br/>"
            f"{full_addr}<br/>"
            f"{city_line}<br/>"
            f"GSTIN/UIN&nbsp;&nbsp;&nbsp;&nbsp;: {d_gst}<br/>"
            f"PAN/IT No.&nbsp;&nbsp;: {d_pan}<br/>"
            f"State Name&nbsp;&nbsp;: {d_state}",
            s8
        )

        buyer_para = Paragraph(
            f"Buyer (Bill to)<br/>"
            f"<b>{d_name}</b><br/>"
            f"{full_addr}<br/>"
            f"{city_line}<br/>"
            f"GSTIN/UIN&nbsp;&nbsp;&nbsp;&nbsp;: {d_gst}<br/>"
            f"PAN/IT No.&nbsp;&nbsp;: {d_pan}<br/>"
            f"State Name&nbsp;&nbsp;: {d_state}",
            s8
        )

        left_inner = Table(
            [[seller_para], [consignee_para], [buyer_para]],
            colWidths=[LEFT_COL - 2]
        )
        left_inner.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ('LINEBELOW', (0, 0), (0, 0), 0.5, colors.black),
            ('LINEBELOW', (0, 1), (0, 1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))

        po_num = po.po_number or ''
        po_date_str = po.po_date.strftime('%d-%b-%y') if po.po_date else ''
        rc = RIGHT_COL - 2
        c1 = rc * 0.40
        c2 = rc * 0.30
        c3 = rc * 0.30

        right_rows = [
            [Paragraph("Credit Note No.", s7), Paragraph("e-Way Bill No.", s7), Paragraph("Dated", s7)],
            [Paragraph(f"<b>{po_num}</b>", s8b), "", Paragraph(f"<b>{po_date_str}</b>", s8b)],
            [Paragraph("Mode/Terms of Payment", s7), Paragraph("<b>Cash</b>", s7), ""],
            [Paragraph(f"<b>Cash&Carry</b>")],
            [Paragraph("Original Invoice No. &amp; Date.", s7), "", Paragraph("Other References", s7)],
            [Paragraph(f"<b>{po_num}/2025-26 dt. {po_date_str}</b>", s7b), "", ""],
            [Paragraph("Dispatch Doc No.", s7), "", ""],
            ["", "", ""],
            [Paragraph("Dispatched through", s7), ""    ],
            ["", "", ""],
            [Paragraph("Terms of Delivery", s7), "", ""],
            ["", "", ""],
        ]

        right_table = Table(right_rows, colWidths=[c1, c2, c3])
        right_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 1),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ('SPAN', (0, 2), (2, 2)),
            ('SPAN', (0, 3), (2, 3)),
            ('SPAN', (0, 4), (1, 4)),
            ('SPAN', (0, 5), (1, 5)),
            ('SPAN', (0, 6), (1, 6)),
            ('SPAN', (0, 7), (1, 7)),
            ('SPAN', (0, 8), (2, 8)),
            ('SPAN', (0, 9), (2, 9)),
            ('SPAN', (0, 10), (1, 10)),
            ('SPAN', (0, 11), (1, 11)),
            ('SPAN', (0, 12), (2, 12)),
            ('SPAN', (0, 13), (2, 13)),
        ]))

        main_row = Table([[left_inner, right_table]], colWidths=[LEFT_COL, RIGHT_COL])
        main_row.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(main_row)

        cw = [
            usable * 0.045,
            usable * 0.30,
            usable * 0.10,
            usable * 0.065,
            usable * 0.09,
            usable * 0.11,
            usable * 0.05,
            usable * 0.065,
            usable * 0.175,
        ]

        prod_header = [
            Paragraph("SI<br/>No.", s7bc),
            Paragraph("Description of Goods", s8bc),
            Paragraph("HSN/SAC", s7bc),
            Paragraph("GST<br/>Rate", s7bc),
            Paragraph("Quantity", s7bc),
            Paragraph("Rate", s7bc),
            Paragraph("per", s7bc),
            Paragraph("Disc. %", s7bc),
            Paragraph("Amount", s7bc),
        ]

        prod_data = [prod_header]
        total_qty = 0
        total_val = 0.0

        items = po.product_items or []
        for i, item in enumerate(items, 1):
            qty = int(item.get('quantity', 0))

            unit_price = float(item.get('unit_distributor_price', 0) or 0)
            if unit_price == 0:
                unit_price = float(item.get('mrp', 0) or 0)

            amt = qty * unit_price
            total_qty += qty
            total_val += amt

            prod_name = item.get('product_name') or ''
            if not prod_name:
                try:
                    prod_obj = Product.objects.get(pk=item.get('product_id'))
                    prod_name = prod_obj.product_name
                except Exception:
                    prod_name = 'N/A'

            hsn = item.get('hsn_code') or ''
            if not hsn:
                try:
                    prod_obj = Product.objects.get(pk=item.get('product_id'))
                    hsn = prod_obj.hsn_code or ''
                except Exception:
                    hsn = ''
           
            prod_data.append([
                Paragraph(str(i), s8c),
                Paragraph(prod_name, s8b),
                Paragraph(hsn, s8c),
                Paragraph("18 %", s8c),
                Paragraph(f"<b>{qty} PCS</b>", s8c),
                Paragraph(f"{self.indian_format(unit_price)}", s8r),
                Paragraph("PCS", s8c),
                "",
                Paragraph(f"<b>{self.indian_format(amt)}</b>", s8br),
            ])

        prod_data.append([
            "", Paragraph("<b>Total</b>", s8br),"", "", "", "", "", "",
            Paragraph(f"<b>{self.indian_format(total_val)}</b>", s8br),
        ])

        igst_amt = round(total_val * 0.18, 2)
        prod_data.append([
            "", Paragraph("<b><i>IGST OUTPUT</i></b>", s8bi), "", "", "", "", "", "",
            Paragraph(f"<b>{self.indian_format(igst_amt)}</b>", s8br),
        ])

        num_blanks = max(8 - len(items), 2)
        for _ in range(num_blanks):
            prod_data.append([""] * 9)

        final_total = round(total_val + igst_amt, 2)
        prod_data.append([
            "",
            Paragraph("<b>Total</b>", s8br),
            "", "",
            Paragraph(f"<b>{total_qty} PCS</b>", s8bc),
            "", "", "",
            Paragraph(f"<b>Rs. {self.indian_format(final_total)}</b>", s8br),
        ])

        prod_table = Table(prod_data, colWidths=cw)
        prod_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F0F0F0')),
        ]))
        elements.append(prod_table)

        words_row = Table(
            [[
                Paragraph(f"Amount Chargeable (in words)<br/><b>INR {self.num_to_words(final_total)}</b>", s8),
                Paragraph("<i>E. &amp; O.E</i>", s8r),
            ]],
            colWidths=[usable * 0.75, usable * 0.25]
        )
        words_row.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(words_row)

        tax_cw = [usable * 0.25, usable * 0.15, usable * 0.20, usable * 0.20, usable * 0.20]

        tax_header = ["", "", Paragraph("<b>IGST</b>", s7bc), "", ""]
        tax_sub = [
            Paragraph("Taxable<br/>Value", s7bc),
            "",
            Paragraph("Rate", s7bc),
            Paragraph("Amount", s7bc),
            Paragraph("Total<br/>Tax Amount", s7bc),
        ]
        tax_vals = [
            Paragraph(f"{self.indian_format(total_val)}", s8r),
            "",
            Paragraph("18%", s8c),
            Paragraph(f"{self.indian_format(igst_amt)}", s8r),
            Paragraph(f"{self.indian_format(igst_amt)}", s8r),
        ]
        tax_total = [
            Paragraph(f"<b>Total: {self.indian_format(total_val)}</b>", s8r),
            "",
            "",
            Paragraph(f"<b>{self.indian_format(igst_amt)}</b>", s8r),
            Paragraph(f"<b>{self.indian_format(igst_amt)}</b>", s8r),
        ]

        tax_table = Table([tax_header, tax_sub, tax_vals, tax_total], colWidths=tax_cw)
        tax_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('SPAN', (0, 0), (1, 0)),
            ('SPAN', (2, 0), (3, 0)),
            ('SPAN', (0, 1), (1, 1)),
            ('SPAN', (0, 2), (1, 2)),
            ('SPAN', (0, 3), (1, 3)),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(tax_table)

        tax_words = Table(
            [[Paragraph(f"Tax Amount (in words) : &nbsp;&nbsp;<b>INR {self.num_to_words(igst_amt)}</b>", s8)]],
            colWidths=[usable]
        )
        tax_words.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(tax_words)

        remarks_text = po.remarks or ''

        remarks_lines = [f"<b><i>Remarks:</i></b>"]
        if remarks_text:
            remarks_lines.append(remarks_text)
        if company_pan:
            remarks_lines.append(f"Company's PAN : <b>{company_pan}</b>")
        remarks_para = Paragraph("<br/>".join(remarks_lines), s8)

        branch_ifsc = ""
        if company_branch and company_ifsc:
            branch_ifsc = f"{company_branch} &amp; {company_ifsc}"
        elif company_branch:
            branch_ifsc = str(company_branch)
        elif company_ifsc:
            branch_ifsc = str(company_ifsc)

        bank_lines = ["<b>Company's Bank Details</b>"]
        if company_bank_name:
            bank_lines.append(f"Bank Name : <b>{company_bank_name}</b>")
        if company_account_no:
            bank_lines.append(f"A/c No. : <b>{company_account_no}</b>")
        if branch_ifsc:
            bank_lines.append(f"Branch &amp; IFS Code : <b>{branch_ifsc}</b>")

        bank_para = Paragraph("<br/>".join(bank_lines), s8)

        signature_img = self._build_safe_image(company_signature_path, width=90, height=35)

        sign_lines = [f"<b>for {company_name}</b>"]
        if company_authorized_person:
            sign_lines.append(f"<b>{company_authorized_person}</b>")
        if company_designation:
            sign_lines.append(company_designation)
        sign_lines.append("Authorised Signatory")
        sign_text_para = Paragraph("<br/>".join(sign_lines), s8bc)

        sign_block = Table(
            [
                [signature_img],
                [sign_text_para]
            ],
            colWidths=[usable * 0.24]
        )
        sign_block.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))

        footer_top = Table(
            [[remarks_para, bank_para]],
            colWidths=[usable * 0.50, usable * 0.50]
        )
        footer_top.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ('LINEBEFORE', (1, 0), (1, 0), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(footer_top)

        footer_bottom = Table(
            [["", sign_block]],
            colWidths=[usable * 0.50, usable * 0.50]
        )
        footer_bottom.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ('LINEBEFORE', (1, 0), (1, 0), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(footer_bottom)

        elements.append(Spacer(1, 4))
        elements.append(Paragraph("<b>This is a Computer Generated Document</b>", s8bc))

        doc.build(elements)
        buffer.seek(0)
        return buffer

class HolidayAPIView(APIView):

    # ---------- GET ----------
    def get(self, request, pk=None):
        if pk:
            holiday = get_object_or_404(Holiday, pk=pk)
            serializer = HolidaySerializer(holiday)
            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        holidays = Holiday.objects.all()
        serializer = HolidaySerializer(holidays, many=True)

        return Response({
            "success": True,
            "message":"Holiday list fetched successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    # ---------- POST ----------
    def post(self, request):
        serializer = HolidaySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Holidays created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- PATCH ----------
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        holiday = get_object_or_404(Holiday, pk=pk)
        serializer = HolidaySerializer(holiday, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Holidays updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- DELETE ----------
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        holiday = get_object_or_404(Holiday, pk=pk)
        holiday.delete()

        return Response({
            "success": True,
            "message": "Holidays deleted successfully"
        }, status=status.HTTP_200_OK)


class LeaveBalanceAPIView(APIView):

    # ---------- GET ----------
    def get(self, request, pk=None):

        if pk:
            leave_balance = get_object_or_404(LeaveBalance, pk=pk)
            serializer = LeaveBalanceSerializer(leave_balance)
            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        leaves_balance = LeaveBalance.objects.all()

        # 🔹 Filter by employee_id
        employee_id = request.GET.get("employee_id")
        if employee_id:
            leaves_balance = leaves_balance.filter(employee_id=employee_id)

        serializer = LeaveBalanceSerializer(leaves_balance, many=True)

        return Response({
            "success": True,
            "message":"Leave Balance list fetched successfully",
            "count": len(serializer.data),
            "data": serializer.data
        }, status=status.HTTP_200_OK)


    # ---------- POST ----------
    def post(self, request):
        serializer = LeaveBalanceSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Leaves  created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- PATCH ----------
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        leave_balance = get_object_or_404(LeaveBalance, pk=pk)
        serializer = LeaveBalanceSerializer(
            leave_balance,
            data=request.data,
            partial=True
        )

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Leaves  updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- DELETE ----------
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        leave_balance = get_object_or_404(LeaveBalance, pk=pk)
        leave_balance.delete()

        return Response({
            "success": True,
            "message": "Leaves deleted successfully"
        }, status=status.HTTP_200_OK)


from django.shortcuts import get_object_or_404
from django.utils import timezone

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status


class LeaveRequestAPI(APIView):

    # ---------- GET (List / Detail) ----------
    def get(self, request, pk=None):

        try:

            if pk:

                leave = LeaveRequest.objects.filter(
                    pk=pk
                ).first()

                if not leave:
                    return Response({
                        "success": False,
                        "message": "Leave request not found"
                    }, status=status.HTTP_404_NOT_FOUND)

                serializer = LeaveRequestSerializer(leave)

                return Response({
                    "success": True,
                    "data": serializer.data
                }, status=status.HTTP_200_OK)

            leaves = LeaveRequest.objects.all().order_by("-created_at")

            serializer = LeaveRequestSerializer(
                leaves,
                many=True
            )

            return Response({
                "success": True,
                "message": "Leave request list fetched successfully",
                "count": len(serializer.data),
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:

            return Response({
                "success": False,
                "message": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ---------- POST (Create) ----------
    def post(self, request):

        try:

            serializer = LeaveRequestSerializer(
                data=request.data
            )

            if serializer.is_valid():

                serializer.save()

                return Response({
                    "success": True,
                    "message": "Leaves created successfully",
                    "data": serializer.data
                }, status=status.HTTP_201_CREATED)

            return Response({
                "success": False,
                "message": "Invalid data",
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:

            return Response({
                "success": False,
                "message": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ---------- PATCH (Update / Approve / Reject) ----------
    def patch(self, request, pk=None):

        try:

            if not pk:
                return Response({
                    "success": False,
                    "message": "pk is required"
                }, status=status.HTTP_400_BAD_REQUEST)

            leave = LeaveRequest.objects.filter(
                pk=pk
            ).first()

            if not leave:
                return Response({
                    "success": False,
                    "message": "Leave request not found"
                }, status=status.HTTP_404_NOT_FOUND)

            serializer = LeaveRequestSerializer(
                leave,
                data=request.data,
                partial=True
            )

            if serializer.is_valid():

                leave = serializer.save()

                # Automatically set approved info
                if (
                    leave.status == "approved"
                    and request.data.get("approved_by")
                ):
                    leave.approved_by_id = request.data.get("approved_by")
                    leave.approved_at = timezone.now()
                    leave.save()

                return Response({
                    "success": True,
                    "message": "Leave request updated successfully",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)

            return Response({
                "success": False,
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:

            return Response({
                "success": False,
                "message": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ---------- DELETE ----------
    def delete(self, request, pk=None):

        try:

            if not pk:
                return Response({
                    "success": False,
                    "message": "pk is required"
                }, status=status.HTTP_400_BAD_REQUEST)

            leave = LeaveRequest.objects.filter(
                pk=pk
            ).first()

            if not leave:
                return Response({
                    "success": False,
                    "message": "Leave request not found"
                }, status=status.HTTP_404_NOT_FOUND)

            if leave.status == "approved":

                balance = LeaveBalance.objects.filter(
                    employee_id=leave.employee_id,
                    leave_type=leave.leave_type
                ).first()

                if balance:

                    balance.used_days -= leave.total_leaves

                    if balance.used_days < 0:
                        balance.used_days = 0

                    balance.save()

            leave.delete()

            return Response({
                "success": True,
                "message": "Leave request deleted successfully"
            }, status=status.HTTP_200_OK)

        except Exception as e:

            return Response({
                "success": False,
                "message": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ExpenseClaimsAPIView(APIView):

    # ---------- GET (All or Single) ----------
    def get(self, request, pk=None):
        if pk:
            claim = get_object_or_404(ExpenseClaim, pk=pk)
            serializer = ExpenseClaimSerializer(claim)
            return Response(
                {"success": True, "data": serializer.data},
                status=status.HTTP_200_OK
            )

        claims = ExpenseClaim.objects.all().order_by("-created_at")
        serializer = ExpenseClaimSerializer(claims, many=True)

        return Response(
            {
                "success": True,
                "message":"Expense Claim list fetched successfully",
                "count": claims.count(),
                "data": serializer.data
            },
            status=status.HTTP_200_OK
        )

    # ---------- POST ----------
    def post(self, request):
        serializer = ExpenseClaimSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Claim created successfully",
                    "data": serializer.data
                },
                status=status.HTTP_201_CREATED
            )

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- PATCH ----------
    def patch(self, request, pk=None):
        if not pk:
            return Response(
                {"success": False, "message": "pk is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        claim = get_object_or_404(ExpenseClaim, pk=pk)
        serializer = ExpenseClaimSerializer(
            claim, data=request.data, partial=True
        )

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Claim updated successfully",
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )

        return Response(
            {"success": False, "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )

    # ---------- DELETE ----------
    def delete(self, request, pk=None):
        if not pk:
            return Response(
                {"success": False, "message": "pk is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        claim = get_object_or_404(ExpenseClaim, pk=pk)
        claim.delete()

        return Response(
            {
                "success": True,
                "message": "Claim deleted successfully"
            },
            status=status.HTTP_200_OK
        )

from django.utils.dateparse import parse_datetime

class LeadFollowUpPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class LeadFollowUpAPIView(APIView):

    def get(self, request, pk=None):

        # 🔹 Single
        if pk:
            followup = get_object_or_404(LeadFollowUp, pk=pk)
            serializer = LeadFollowUpSerializer(followup)

            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔹 Queryset (IMPORTANT: ordered)
        qs = LeadFollowUp.objects.all().order_by("-created_at")

        # ---------- 🔹 FILTERS ----------
        lead_id = request.query_params.get("lead_id", "").strip()
        employee_id = request.query_params.get("employee_id", "").strip()
        status_param = request.query_params.get("status", "").strip()
        created_at_param = request.query_params.get("created_at", "").strip()

        if lead_id:
            qs = qs.filter(lead_id=lead_id)

        if employee_id:
            qs = qs.filter(employee_id=employee_id)

        # 🔹 Boolean filter
        if status_param.lower() == "true":
            qs = qs.filter(status=True)
        elif status_param.lower() == "false":
            qs = qs.filter(status=False)

        # 🔹 Datetime filter
        if created_at_param:
            dt_obj = parse_datetime(created_at_param)
            if dt_obj:
                qs = qs.filter(created_at=dt_obj)

        # 🔹 Pagination
        paginator = LeadFollowUpPagination()
        paginated_qs = paginator.paginate_queryset(qs, request)

        serializer = LeadFollowUpSerializer(paginated_qs, many=True)

        return Response({
            "success": True,
            "count": qs.count(),
            "message":"Leadfollow up fatch sucessfully..",
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        }, status=status.HTTP_200_OK)


    # ---------- POST ----------
    def post(self, request):
        serializer = LeadFollowUpSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Follow-up created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- PATCH ----------
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        leadfollowup = get_object_or_404(LeadFollowUp, pk=pk)
        serializer = LeadFollowUpSerializer(leadfollowup, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Lead follow-up updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- DELETE ----------
    def delete(self, request, pk):
        followup = get_object_or_404(LeadFollowUp, pk=pk)
        followup.delete()

        return Response({
            "success": True,
            "message": "Follow-up deleted successfully"
        }, status=status.HTTP_204_NO_CONTENT)

class TodayFollowUpAPI(APIView):

    def get(self, request, employee_id):

        today = timezone.now().date()

        followups = LeadFollowUp.objects.filter(
            employee_id=employee_id,
            followup_date=today
        ).order_by("-created_at")

        serializer = LeadFollowUpSerializer(followups, many=True)

        return Response({
            "success": True,
            "message":"Today Followup list fetched successfully",
            "employee_id": employee_id,
            "date": today,
            "count": followups.count(),
            "data": serializer.data
        }, status=status.HTTP_200_OK)

class WareHouseAPIView(APIView):

    # ---------- GET ----------
    def get(self, request, pk=None):

        if pk:
            warehouse = get_object_or_404(WareHouse, pk=pk)
            serializer = WareHouseSerializer(warehouse)
            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔹 Get query param safely
        name = request.query_params.get('name', '').strip()

        warehouses = WareHouse.objects.all()

        # 🔹 Apply filter only if name exists
        if name:
            warehouses = warehouses.filter(name__icontains=name)

        serializer = WareHouseSerializer(warehouses, many=True)

        return Response({
            "success": True,
            "message":"Warehouse list fetched successfully",
            "count": warehouses.count(),
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    # ---------- POST ----------
    def post(self, request):
        serializer = WareHouseSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Warehouse created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- PATCH ----------
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        warehouse = get_object_or_404(WareHouse, pk=pk)
        serializer = WareHouseSerializer(warehouse, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Warehouse updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- DELETE ----------
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "Id is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        warehouse = get_object_or_404(WareHouse, pk=pk)
        warehouse.delete()

        return Response({
            "success": True,
            "message": "Warehouse deleted successfully"
        }, status=status.HTTP_200_OK)


class LocationAPIView(APIView):

    # ---------- GET ----------
    def get(self, request, pk=None):
        if pk:
            location = get_object_or_404(Location, pk=pk)
            serializer = LocationSerializer(location)
            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        locations = Location.objects.all()

        # 🔍 Query Params
        warehouse_id = request.query_params.get('warehouse_id')
        warehouse_name = request.query_params.get('warehouse_name')
        status_param = request.query_params.get('status')
        name = request.query_params.get('name', '').strip()   # ✅ NEW

        # ✅ Filter by warehouse_id
        if warehouse_id:
            locations = locations.filter(warehouse_id_id=warehouse_id)

        # ✅ Filter by warehouse_name
        if warehouse_name:
            locations = locations.filter(
                warehouse_id__name__icontains=warehouse_name
            )

        # ✅ Filter by location name ✅ NEW
        if name:
            locations = locations.filter(name__icontains=name)

        # ✅ Filter by status
        if status_param:
            status_param = status_param.lower()

            if status_param in ['active', 'deactive']:
                locations = locations.filter(status=status_param)
            else:
                return Response({
                    "success": False,
                    "message": "Status must be 'active' or 'deactive'"
                }, status=status.HTTP_400_BAD_REQUEST)

        serializer = LocationSerializer(locations, many=True)

        return Response({
            "success": True,
            "message":"Location list fetched successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    # ---------- POST ----------
    def post(self, request):
        serializer = LocationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Location created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)



    # ---------- PATCH ----------
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        location = get_object_or_404(Location, pk=pk)
        serializer = LocationSerializer(location, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Location updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- DELETE ----------
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "Id is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        location = get_object_or_404(Location, pk=pk)
        location.delete()

        return Response({
            "success": True,
            "message": "Location deleted successfully"
        }, status=status.HTTP_200_OK)



class DistributerOrdersAPIView(APIView):

    # ---------- GET ----------
    def get(self, request, pk=None):

        # -------- Single Record --------
        if pk:
            order = get_object_or_404(DistributerOrders, pk=pk)
            serializer = DistributerOrdersSerializer(order)

            return Response(
                {
                    "success": True,
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )

        # -------- List --------
        orders = DistributerOrders.objects.all().order_by("-id")
        serializer = DistributerOrdersSerializer(orders, many=True)

        return Response(
            {
                "success": True,
                "message":"Distributor order list fetched successfully",
                "count": orders.count(),
                "data": serializer.data
            },
            status=status.HTTP_200_OK
        )
 

    # ---------- DELETE (Optional) ----------
    def delete(self, request, pk=None):

        order = get_object_or_404(DistributerOrders, pk=pk)
        order.delete()

        return Response(
            {
                "success": True,
                "message": "Distributor order deleted successfully"
            },
            status=status.HTTP_200_OK
        )

class PicProductPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


class PicProductAPIView(APIView):

    def get(self, request, pk=None):

        # 🔹 Detail
        if pk:
            obj = get_object_or_404(Pic_product, pk=pk)  # ✅ fixed model name
            serializer = PicProductSerializer(obj)

            return Response({
                "status": "success",
                "message": "Pic product fetched successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔹 Queryset
        qs = Pic_product.objects.all().order_by("-id")

        # 🔹 Filters
        product_id = request.query_params.get("product", "").strip()
        if product_id:
            qs = qs.filter(product_id=product_id)

        # 🔹 Pagination
        paginator = PicProductPagination()
        paginated_qs = paginator.paginate_queryset(qs, request)

        serializer = PicProductSerializer(paginated_qs, many=True)

        return Response({
            "status": "success",
            "message": "Pic product list fetched successfully",
            "count": qs.count(),
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        }, status=status.HTTP_200_OK)


    # ---------- POST ----------
    def post(self, request):

        serializer = PicProductSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # ✅ Get PurchaseOrder object directly
        po = serializer.validated_data["purchase_order_id"]

        # ✅ Ensure Purchase Order is APPROVED
        if po.status != "APPROVED":
            return Response({
                "success": False,
                "message": "Purchase Order must be APPROVED"
            }, status=status.HTTP_400_BAD_REQUEST)

        # ✅ Check if already exists (optional safety)
        if Pic_product.objects.filter(purchase_order_id=po).exists():
            return Response({
                "success": False,
                "message": "Pic Product already created for this Purchase Order"
            }, status=status.HTTP_400_BAD_REQUEST)

        # ✅ Create Pic_product
        pic = Pic_product.objects.create(
            purchase_order_id=po
        )

        return Response({
            "success": True,
            "message": "Picked Product created successfully",
            "data": PicProductSerializer(pic).data
        }, status=status.HTTP_201_CREATED)


class PackProductPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100

class PackProductAPIView(APIView):

    def get(self, request, pk=None):

        # 🔹 Detail
        if pk:
            obj = get_object_or_404(Pack_product, pk=pk)
            serializer = PackProductSerializer(obj)

            return Response({
                "status": "success",
                "message": "Pack product fetched successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔹 Queryset
        qs = Pack_product.objects.all().order_by("-id")

        # 🔹 Filters
        product_id = request.query_params.get("product", "").strip()
        if product_id:
            qs = qs.filter(product_id=product_id)

        # 🔹 Pagination
        paginator = PackProductPagination()
        paginated_qs = paginator.paginate_queryset(qs, request)

        serializer = PackProductSerializer(paginated_qs, many=True)

        return Response({
            "status": "success",
            "message": "Pack product list fetched successfully",
            "count": qs.count(),
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        }, status=status.HTTP_200_OK)


    # ---------- POST ----------
    def post(self, request):

        serializer = PackProductSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # ✅ Get PurchaseOrder object directly
        po = serializer.validated_data["purchase_order_id"]

        # ✅ Ensure Purchase Order is APPROVED
        if po.status != "APPROVED":
            return Response({
                "success": False,
                "message": "Purchase Order must be APPROVED"
            }, status=status.HTTP_400_BAD_REQUEST)

        # ✅ Check if already exists (optional safety)
        if Pack_product.objects.filter(purchase_order_id=po).exists():
            return Response({
                "success": False,
                "message": "Pack Product already created for this Purchase Order"
            }, status=status.HTTP_400_BAD_REQUEST)

        # ✅ Create Pack_product
        pack = Pack_product.objects.create(
            purchase_order_id=po
        )

        return Response({
            "success": True,
            "message": "Packed Product created successfully",
            "data": PackProductSerializer(pack).data
        }, status=status.HTTP_201_CREATED)




class ContactPagination(PageNumberPagination):
    page_size = 10  # default items per page
    page_size_query_param = 'page_size'  # allow dynamic page size
    max_page_size = 100


class ContactAPIView(APIView):
    def get(self, request, pk=None):

        # 🔹 Single record
        if pk:
            contact = get_object_or_404(Contact, pk=pk)
            serializer = ContactSerializer(contact)
            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔹 List with filtering
        contacts = Contact.objects.all().order_by('-id')

        # ✅ Filters
        name = request.query_params.get('name')
        email = request.query_params.get('email')
        mobile = request.query_params.get('mobile')

        if name:
            contacts = contacts.filter(name__icontains=name)

        if email:
            contacts = contacts.filter(email__icontains=email)

        if mobile:
            contacts = contacts.filter(mobile__icontains=mobile)

        # 🔹 Pagination
        paginator = ContactPagination()
        paginated_contacts = paginator.paginate_queryset(contacts, request)

        serializer = ContactSerializer(paginated_contacts, many=True)

        # ✅ ONLY CHANGE IS HERE
        return Response({
            "success": True,
            "message":"Contact fetched successfully",
            "count": contacts.count(),
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        })


    # ---------- POST ----------
    def post(self, request):
        serializer = ContactSerializer(data=request.data)

        if not serializer.is_valid():
            return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

        contact = serializer.save()

        try:
            # ==============================
            # ✅ 1. EMAIL TO ADMIN
            # ==============================
            admin_subject = "📩 New Contact Form Submission"

            admin_message = f"""
            A new contact request has been submitted:

            Name: {contact.name}
            Email: {contact.email}
            Mobile: {contact.mobile}

            Message:
            {contact.message}

            Please respond to the user as soon as possible.
            """

            send_mail(
                admin_subject,
                admin_message,
                settings.EMAIL_HOST_USER,
                ['jcchitroda24@gmail.com'],   # ✅ admin email
                fail_silently=False,
            )

            # ==============================
            # ✅ 2. EMAIL TO CUSTOMER
            # ==============================
            customer_subject = "✅ We Received Your Message"

            customer_message = f"""
            Dear {contact.name},

            Thank you for contacting us!

            We have successfully received your message:
            "{contact.message}"

            Our team will get back to you shortly.

            Best Regards,
            HOGO AUTO FILMS
            Support Team
            """

            send_mail(
                customer_subject,
                customer_message,
                settings.EMAIL_HOST_USER,
                [contact.email],   # ✅ customer email
                fail_silently=False,
            )

        except Exception as e:
            return Response({
                "success": True,
                "message": "Contact created but email failed",
                "error": str(e),
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "success": True,
            "message": "Contact created & emails sent successfully",
            "data": serializer.data
        }, status=status.HTTP_201_CREATED)



    # ---------- PATCH ----------
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        contact = get_object_or_404(Contact, pk=pk)
        serializer = ContactSerializer(contact, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Contact updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


    # ---------- DELETE ----------
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "Id is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        contact = get_object_or_404(Contact, pk=pk)
        contact.delete()

        return Response({
            "success": True,
            "message": "Contact deleted successfully"
        }, status=status.HTTP_200_OK)



class PoPaymentAPIView(APIView):

    def get(self, request, pk=None):

        date = request.GET.get("date")
        month = request.GET.get("month")
        year = request.GET.get("year")

        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        start_month = request.GET.get("start_month")
        end_month = request.GET.get("end_month")

        start_year = request.GET.get("start_year")
        end_year = request.GET.get("end_year")

        approved_by = request.GET.get("approved_by")  # ✅ NEW

        payments = Po_payment.objects.all()

        if pk:
            payments = payments.filter(purchase_order_id=pk)

        # ✅ Approved by filter
        if approved_by:
            payments = payments.filter(approved_by=approved_by)

        # Date filter
        if date:
            payments = payments.filter(payment_date=date)

        # Month + year filter
        if month and year:
            payments = payments.filter(payment_date__month=month, payment_date__year=year)

        # Date range filter
        if start_date and end_date:
            payments = payments.filter(payment_date__range=[start_date, end_date])

        # Month range filter
        if start_month and end_month and year:
            payments = payments.filter(
                payment_date__month__gte=start_month,
                payment_date__month__lte=end_month,
                payment_date__year=year
            )

        # Year range filter
        if start_year and end_year:
            payments = payments.filter(
                payment_date__year__gte=start_year,
                payment_date__year__lte=end_year
            )

        result = {}

        for payment in payments:
            po_id = payment.purchase_order.id

            if po_id not in result:
                result[po_id] = {
                    "purchase_order": po_id,
                    "po_number": payment.purchase_order.po_number,
                    "distributor": payment.distributor.id,
                    "distributor_name": str(payment.distributor),
                    "payments": []
                }

            serializer = PoPaymentSerializer(payment, context={"request": request})
            payment_data = serializer.data

            result[po_id]["payments"].append({
                "id": payment_data["id"],
                "amount": payment_data["amount"],
                "payment_date": payment_data["payment_date"],
                "images": payment_data["images"],
                "approved_by": payment_data["approved_by"],
                "admin_name": payment_data["admin_name"],
                "status": payment_data["status"]
            })

        return Response({
            "success": True,
            "message": "PO Payment list fetched successfully",
            "data": list(result.values())
        }, status=200)



    # POST payment
    def post(self, request):
        serializer = PoPaymentSerializer(data=request.data, context={"request": request})

        if serializer.is_valid():
            try:
                payment = serializer.save()
                return Response({
                    "message": "Payment added successfully",
                    "data": PoPaymentSerializer(payment, context={"request": request}).data
                }, status=status.HTTP_201_CREATED)
            except DistributorInformation.DoesNotExist:
                return Response({"message": "Distributor not found"}, status=400)
            except PurchaseOrder.DoesNotExist:
                return Response({"message": "Purchase Order not found"}, status=400)
            except Admin.DoesNotExist:
                return Response({"message": "Admin not found"}, status=400)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    # PATCH payment
    def patch(self, request, pk=None):
        if not pk:
            return Response({"message": "ID is required for update"}, status=400)

        try:
            payment = Po_payment.objects.get(id=pk)
        except Po_payment.DoesNotExist:
            return Response({"message": f"Payment with ID {pk} not found"}, status=404)

        serializer = PoPaymentSerializer(payment, data=request.data, partial=True, context={"request": request})
        if serializer.is_valid():
            serializer.save()
            return Response({
                "message": "Payment updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "message": "Validation failed",
            "errors": serializer.errors
        }, status=400)

    # DELETE payment
    def delete(self, request, pk=None):
        if not pk:
            return Response({"success":False,
                "message": "ID is required for delete"}, status=400)

        try:
            payment = Po_payment.objects.get(id=pk)
        except Po_payment.DoesNotExist:
            return Response({"message": f"Payment with ID {pk} not found"}, status=404)

        payment.delete()
        return Response({"success":True,
            "message": "Payment deleted successfully"}, status=204)



class TravelPlanAPI(APIView):

    def get(self, request, id=None):

        if id:
            plan = get_object_or_404(TravelPlan, id=id)
            serializer = TravelPlanSerializer(plan)

            return Response({
                "success": True,
                "message": "Travel Plan Details",
                "count": 1,
                "data": serializer.data
            })

        plans = TravelPlan.objects.all()

        year = request.GET.get("year")
        month = request.GET.get("month")
        employee_id = request.GET.get("employee_id") or request.GET.get("employee")
        region = request.GET.get("region")

        if employee_id:
            employee_id = str(employee_id).strip()
            plans = plans.filter(employee_id=employee_id)

        # ✅ Region filter
        if region:
            region = str(region).strip()
            plans = plans.filter(region__iexact=region)

        plans = list(plans)

        if year:
            plans = [
                p for p in plans
                if p.month and p.month.split("-")[-1] == str(year)
            ]

        if month:
            plans = [
                p for p in plans
                if p.month and p.month.split("-")[0].strip().lower() == month.strip().lower()
            ]

        def get_month_order(plan):
            try:
                if not plan.month:
                    return (9999, 12)

                month_str, yr = plan.month.strip().lower().split("-")

                month_map = {
                    m.lower(): i for i, m in enumerate(calendar.month_name) if m
                }

                month_number = month_map.get(month_str, 12)

                return (int(yr), month_number)

            except:
                return (9999, 12)

        plans = sorted(plans, key=get_month_order)

        serializer = TravelPlanSerializer(plans, many=True)

        return Response({
            "success": True,
            "message": "Travel Plan List (Filtered)",
            "count": len(serializer.data),
            "applied_filters": {
                "employee_id": employee_id,
                "month": month,
                "year": year,
                "region": region
            },
            "data": serializer.data
        })


    def post(self, request):
        serializer = TravelPlanSerializer(data=request.data)

        if not serializer.is_valid():

            errors = serializer.errors

            # ✅ get first error cleanly
            message = None

            if "non_field_errors" in errors:
                message = errors["non_field_errors"][0]
            else:
                message = list(errors.values())[0][0]

            return Response({
                "success": False,
                "message": message
            }, status=status.HTTP_400_BAD_REQUEST)

        serializer.save()

        return Response({
            "success": True,
            "message": "Travel plan created successfully",
            "data": serializer.data
        }, status=status.HTTP_201_CREATED)

    def patch(self, request, id=None):

        if not id:
            return Response({
                "success": False,
                "message": "id is required"
            })

        plan = get_object_or_404(TravelPlan, id=id)

        serializer = TravelPlanSerializer(
            plan,
            data=request.data,
            partial=True
        )

        if serializer.is_valid():
            serializer.save()

            return Response({
                "success": True,
                "message": "Travel Plan Updated Successfully",
                "data": serializer.data
            })

        return Response({
            "success": False,
            "message": "Validation Error",
            "errors": serializer.errors
        })

    def delete(self, request, id=None):

        if not id:
            return Response({
                "success": False,
                "message": "id is required"
            })

        plan = get_object_or_404(TravelPlan, id=id)
        plan.delete()

        return Response({
            "success": True,
            "message": "Travel Plan Deleted Successfully"
        })

class DailyPlanView(APIView):

    # GET
    def get(self, request, pk=None):

        if pk:
            try:
                daily_plan = DailyPlan.objects.get(pk=pk)
            except DailyPlan.DoesNotExist:
                return Response({
                    "success": False,
                    "message": "DailyPlan not found"
                }, status=404)

            serializer = DailyPlanSerializer(daily_plan)

            return Response({
                "success": True,
                "message": "DailyPlan details fetched successfully",
                "count": 1,
                "data": serializer.data
            })

        daily_plans = DailyPlan.objects.all()
        serializer = DailyPlanSerializer(daily_plans, many=True)

        return Response({
            "success": True,
            "message": "DailyPlan list fetched successfully",
            "count": len(serializer.data),
            "data": serializer.data
        })
    # POST
    def post(self, request):

        serializer = DailyPlanSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "message": "DailyPlan created successfully",
                "data": serializer.data
            }, status=201)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # PATCH
    def patch(self, request, pk=None):

        if not pk:
            return Response({
                "success": False,
                "message": "id is required"
            })

        try:
            daily_plan = DailyPlan.objects.get(pk=pk)
        except DailyPlan.DoesNotExist:
            return Response({"message": "DailyPlan not found"}, status=404)

        serializer = DailyPlanSerializer(daily_plan, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "message": "DailyPlan updated successfully",
                "data": serializer.data
            })

        return Response(serializer.errors, status=400)

    # DELETE
    def delete(self, request, pk=None):

        if not pk:
            return Response({"success":False,
                "message": "id is required"}, status=400)

        try:
            daily_plan = DailyPlan.objects.get(pk=pk)
        except DailyPlan.DoesNotExist:
            return Response({"message": "DailyPlan not found"}, status=404)

        daily_plan.delete()

        return Response({
            "success":True,
            "message": "DailyPlan deleted successfully"
        }, status=200)



class SalaryPaymentAPI(APIView):

    def get(self, request, pk=None, id=None):
        try:
            # ✅ Fix: support both pk and id
            record_id = pk if pk is not None else id

            # ✅ Single record
            if record_id:
                try:
                    salary_payment = Salary_payment.objects.get(pk=record_id)
                    serializer = SalaryPaymentSerializer(salary_payment)
                    return Response({
                        "success": True,
                        "message": "Salary payment fetched successfully",
                        "data": serializer.data
                    }, status=status.HTTP_200_OK)
                except Salary_payment.DoesNotExist:
                    return Response({
                        "success": False,
                        "message": "Salary payment not found"
                    }, status=status.HTTP_404_NOT_FOUND)

            # ✅ List records
            queryset = Salary_payment.objects.all().order_by("-id")

            employee = request.GET.get("employee")
            month = request.GET.get("month")
            year = request.GET.get("year")

            if employee:
                queryset = queryset.filter(employee_id=employee)

            if month:
                try:
                    queryset = queryset.filter(month=int(month))
                except ValueError:
                    return Response({
                        "success": False,
                        "message": "Month must be a valid number"
                    }, status=status.HTTP_400_BAD_REQUEST)

            if year:
                try:
                    queryset = queryset.filter(year=int(year))
                except ValueError:
                    return Response({
                        "success": False,
                        "message": "Year must be a valid number"
                    }, status=status.HTTP_400_BAD_REQUEST)

            serializer = SalaryPaymentSerializer(queryset, many=True)
            return Response({
                "success": True,
                "message": "Salary payment list fetched",
                "count": len(serializer.data),
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "success": False,
                "message": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def post(self, request):
        serializer = SalaryPaymentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Salary payment created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

class BulkSalaryPaymentAPI(APIView):
    """
    Automated system to pay salary to all active employees for a given month/year.
    """
    def post(self, request):
        month = request.data.get("month")
        year = request.data.get("year")

        if not month or not year:
            return Response({
                "success": False,
                "message": "month and year are required"
            }, status=status.HTTP_400_BAD_REQUEST)

        active_employees = Employee.objects.filter(status="Active")
        results = []
        errors = []

        for emp in active_employees:
            # Check if EmployeeSalary exists
            if not EmployeeSalary.objects.filter(employee_id=emp.id).exists():
                errors.append(f"Employee {emp.employee_code}: Salary structure not found")
                continue

            # Delete existing to allow re-calculation
            Salary_payment.objects.filter(employee=emp, month=month, year=year).delete()

            serializer = SalaryPaymentSerializer(data={"employee": emp.id, "month": month, "year": year})
            if serializer.is_valid():
                serializer.save()
                results.append({
                    "employee_code": emp.employee_code,
                    "name": f"{emp.first_name} {emp.last_name}",
                    "net_salary": serializer.data["net_salary"]
                })
            else:
                errors.append(f"Employee {emp.employee_code}: {serializer.errors}")

        return Response({
            "success": True,
            "message": f"Processed {len(results)} payments. {len(errors)} errors.",
            "data": results,
            "errors": errors
        }, status=status.HTTP_200_OK)

    # PATCH (Update)
    def patch(self, request, id):

        payment = get_object_or_404(Salary_payment, id=id)

        serializer = SalaryPaymentSerializer(payment, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()

            return Response({
                "success": True,
                "message": "Salary payment updated successfully",
                "data": serializer.data
            })

        return Response({
            "success": False,
            "message": "Validation error",
            "errors": serializer.errors
        })

    # DELETE
    def delete(self, request, id):

        payment = get_object_or_404(Salary_payment, id=id)
        payment.delete()

        return Response({
            "success": True,
            "message": "Salary payment deleted successfully"
        })


class DistributorTargetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


class DistributorTargetAPIView(APIView):

    def get(self, request, pk=None):

        # 🔹 Detail API
        if pk:
            obj = get_object_or_404(DistributorInformation_Target, pk=pk)
            serializer = DistributorTargetSerializer(obj)

            return Response({
                "status": "success",
                "message": "Data fetched successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔹 Base queryset
        qs = DistributorInformation_Target.objects.all().order_by('-id')

        # 🔹 Filter
        distributor_id = request.query_params.get('distributor', "").strip()
        if distributor_id:
            qs = qs.filter(distributor_id=distributor_id)

        # 🔹 Pagination
        paginator = DistributorTargetPagination()
        paginated_qs = paginator.paginate_queryset(qs, request)

        serializer = DistributorTargetSerializer(paginated_qs, many=True)

        return Response({
            "status": "success",
            "message": "Data fetched successfully",
            "count": qs.count(),
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = DistributorTargetSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "status": "success",
                    "message": "Created successfully",
                    "data": serializer.data
                },
                status=status.HTTP_201_CREATED
            )

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    # ✅ PATCH (Partial Update)
    def patch(self, request, pk=None):

        if not pk:
            return Response(
                {"status": "error", "message": "ID is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        obj = get_object_or_404(DistributorInformation_Target, pk=pk)

        serializer = DistributorTargetSerializer(
            obj, data=request.data, partial=True   # 👈 important
        )

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "status": "success",
                    "message": "Updated successfully",
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



    def delete(self, request, pk=None):

        if not pk:
            return Response(
                {"status": "error", "message": "ID is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        obj = get_object_or_404(DistributorInformation_Target, pk=pk)
        obj.delete()

        return Response(
            {
                "status": "success",
                "message": "Deleted successfully"
            },
            status=status.HTTP_200_OK)



class ProductTargetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


class ProductTargetAPIView(APIView):

    def get(self, request, pk=None):

        # 🔹 Single
        if pk:
            obj = get_object_or_404(Product_Target, pk=pk)
            serializer = ProductTargetSerializer(obj)

            return Response({
                "status": "success",
                "message": "Data fetched successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔹 Queryset
        qs = Product_Target.objects.all().order_by('-id')

        # 🔹 (Optional Filters)
        product_id = request.query_params.get("product", "").strip()
        if product_id:
            qs = qs.filter(product_id=product_id)

        # 🔹 Pagination
        paginator = ProductTargetPagination()
        paginated_qs = paginator.paginate_queryset(qs, request)

        serializer = ProductTargetSerializer(paginated_qs, many=True)

        return Response({
            "status": "success",
            "message": "Data fetched successfully",
            "count": qs.count(),
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        }, status=status.HTTP_200_OK)



    def post(self, request):
        serializer = ProductTargetSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "status": "success",
                    "message": "Created successfully",
                    "data": serializer.data
                },
                status=status.HTTP_201_CREATED
            )

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    # ✅ PATCH (Partial Update)
    def patch(self, request, pk=None):

        if not pk:
            return Response(
                {"status": "error", "message": "ID is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        obj = get_object_or_404(Product_Target, pk=pk)

        serializer = ProductTargetSerializer(
            obj, data=request.data, partial=True   # 👈 important
        )

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "status": "success",
                    "message": "Updated successfully",
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk=None):

        if not pk:
            return Response(
                {"status": "error", "message": "ID is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        obj = get_object_or_404(Product_Target, pk=pk)
        obj.delete()

        return Response(
            {
                "status": True,
                "message": "Deleted successfully"
            },
            status=status.HTTP_200_OK)


class CategoryTargetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100

class CategoryTargetAPIView(APIView):

    def get(self, request, pk=None):

        # 🔹 Detail API
        if pk:
            obj = get_object_or_404(Category_Target, pk=pk)
            serializer = CategoryTargetSerializer(obj)

            return Response({
                "status": "success",
                "message": "Data fetched successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔹 Queryset
        qs = Category_Target.objects.all().order_by('-id')

        # 🔹 Filter
        category_id = request.query_params.get('category', "").strip()
        if category_id:
            qs = qs.filter(category_id=category_id)

        # 🔹 Pagination
        paginator = CategoryTargetPagination()
        paginated_qs = paginator.paginate_queryset(qs, request)

        serializer = CategoryTargetSerializer(paginated_qs, many=True)

        return Response({
            "status": "success",
            "message": "Data fetched successfully",
            "count": qs.count(),
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        }, status=status.HTTP_200_OK)




    def post(self, request):
        serializer = CategoryTargetSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "status": "success",
                "message": "Created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk=None):

        if not pk:
            return Response(
                {"status": "error", "message": "ID is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        obj = get_object_or_404(Category_Target, pk=pk)

        serializer = CategoryTargetSerializer(
            obj, data=request.data, partial=True   # 👈 important
        )

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "status": "success",
                    "message": "Updated successfully",
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk=None):

        if not pk:
            return Response(
                {"status": "error", "message": "ID is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        obj = get_object_or_404(Category_Target, pk=pk)
        obj.delete()

        return Response(
            {
                "status":True,
                "message": "Deleted successfully"
            },
            status=status.HTTP_200_OK)




class CompanyProfileAPIView(APIView):

    # ✅ GET (List + Single)
    def get(self, request, pk=None):
        if pk:
            company = get_object_or_404(CompanyProfile, pk=pk)
            serializer = CompanyProfileSerializer(company)
            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        companies = CompanyProfile.objects.all()
        serializer = CompanyProfileSerializer(companies, many=True)

        return Response({
            "success": True,
            "message":"Campany profile list fetched successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)


    # ✅ POST (Create)
    def post(self, request):
        serializer = CompanyProfileSerializer(data=request.data)

        if not serializer.is_valid():   # 🔥 FIRST VALIDATE
            return Response({
                "success": False,
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            serializer.save()
            return Response({
                "success": True,
                "message": "Company created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        except IntegrityError:
            return Response({
                "success": False,
                "errors": {
                    "email": ["Email already exists."]
                }
            }, status=status.HTTP_400_BAD_REQUEST)


    # ✅ PATCH (Update)
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        company = get_object_or_404(CompanyProfile, pk=pk)
        serializer = CompanyProfileSerializer(company, data=request.data, partial=True)

        if not serializer.is_valid():
            return Response({
                "success": False,
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            serializer.save()
            return Response({
                "success": True,
                "message": "Company updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except IntegrityError:
            return Response({
                "success": False,
                "errors": {
                    "email": ["Email already exists."]
                }
            }, status=status.HTTP_400_BAD_REQUEST)


    # ✅ DELETE
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        company = get_object_or_404(CompanyProfile, pk=pk)
        company.delete()

        return Response({
            "success": True,
            "message": "Company deleted successfully"
        }, status=status.HTTP_200_OK)


class ShipmentAPIView(APIView):

    # ✅ GET (Single + All)
    def get(self, request, pk=None):
        try:
            if pk:
                shipment = get_object_or_404(Shipment, pk=pk)
                serializer = ShipmentSerializer(shipment)
                return Response({
                    "success": True,
                    "message": "Shipment fetched successfully.",
                    "count": 1,
                    "data": serializer.data
                }, status=status.HTTP_200_OK)

            shipments = Shipment.objects.all().order_by("-id")

            # ✅ Filter by order_id
            order_id = request.query_params.get("order_id")

            if order_id and order_id.isdigit():
                shipments = shipments.filter(order_id=int(order_id))

            serializer = ShipmentSerializer(shipments, many=True)

            return Response({
                "success": True,
                "message": "Shipment list fetched successfully.",
                "count": len(serializer.data),
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "success": False,
                "message": "Something went wrong while fetching shipment.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ✅ POST (Create)
    def post(self, request):
        try:
            serializer = ShipmentSerializer(data=request.data)
            if serializer.is_valid():
                shipment = serializer.save()

                return Response({
                    "success": True,
                    "message": "Shipment created successfully.",
                    "data": ShipmentSerializer(shipment).data
                }, status=status.HTTP_201_CREATED)

            return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({
                "success": False,
                "message": "Something went wrong while creating shipment.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ✅ PATCH (Update)
    def patch(self, request, pk=None):
        try:
            if not pk:
                return Response({
                    "success": False,
                    "message": "Shipment ID is required for update."
                }, status=status.HTTP_400_BAD_REQUEST)

            shipment = get_object_or_404(Shipment, pk=pk)
            serializer = ShipmentSerializer(shipment, data=request.data, partial=True)

            if serializer.is_valid():
                shipment = serializer.save()

                return Response({
                    "success": True,
                    "message": "Shipment updated successfully.",
                    "data": ShipmentSerializer(shipment).data
                }, status=status.HTTP_200_OK)

            return Response({
                "success": False,
                "message": "Shipment update failed.",
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({
                "success": False,
                "message": "Something went wrong while updating shipment.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ✅ DELETE (Delete)
    def delete(self, request, pk=None):
        try:
            if not pk:
                return Response({
                    "success": False,
                    "message": "Shipment ID is required for delete."
                }, status=status.HTTP_400_BAD_REQUEST)

            shipment = get_object_or_404(Shipment, pk=pk)
            shipment.delete()

            return Response({
                "success": True,
                "message": "Shipment deleted successfully."
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "success": False,
                "message": "Something went wrong while deleting shipment.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class CarBrandAPIView(APIView):

    # ✅ GET (List + Single)
    def get(self, request, pk=None):
        if pk:
            brand = get_object_or_404(CarBrand, pk=pk)
            serializer = CarBrandSerializer(brand)
            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        brands = CarBrand.objects.all()
        serializer = CarBrandSerializer(brands, many=True)

        return Response({
            "success": True,
            "message":"Car Brand list fetched successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)


    # ✅ POST (Create)
    def post(self, request):
        serializer = CarBrandSerializer(data=request.data)

        if not serializer.is_valid():
            return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

        serializer.save()
        return Response({
            "success": True,
            "message": "Car brand created successfully",
            "data": serializer.data
        }, status=status.HTTP_201_CREATED)


    # ✅ PATCH (Update)
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        brand = get_object_or_404(CarBrand, pk=pk)
        serializer = CarBrandSerializer(brand, data=request.data, partial=True)

        if not serializer.is_valid():
            return Response({
                "success": False,
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        serializer.save()
        return Response({
            "success": True,
            "message": "Car brand updated successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)


    # ✅ DELETE
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        brand = get_object_or_404(CarBrand, pk=pk)
        brand.delete()

        return Response({
            "success": True,
            "message": "Car brand deleted successfully"
        }, status=status.HTTP_200_OK)


class CarModelAPIView(APIView):

    # ✅ GET (List + Single)
    def get(self, request, pk=None):
        if pk:
            car = get_object_or_404(CarModel, pk=pk)
            serializer = CarModelSerializer(car)
            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        cars = CarModel.objects.all()
        serializer = CarModelSerializer(cars, many=True)

        return Response({
            "success": True,
            "message":"Car Model list fetched successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)


    # ✅ POST (Create)
    def post(self, request):
        serializer = CarModelSerializer(data=request.data)

        if not serializer.is_valid():
            return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

        serializer.save()
        return Response({
            "success": True,
            "message": "Car model created successfully",
            "data": serializer.data
        }, status=status.HTTP_201_CREATED)


    # ✅ PATCH (Update)
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        car = get_object_or_404(CarModel, pk=pk)
        serializer = CarModelSerializer(car, data=request.data, partial=True)

        if not serializer.is_valid():
            return Response({
                "success": False,
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        serializer.save()
        return Response({
            "success": True,
            "message": "Car model updated successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)


    # ✅ DELETE
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        car = get_object_or_404(CarModel, pk=pk)
        car.delete()

        return Response({
            "success": True,
            "message": "Car model deleted successfully"
        }, status=status.HTTP_200_OK)



class QuotePagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class GetQuoteAPIView(APIView):

    def get(self, request, pk=None):

        # 🔹 Single record
        if pk:
            quote = get_object_or_404(GetQuote, pk=pk)
            serializer = GetQuoteSerializer(quote)
            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        # 🔹 Base queryset
        quotes = GetQuote.objects.all().order_by('-id')

        # ✅ Filters
        full_name = request.query_params.get('full_name')
        email = request.query_params.get('email')
        service = request.query_params.get('service')
        contact = request.query_params.get('contact')
        brand_id = request.query_params.get('brand_id')
        model_id = request.query_params.get('model_id')

        if full_name:
            quotes = quotes.filter(full_name__icontains=full_name)

        if email:
            quotes = quotes.filter(email__icontains=email)

        if service:
            quotes = quotes.filter(service__icontains=service)

        if contact:
            quotes = quotes.filter(contact__icontains=contact)

        if brand_id:
            quotes = quotes.filter(brand_id=brand_id)

        if model_id:
            quotes = quotes.filter(model_id=model_id)

        # 🔹 Pagination
        paginator = QuotePagination()
        paginated_quotes = paginator.paginate_queryset(quotes, request)

        serializer = GetQuoteSerializer(paginated_quotes, many=True)

        # ✅ ONLY CHANGE HERE
        return Response({
            "success": True,
            "message":"Quotes fetched successfully",
            "count": quotes.count(),
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        })


    def post(self, request):
        serializer = GetQuoteSerializer(data=request.data)

        if not serializer.is_valid():
            return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

        quote = serializer.save()

        try:
            # ==============================
            # ✅ 1. EMAIL TO ADMIN
            # ==============================
            admin_subject = "🚨 New Quote Request Received"

            admin_message = f"""
            A new quote request has been submitted:

            Name: {quote.full_name}
            Email: {quote.email}
            Contact: {quote.contact}
            Address: {quote.Address}

            Brand: {quote.brand_id.name}
            Model: {quote.model_id.name}
            Service: {quote.service}

            Please contact the customer as soon as possible.
            """

            send_mail(
                admin_subject,
                admin_message,
                settings.EMAIL_HOST_USER,
                ['jcchitroda24@gmail.com'],   # ✅ admin email
                fail_silently=False,
            )


            # ==============================
            # ✅ 2. EMAIL TO CUSTOMER
            # ==============================
            customer_subject = "✅ Your Quote Request Has Been Received"

            customer_message = f"""
            Dear {quote.full_name},

            Thank you for contacting us!

            We have successfully received your quote request for:

            Service: {quote.service}
            Car: {quote.brand_id.name} - {quote.model_id.name}

            Our team will review your request and get back to you shortly with the best quote.

            Thank you for choosing us!

            Best Regards,
            HOGO AUTO FILMS
            Support Team
            """

            send_mail(
                customer_subject,
                customer_message,
                settings.EMAIL_HOST_USER,
                [quote.email],   # ✅ customer email
                fail_silently=False,
            )

        except Exception as e:
            return Response({
                "success": True,
                "message": "Quote created but email failed",
                "error": str(e),
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "success": True,
            "message": "Quote created & emails sent successfully",
            "data": serializer.data
        }, status=status.HTTP_201_CREATED)

    # ✅ PATCH (Update)
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        quote = get_object_or_404(GetQuote, pk=pk)
        serializer = GetQuoteSerializer(quote, data=request.data, partial=True)

        if not serializer.is_valid():
            return Response({
                "success": False,
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        serializer.save()
        return Response({
            "success": True,
            "message": "Quote updated successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)


    # ✅ DELETE
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        quote = get_object_or_404(GetQuote, pk=pk)
        quote.delete()

        return Response({
            "success": True,
            "message": "Quote deleted successfully"
        }, status=status.HTTP_200_OK)

class RegionAPIView(APIView):

    # ✅ GET (List + Single)
    def get(self, request, pk=None):
        if pk:
            obj = get_object_or_404(Region, pk=pk)
            serializer = RegionSerializer(obj)
            return Response({
                "status": "success",
                "data": serializer.data
            })

        objs = Region.objects.all().order_by('-id')
        serializer = RegionSerializer(objs, many=True)
        return Response({
            "status": "success",
            "message":"Region list fetched successfully",
            "data": serializer.data
        })


    # ✅ POST (Create)
    def post(self, request):
        serializer = RegionSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "status": "success",
                "message": "Created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


    # ✅ PATCH (Update partial)
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "status": "error",
                "message": "ID is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        obj = get_object_or_404(Region, pk=pk)
        serializer = RegionSerializer(obj, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "status": "success",
                "message": "Updated successfully",
                "data": serializer.data
            })

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


    # ✅ DELETE
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "status": False,
                "message": "ID is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        obj = get_object_or_404(Region, pk=pk)
        obj.delete()

        return Response({
            "status":True,
            "message": "Deleted successfully"
        }, status=status.HTTP_200_OK)



class RegionTargetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100



class RegionTargetAPIView(APIView):

    def get(self, request, id=None):

        # 🔹 Single Record
        if id:
            try:
                target = Region_Target.objects.get(id=id)
            except Region_Target.DoesNotExist:
                return Response({
                    "success": False,
                    "message": "Region target not found"
                }, status=status.HTTP_404_NOT_FOUND)

            serializer = RegionTargetSerializer(target)
            return Response({
                "success": True,
                "message": "Region target fetched successfully",
                "data": serializer.data
            })

        # 🔹 Queryset
        qs = Region_Target.objects.all().order_by("-id")

        # 🔹 Filters
        region = request.query_params.get("region", "").strip()
        year = request.query_params.get("year", "").strip()
        month = request.query_params.get("month", "").strip()

        # ✅ Filter by region (name OR id)
        if region:
            if region.isdigit():
                qs = qs.filter(region__id=region)   # filter by id
            else:
                qs = qs.filter(region__name__icontains=region)  # filter by name

        # ✅ Filter by year
        if year:
            qs = qs.filter(month__year=year)

        # ✅ Filter by month
        if month:
            qs = qs.filter(month__month=month)

        # 🔹 Pagination
        paginator = RegionTargetPagination()
        paginated_qs = paginator.paginate_queryset(qs, request)

        serializer = RegionTargetSerializer(paginated_qs, many=True)

        return Response({
            "success": True,
            "message": "Region target list fetched successfully",
            "count": qs.count(),
            "current_page": paginator.page.number,
            "total_pages": paginator.page.paginator.num_pages,
            "next": paginator.get_next_link(),
            "previous": paginator.get_previous_link(),
            "data": serializer.data
        })

    def post(self, request):
        serializer = RegionTargetSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "status": "success",
                    "message": "Created successfully",
                    "data": serializer.data
                },
                status=status.HTTP_201_CREATED
            )

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    # ✅ PATCH (Partial Update)
    def patch(self, request, pk=None):

        if not pk:
            return Response(
                {"status": "error", "message": "ID is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        obj = get_object_or_404(Region_Target, pk=pk)

        serializer = RegionTargetSerializer(
            obj, data=request.data, partial=True   # 👈 important
        )

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "status": "success",
                    "message": "Updated successfully",
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk=None):

        if not pk:
            return Response(
                {"status": False, "message": "ID is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        obj = get_object_or_404(Region_Target, pk=pk)
        obj.delete()

        return Response(
            {
                "status": True,
                "message": "Deleted successfully"
            },
            status=status.HTTP_200_OK)



     
# class CategoryTargetReportAPIView(APIView):
 
#     def _to_decimal(self, value, default="0.00"):
#         try:
#             if value in [None, "", "null"]:
#                 return Decimal(default)
#             return Decimal(str(value))
#         except (InvalidOperation, TypeError, ValueError):
#             return Decimal(default)
 
#     def _safe_str(self, value, default=""):
#         if value in [None, "", "null"]:
#             return default
#         return str(value)
 
#     def _get_product_name(self, product):
#         for field in ["name", "product_name", "title", "item_name"]:
#             value = getattr(product, field, None)
#             if value not in [None, "", "null"]:
#                 return str(value)
#         return str(product) if product else ""
 
#     def _get_item_product_id(self, item):
#         for key in ["product_id", "product", "product_data", "id"]:
#             value = item.get(key)
#             if value not in [None, "", "null"]:
#                 if isinstance(value, dict):
#                     return str(value.get("id") or value.get("product_id") or "")
#                 return str(value)
#         return None
 
#     def _get_order_amount_without_gst(self, item):
#         qty = self._to_decimal(item.get("quantity", 0))
 
#         if item.get("total_price") not in [None, "", "null"]:
#             return self._to_decimal(item.get("total_price"))
 
#         if item.get("subtotal") not in [None, "", "null"]:
#             return self._to_decimal(item.get("subtotal"))
 
#         if item.get("taxable_amount") not in [None, "", "null"]:
#             return self._to_decimal(item.get("taxable_amount"))
 
#         if item.get("total_amount") not in [None, "", "null"]:
#             return self._to_decimal(item.get("total_amount"))
 
#         if item.get("total_distributor_price") not in [None, "", "null"] and self._to_decimal(item.get("total_distributor_price")) > 0:
#             return self._to_decimal(item.get("total_distributor_price"))
 
#         if item.get("unit_distributor_price") not in [None, "", "null"] and self._to_decimal(item.get("unit_distributor_price")) > 0:
#             return self._to_decimal(item.get("unit_distributor_price")) * qty
 
#         if item.get("unit_price") not in [None, "", "null"]:
#             return self._to_decimal(item.get("unit_price")) * qty
 
#         if item.get("price") not in [None, "", "null"]:
#             return self._to_decimal(item.get("price")) * qty
 
#         if item.get("mrp") not in [None, "", "null"]:
#             return self._to_decimal(item.get("mrp")) * qty
 
#         return Decimal("0.00")
 
#     def _get_invoice_amount_with_gst(self, item):
#         qty = self._to_decimal(item.get("quantity", 0))
 
#         if item.get("final_amount") not in [None, "", "null"]:
#             return self._to_decimal(item.get("final_amount"))
 
#         if item.get("total_with_gst") not in [None, "", "null"]:
#             return self._to_decimal(item.get("total_with_gst"))
 
#         if item.get("amount_with_gst") not in [None, "", "null"]:
#             return self._to_decimal(item.get("amount_with_gst"))
 
#         if item.get("invoice_total") not in [None, "", "null"]:
#             return self._to_decimal(item.get("invoice_total"))
 
#         taxable = Decimal("0.00")
#         gst = Decimal("0.00")
 
#         if item.get("total_price") not in [None, "", "null"]:
#             taxable = self._to_decimal(item.get("total_price"))
#         elif item.get("subtotal") not in [None, "", "null"]:
#             taxable = self._to_decimal(item.get("subtotal"))
#         elif item.get("taxable_amount") not in [None, "", "null"]:
#             taxable = self._to_decimal(item.get("taxable_amount"))
#         elif item.get("total_amount") not in [None, "", "null"]:
#             taxable = self._to_decimal(item.get("total_amount"))
#         elif item.get("total_distributor_price") not in [None, "", "null"] and self._to_decimal(item.get("total_distributor_price")) > 0:
#             taxable = self._to_decimal(item.get("total_distributor_price"))
#         elif item.get("unit_distributor_price") not in [None, "", "null"] and self._to_decimal(item.get("unit_distributor_price")) > 0:
#             taxable = self._to_decimal(item.get("unit_distributor_price")) * qty
#         elif item.get("unit_price") not in [None, "", "null"]:
#             taxable = self._to_decimal(item.get("unit_price")) * qty
#         elif item.get("price") not in [None, "", "null"]:
#             taxable = self._to_decimal(item.get("price")) * qty
#         elif item.get("mrp") not in [None, "", "null"]:
#             taxable = self._to_decimal(item.get("mrp")) * qty
 
#         if item.get("gst_amount") not in [None, "", "null"]:
#             gst = self._to_decimal(item.get("gst_amount"))
#         elif item.get("gst") not in [None, "", "null"]:
#             gst = self._to_decimal(item.get("gst"))
#         else:
#             gst = (
#                 self._to_decimal(item.get("cgst_amount", 0)) +
#                 self._to_decimal(item.get("sgst_amount", 0)) +
#                 self._to_decimal(item.get("igst_amount", 0))
#             )
 
#         return taxable + gst
 
#     def get(self, request):
#         category_id = request.GET.get("category")
#         month_value = request.GET.get("month")
 
#         if not category_id or not month_value:
#             return Response({
#                 "success": False,
#                 "message": "category and month are required"
#             }, status=status.HTTP_400_BAD_REQUEST)
 
#         try:
#             if len(month_value) == 7:
#                 parsed_date = datetime.strptime(month_value, "%Y-%m")
#             elif len(month_value) == 10:
#                 parsed_date = datetime.strptime(month_value, "%Y-%m-%d")
#             else:
#                 return Response({
#                     "success": False,
#                     "message": "Month format must be YYYY-MM or YYYY-MM-DD"
#                 }, status=status.HTTP_400_BAD_REQUEST)
#         except ValueError:
#             return Response({
#                 "success": False,
#                 "message": "Invalid month format. Use YYYY-MM or YYYY-MM-DD"
#             }, status=status.HTTP_400_BAD_REQUEST)
 
#         year = parsed_date.year
#         month_num = parsed_date.month
 
#         category_target = Category_Target.objects.filter(
#             category_id=category_id,
#             month__year=year,
#             month__month=month_num
#         ).first()
 
#         if not category_target:
#             return Response({
#                 "success": False,
#                 "message": "No target found for this category and month"
#             }, status=status.HTTP_404_NOT_FOUND)
 
#         total_target = self._to_decimal(category_target.target, default="0.00")
 
#         products = Product.objects.filter(category_id=category_id).order_by("id")
#         product_count = products.count()
 
#         if product_count == 0:
#             return Response({
#                 "success": True,
#                 "category": self._safe_str(category_target.category),
#                 "month": category_target.month.strftime("%Y-%m") if category_target.month else "",
#                 "target": self._safe_str(category_target.target, "0"),
#                 "summary": {
#                     "total_order_without_gst": "0.00",
#                     "total_invoice_with_gst": "0.00"
#                 },
#                 "products": []
#             }, status=status.HTTP_200_OK)
 
#         per_product_target = total_target / Decimal(product_count) if product_count else Decimal("0.00")
 
#         from django.db.models import Q
#         # Deep Search: Get POs for the specified month
#         purchase_orders = PurchaseOrder.objects.filter(
#             Q(po_date__year=year, po_date__month=month_num) |
#             Q(po_date__isnull=True, created_at__year=year, created_at__month=month_num)
#         ).order_by("-id")
 
#         total_order = Decimal("0.00")
#         total_invoice = Decimal("0.00")
#         total_products = 0
#         product_data = []
 
#         gst_rate = Decimal("0.18")
 
#         for product in products:
#             order_total = Decimal("0.00")
#             invoice_total = Decimal("0.00")
#             product_qty = 0
 
#             # Use product's actual price
#             base_price = Decimal(str(product.unit_distributor_price)) if getattr(product, 'unit_distributor_price', None) else product.mrp
 
#             for po in purchase_orders:
#                 items = po.product_items or []
 
#                 if not isinstance(items, list):
#                     continue
 
#                 for item in items:
#                     if not isinstance(item, dict):
#                         continue
 
#                     item_product_id = self._get_item_product_id(item)
 
#                     if str(item_product_id) == str(product.id):
#                         qty = self._to_decimal(item.get("quantity", item.get("qty", 0)))

#                         product_qty += int(qty)

#                         item_order_amt = qty * base_price
#                         order_total += item_order_amt

#                         item_invoice_amt = item_order_amt * (Decimal("1") + gst_rate)
#                         invoice_total += item_invoice_amt
 
#             total_order += order_total
#             total_invoice += invoice_total
#             total_products += product_qty
 
#             product_data.append({
#                 "product_id": product.id if product.id is not None else 0,
#                 "product_name": self._safe_str(self._get_product_name(product), ""),
#                 "allotted_target": str(per_product_target.quantize(Decimal("0.01"))),
#                 "total_products": product_qty,
#                 "order_total_without_gst": str(order_total.quantize(Decimal("0.01"))),
#                 "actual_invoice_total_with_gst": str(invoice_total.quantize(Decimal("0.01"))),
#             })
 
#         return Response({
#             "success": True,
#             "category": self._safe_str(category_target.category),
#             "month": category_target.month.strftime("%Y-%m") if category_target.month else "",
#             "target": self._safe_str(category_target.target, "0"),
#             "summary": {
#                 "total_products": total_products,
#                 "total_order_without_gst": str(total_order.quantize(Decimal("0.01"))),
#                 "total_invoice_with_gst": str(total_invoice.quantize(Decimal("0.01")))
#             },
#             "products": product_data
#         }, status=status.HTTP_200_OK)
 
import json
import calendar
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.db.models import Q
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

class CategoryTargetReportAPIView(APIView):

    def _to_decimal(self, value, default="0.00"):
        try:
            if value in [None, "", "null"]:
                return Decimal(default)
            return Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    def _safe_str(self, value, default=""):
        if value in [None, "", "null"]:
            return default
        return str(value)

    def _get_product_name(self, product):
        for field in ["name", "product_name", "title", "item_name"]:
            value = getattr(product, field, None)
            if value not in [None, "", "null"]:
                return str(value)
        return str(product) if product else ""

    def _get_item_product_id(self, item):
        for key in ["product_id", "product", "product_data", "id"]:
            value = item.get(key)
            if value not in [None, "", "null"]:
                if isinstance(value, dict):
                    return str(value.get("id") or value.get("product_id") or "")
                return str(value)
        return None

    def _get_order_amount_without_gst(self, item):
        qty = self._to_decimal(item.get("quantity", 0))

        if item.get("total_price") not in [None, "", "null"]:
            return self._to_decimal(item.get("total_price"))

        if item.get("subtotal") not in [None, "", "null"]:
            return self._to_decimal(item.get("subtotal"))

        if item.get("taxable_amount") not in [None, "", "null"]:
            return self._to_decimal(item.get("taxable_amount"))

        if item.get("total_amount") not in [None, "", "null"]:
            return self._to_decimal(item.get("total_amount"))

        if item.get("total_distributor_price") not in [None, "", "null"] and self._to_decimal(item.get("total_distributor_price")) > 0:
            return self._to_decimal(item.get("total_distributor_price"))

        if item.get("unit_distributor_price") not in [None, "", "null"] and self._to_decimal(item.get("unit_distributor_price")) > 0:
            return self._to_decimal(item.get("unit_distributor_price")) * qty

        if item.get("unit_price") not in [None, "", "null"]:
            return self._to_decimal(item.get("unit_price")) * qty

        if item.get("price") not in [None, "", "null"]:
            return self._to_decimal(item.get("price")) * qty

        if item.get("mrp") not in [None, "", "null"]:
            return self._to_decimal(item.get("mrp")) * qty

        return Decimal("0.00")

    def _get_invoice_amount_with_gst(self, item):
        qty = self._to_decimal(item.get("quantity", 0))

        unit_price = self._to_decimal(
            item.get("unit_distributor_price", item.get("price", 0))
        )

        product_id = item.get("product_id")

        gst_percentage = Decimal("0.00")

        if product_id:
            try:
                product = Product.objects.get(id=product_id)

                if product.GST:
                    gst_percentage = self._to_decimal(
                        str(product.GST).replace("%", "")
                    )

            except Product.DoesNotExist:
                pass

        taxable = qty * unit_price

        gst_amount = (
            taxable * gst_percentage / Decimal("100")
        )

        return (taxable + gst_amount).quantize(
            Decimal("0.01")
        )

    def get(self, request):
        category_id = request.GET.get("category")
        month_value = request.GET.get("month")

        if not category_id or not month_value:
            return Response({
                "success": False,
                "message": "category and month are required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            if len(month_value) == 7:
                parsed_date = datetime.strptime(month_value, "%Y-%m")
            elif len(month_value) == 10:
                parsed_date = datetime.strptime(month_value, "%Y-%m-%d")
            else:
                return Response({
                    "success": False,
                    "message": "Month format must be YYYY-MM or YYYY-MM-DD"
                }, status=status.HTTP_400_BAD_REQUEST)
        except ValueError:
            return Response({
                "success": False,
                "message": "Invalid month format. Use YYYY-MM or YYYY-MM-DD"
            }, status=status.HTTP_400_BAD_REQUEST)

        year = parsed_date.year
        month_num = parsed_date.month

        category_target = Category_Target.objects.filter(
            category_id=category_id,
            month__year=year,
            month__month=month_num
        ).first()

        if not category_target:
            return Response({
                "success": False,
                "message": "No target found for this category and month"
            }, status=status.HTTP_404_NOT_FOUND)

        total_target = self._to_decimal(category_target.target, default="0.00")

        products = Product.objects.filter(category_id=category_id).order_by("id")
        product_count = products.count()

        if product_count == 0:
            return Response({
                "success": True,
                "category": self._safe_str(category_target.category),
                "month": category_target.month.strftime("%Y-%m") if category_target.month else "",
                "target": self._safe_str(category_target.target, "0"),
                "summary": {
                    "total_order_without_gst": "0.00",
                    "total_invoice_with_gst": "0.00"
                },
                "products": []
            }, status=status.HTTP_200_OK)

        per_product_target = total_target / Decimal(product_count) if product_count else Decimal("0.00")

        from django.db.models import Q
        # Deep Search: Get POs for the specified month
        purchase_orders = PurchaseOrder.objects.filter(
            Q(po_date__year=year, po_date__month=month_num) |
            Q(po_date__isnull=True, created_at__year=year, created_at__month=month_num)
        ).order_by("-id")

        total_order = Decimal("0.00")
        total_invoice = Decimal("0.00")
        total_products = Decimal("0")
        product_data = []

        # gst_rate = Decimal("0.18")

        for product in products:
            order_total = Decimal("0.00")
            invoice_total = Decimal("0.00")
            po_total_products = Decimal("0")

            # Use product's actual price
            base_price = Decimal(str(product.unit_distributor_price)) if getattr(product, 'unit_distributor_price', None) else product.mrp

            for po in purchase_orders:
                items = po.product_items or []

                if not isinstance(items, list):
                    continue

                for item in items:
                    if not isinstance(item, dict):
                        continue

                    item_product_id = self._get_item_product_id(item)

                    if str(item_product_id) == str(product.id):
                        qty = self._to_decimal(item.get("quantity", item.get("qty", 0)))
                        po_total_products += qty

                        # Without GST
                        item_order_amt = self._get_order_amount_without_gst(item)
                        order_total += item_order_amt

                        # With GST
                        item_invoice_amt = self._get_invoice_amount_with_gst(item)
                        invoice_total += item_invoice_amt

            total_order += order_total
            total_invoice += invoice_total
            total_products += po_total_products
            product_data.append({
                "product_id": product.id if product.id is not None else 0,
                "product_name": self._safe_str(self._get_product_name(product), ""),
                "allotted_target": str(per_product_target.quantize(Decimal("0.01"))),
                "order_total_without_gst": str(order_total.quantize(Decimal("0.01"))),
                "actual_invoice_total_with_gst": str(invoice_total.quantize(Decimal("0.01"))),
                "total_products": str(po_total_products.quantize(Decimal("1"))),
            })

        return Response({
            "success": True,
            "category": self._safe_str(category_target.category),
            "month": category_target.month.strftime("%Y-%m") if category_target.month else "",
            "target": self._safe_str(category_target.target, "0"),
            "summary": {
                "total_order_without_gst": str(total_order.quantize(Decimal("0.01"))),
                "total_invoice_with_gst": str(total_invoice.quantize(Decimal("0.01"))),
                "total_products": str(total_products.quantize(Decimal("1")))
            },
            "products": product_data
        }, status=status.HTTP_200_OK)




from decimal import Decimal, InvalidOperation
from datetime import datetime
import calendar

from django.db.models import Q
from django.http import HttpResponse

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter




# class CategoryTargetReportExcelAPIView(APIView):

#     # --------------------------------
#     # SAFE DECIMAL
#     # --------------------------------
#     def _to_decimal(self, value, default="0.00"):

#         try:

#             if value in [None, "", "null"]:
#                 return Decimal(default)

#             return Decimal(str(value))

#         except (InvalidOperation, TypeError, ValueError):

#             return Decimal(default)

#     # --------------------------------
#     # GET PRODUCT ID FROM ITEM
#     # --------------------------------
#     def _get_item_product_id(self, item):

#         for key in [
#             "product_id",
#             "product",
#             "product_data",
#             "id"
#         ]:

#             value = item.get(key)

#             if value not in [None, "", "null"]:

#                 if isinstance(value, dict):

#                     return str(
#                         value.get("id")
#                         or value.get("product_id")
#                         or ""
#                     )

#                 return str(value)

#         return None

#     # --------------------------------
#     # MAIN API
#     # --------------------------------
#     def get(self, request):

#         category_id = (
#             request.GET.get("category")
#             or request.GET.get("category_id")
#         )

#         month_value = request.GET.get("month")
#         year_value = request.GET.get("year")

#         if not category_id:

#             return Response(
#                 {
#                     "success": False,
#                     "message": "category is required"
#                 },
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         # --------------------------------
#         # YEAR + MONTH FILTER
#         # --------------------------------
#         current_date = datetime.now()

#         # DEFAULT YEAR
#         year = current_date.year
#         selected_month = None

#         # --------------------------------
#         # YEAR FILTER
#         # --------------------------------
#         if year_value:

#             try:

#                 year = int(year_value)

#             except ValueError:

#                 return Response(
#                     {
#                         "success": False,
#                         "message": "Invalid year format"
#                     },
#                     status=status.HTTP_400_BAD_REQUEST
#                 )

#         # --------------------------------
#         # MONTH FILTER
#         # --------------------------------
#         if month_value:

#             try:

#                 # MM
#                 if len(month_value) == 2:

#                     selected_month = int(month_value)

#                 # YYYY-MM
#                 elif len(month_value) == 7:

#                     parsed_date = datetime.strptime(
#                         month_value,
#                         "%Y-%m"
#                     )

#                     year = parsed_date.year
#                     selected_month = parsed_date.month

#                 # YYYY-MM-DD
#                 elif len(month_value) == 10:

#                     parsed_date = datetime.strptime(
#                         month_value,
#                         "%Y-%m-%d"
#                     )

#                     year = parsed_date.year
#                     selected_month = parsed_date.month

#                 else:

#                     return Response(
#                         {
#                             "success": False,
#                             "message": (
#                                 "Month format must be "
#                                 "MM or YYYY-MM or YYYY-MM-DD"
#                             )
#                         },
#                         status=status.HTTP_400_BAD_REQUEST
#                     )

#             except ValueError:

#                 return Response(
#                     {
#                         "success": False,
#                         "message": "Invalid month format"
#                     },
#                     status=status.HTTP_400_BAD_REQUEST
#                 )

#         # --------------------------------
#         # CATEGORY
#         # --------------------------------
#         category_obj = Category.objects.filter(
#             id=category_id
#         ).first()

#         if not category_obj:

#             return Response(
#                 {
#                     "success": False,
#                     "message": "Category not found"
#                 },
#                 status=status.HTTP_404_NOT_FOUND
#             )

#         category_name = category_obj.name or "Category"

#         # --------------------------------
#         # PRODUCTS FILTER BY CATEGORY
#         # --------------------------------
#         products = Product.objects.filter(
#             category_id=category_id
#         )

#         if not products.exists():

#             return Response(
#                 {
#                     "success": False,
#                     "message": "No products found for this category"
#                 },
#                 status=status.HTTP_404_NOT_FOUND
#             )

#         # --------------------------------
#         # PURCHASE ORDERS
#         # --------------------------------
#         purchase_orders = PurchaseOrder.objects.filter(
#             Q(po_date__year=year) |
#             Q(
#                 po_date__isnull=True,
#                 created_at__year=year
#             )
#         ).order_by("-id")

#         # --------------------------------
#         # EXCEL
#         # --------------------------------
#         workbook = Workbook()

#         worksheet = workbook.active

#         worksheet.title = "Category Target Report"

#         # --------------------------------
#         # TITLE
#         # --------------------------------
#         worksheet.merge_cells("A1:F1")

#         title_cell = worksheet["A1"]

#         if selected_month:

#             report_title = (
#                 f"{category_name} Target Report - "
#                 f"{calendar.month_name[selected_month]} {year}"
#             )

#         else:

#             report_title = (
#                 f"{category_name} Target Report - {year}"
#             )

#         title_cell.value = report_title

#         title_cell.font = Font(
#             bold=True,
#             size=16,
#             color="FFFFFF"
#         )

#         title_cell.fill = PatternFill(
#             start_color="C00000",
#             end_color="C00000",
#             fill_type="solid"
#         )

#         title_cell.alignment = Alignment(
#             horizontal="center",
#             vertical="center"
#         )

#         # --------------------------------
#         # HEADERS
#         # --------------------------------
#         headers = [
#             "SR No",
#             "Month",
#             "Target",
#             "Total Product",
#             "Order (excl. GST)",
#             "Invoice (incl. GST)",
#         ]

#         header_fill = PatternFill(
#             start_color="C00000",
#             end_color="C00000",
#             fill_type="solid"
#         )

#         for col_num, header in enumerate(headers, 1):

#             cell = worksheet.cell(
#                 row=3,
#                 column=col_num
#             )

#             cell.value = header

#             cell.font = Font(
#                 bold=True,
#                 color="FFFFFF"
#             )

#             cell.fill = header_fill

#             cell.alignment = Alignment(
#                 horizontal="center",
#                 vertical="center"
#             )

#         # --------------------------------
#         # DATA
#         # --------------------------------
#         row_num = 4

#         grand_total_target = Decimal("0.00")
#         grand_total_product = Decimal("0.00")
#         grand_total_order = Decimal("0.00")
#         grand_total_invoice = Decimal("0.00")

#         gst_rate = Decimal("0.18")

#         # --------------------------------
#         # LOOP MONTH WISE
#         # --------------------------------
#         months_to_loop = (
#             [selected_month]
#             if selected_month
#             else range(1, 13)
#         )

#         sr_no = 1

#         for month in months_to_loop:

#             # --------------------------------
#             # MONTH NAME
#             # --------------------------------
#             month_name = (
#                 calendar.month_abbr[month]
#                 + f"-{str(year)[-2:]}"
#             )

#             # --------------------------------
#             # MONTH TARGET
#             # --------------------------------
#             category_target = Category_Target.objects.filter(
#                 category_id=category_id,
#                 month__year=year,
#                 month__month=month
#             ).first()

#             month_target = Decimal("0.00")

#             if category_target:

#                 month_target = self._to_decimal(
#                     category_target.target
#                 )

#             # --------------------------------
#             # PURCHASE ORDERS MONTH FILTER
#             # --------------------------------
#             month_purchase_orders = purchase_orders.filter(
#                 Q(po_date__month=month) |
#                 Q(
#                     po_date__isnull=True,
#                     created_at__month=month
#                 )
#             )
#             month_product_total = Decimal("0.00")
#             month_order_total = Decimal("0.00")
#             month_invoice_total = Decimal("0.00")

#             # --------------------------------
#             # PRODUCT LOOP
#             # --------------------------------
#             for product in products:

#                 base_price = Decimal(
#                     str(
#                         getattr(
#                             product,
#                             "unit_distributor_price",
#                             0
#                         ) or 0
#                     )
#                 )

#                 if base_price <= 0:

#                     base_price = self._to_decimal(
#                         getattr(product, "mrp", 0)
#                     )

#                 for po in month_purchase_orders:

#                     items = po.product_items or []

#                     if not isinstance(items, list):
#                         continue

#                     for item in items:

#                         if not isinstance(item, dict):
#                             continue

#                         item_product_id = (
#                             self._get_item_product_id(item)
#                         )

#                         if str(item_product_id) == str(product.id):

#                             qty = self._to_decimal(
#                                 item.get(
#                                     "quantity",
#                                     item.get("qty", 0)
#                                 )
#                             )
#                             month_product_total += qty
#                             item_order_amt = (
#                                 qty * base_price
#                             )

#                             item_invoice_amt = (
#                                 item_order_amt *
#                                 (Decimal("1") + gst_rate)
#                             )

#                             month_order_total += item_order_amt
#                             month_invoice_total += item_invoice_amt

#             # --------------------------------
#             # GRAND TOTALS
#             # --------------------------------
#             grand_total_target += month_target
#             grand_total_product += month_product_total
#             grand_total_order += month_order_total
#             grand_total_invoice += month_invoice_total

#             # --------------------------------
#             # SR NO
#             # --------------------------------
#             sr_cell = worksheet.cell(
#                 row=row_num,
#                 column=1
#             )

#             sr_cell.value = sr_no

#             sr_cell.alignment = Alignment(
#                 horizontal="center",
#                 vertical="center"
#             )

#             # --------------------------------
#             # MONTH
#             # --------------------------------
#             month_cell = worksheet.cell(
#                 row=row_num,
#                 column=2
#             )

#             month_cell.value = month_name

#             month_cell.alignment = Alignment(
#                 horizontal="center",
#                 vertical="center"
#             )

#             # --------------------------------
#             # TARGET
#             # --------------------------------
#             target_cell = worksheet.cell(
#                 row=row_num,
#                 column=3
#             )

#             target_cell.value = float(
#                 month_target.quantize(
#                     Decimal("0.01")
#                 )
#             )

#             target_cell.number_format = '0.00'

#             target_cell.alignment = Alignment(
#                 horizontal="center",
#                 vertical="center"
#             )

#             # --------------------------------
#             # TOTAL PRODUCT
#             # --------------------------------
#             product_cell = worksheet.cell(
#                 row=row_num,
#                 column=4
#             )

#             product_cell.value = float(
#                 month_product_total.quantize(
#                     Decimal("0.01")
#                 )
#             )

#             product_cell.number_format = '0.00'

#             product_cell.alignment = Alignment(
#                 horizontal="center",
#                 vertical="center"
#             )

#             # --------------------------------
#             # ORDER VALUE
#             # --------------------------------
#             order_cell = worksheet.cell(
#                 row=row_num,
#                 column=5
#             )

#             order_cell.value = float(
#                 month_order_total.quantize(
#                     Decimal("0.01")
#                 )
#             )

#             order_cell.number_format = '0.00'

#             order_cell.alignment = Alignment(
#                 horizontal="center",
#                 vertical="center"
#             )

#             # --------------------------------
#             # INVOICE VALUE
#             # --------------------------------
#             invoice_cell = worksheet.cell(
#                 row=row_num,
#                 column=6
#             )

#             invoice_cell.value = float(
#                 month_invoice_total.quantize(
#                     Decimal("0.01")
#                 )
#             )

#             invoice_cell.number_format = '0.00'

#             invoice_cell.alignment = Alignment(
#                 horizontal="center",
#                 vertical="center"
#             )

#             row_num += 1
#             sr_no += 1

#         # --------------------------------
#         # TOTAL ROW
#         # --------------------------------
#         total_fill = PatternFill(
#             start_color="FFF2CC",
#             end_color="FFF2CC",
#             fill_type="solid"
#         )

#         total_label = worksheet.cell(
#             row=row_num,
#             column=2
#         )

#         total_label.value = "TOTAL"

#         total_label.font = Font(bold=True)

#         total_label.fill = total_fill

#         total_label.alignment = Alignment(
#             horizontal="center",
#             vertical="center"
#         )

#         # TARGET TOTAL
#         total_target_cell = worksheet.cell(
#             row=row_num,
#             column=3
#         )

#         total_target_cell.value = float(
#             grand_total_target.quantize(
#                 Decimal("0.01")
#             )
#         )

#         total_target_cell.number_format = '0.00'

#         total_target_cell.font = Font(bold=True)

#         total_target_cell.fill = total_fill

#         total_target_cell.alignment = Alignment(
#             horizontal="center",
#             vertical="center"
#         )

#         # TOTAL PRODUCT
#         total_product_cell = worksheet.cell(
#             row=row_num,
#             column=4
#         )

#         total_product_cell.value = float(
#             grand_total_product.quantize(
#                 Decimal("0.01")
#             )
#         )

#         total_product_cell.number_format = '0.00'

#         total_product_cell.font = Font(bold=True)

#         total_product_cell.fill = total_fill

#         total_product_cell.alignment = Alignment(
#             horizontal="center",
#             vertical="center"
#         )

#         # ORDER TOTAL
#         total_order_cell = worksheet.cell(
#             row=row_num,
#             column=5
#         )

#         total_order_cell.value = float(
#             grand_total_order.quantize(
#                 Decimal("0.01")
#             )
#         )

#         total_order_cell.number_format = '0.00'

#         total_order_cell.font = Font(bold=True)

#         total_order_cell.fill = total_fill

#         total_order_cell.alignment = Alignment(
#             horizontal="center",
#             vertical="center"
#         )

#         # INVOICE TOTAL
#         total_invoice_cell = worksheet.cell(
#             row=row_num,
#             column=6
#         )

#         total_invoice_cell.value = float(
#             grand_total_invoice.quantize(
#                 Decimal("0.01")
#             )
#         )

#         total_invoice_cell.number_format = '0.00'

#         total_invoice_cell.font = Font(bold=True)

#         total_invoice_cell.fill = total_fill

#         total_invoice_cell.alignment = Alignment(
#             horizontal="center",
#             vertical="center"
#         )
#         # --------------------------------
#         # AUTO WIDTH FIX
#         # --------------------------------
#         for column in range(
#             1,
#             worksheet.max_column + 1
#         ):

#             max_length = 0

#             column_letter = get_column_letter(column)

#             for row in range(
#                 1,
#                 worksheet.max_row + 1
#             ):

#                 cell = worksheet.cell(
#                     row=row,
#                     column=column
#                 )

#                 try:

#                     if cell.value:

#                         max_length = max(
#                             max_length,
#                             len(str(cell.value))
#                         )

#                 except Exception:
#                     pass

#             worksheet.column_dimensions[
#                 column_letter
#             ].width = max_length + 5

#         # --------------------------------
#         # RESPONSE
#         # --------------------------------
#         response = HttpResponse(
#             content_type=(
#                 "application/vnd.openxmlformats-"
#                 "officedocument.spreadsheetml.sheet"
#             )
#         )

#         safe_category_name = (
#             str(category_name)
#             .replace(" ", "_")
#             .replace("/", "_")
#         )

#         if selected_month:

#             file_name = (
#                 f"{safe_category_name}_"
#                 f"{year}_{selected_month}_report.xlsx"
#             )

#         else:

#             file_name = (
#                 f"{safe_category_name}_target_report_{year}.xlsx"
#             )

#         response[
#             "Content-Disposition"
#         ] = (
#             f'attachment; filename="{file_name}"'
#         )

#         workbook.save(response)

#         return response






class CategoryTargetReportExcelAPIView(APIView):

    # ------------------------------------------------------------------ #
    # HELPERS
    # ------------------------------------------------------------------ #

    def _to_decimal(self, value, default="0.00"):
        try:
            if value in [None, "", "null"]:
                return Decimal(default)
            return Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    def _get_item_product_id(self, item):
        for key in ["product_id", "product", "product_data", "id"]:
            value = item.get(key)
            if value not in [None, "", "null"]:
                if isinstance(value, dict):
                    return str(value.get("id") or value.get("product_id") or "")
                return str(value)
        return None

    def _parse_items(self, raw):
        if not raw:
            return []
        if isinstance(raw, list):
            return raw
        if isinstance(raw, str):
            try:
                parsed = json.loads(raw)
                return parsed if isinstance(parsed, list) else []
            except (json.JSONDecodeError, ValueError):
                return []
        return []

    def _get_base_price(self, item, product):
        for key in ["unit_distributor_price", "mrp", "unit_price", "price"]:
            val = item.get(key)
            if val not in [None, "", "null", 0, "0", "0.00"]:
                try:
                    return Decimal(str(val))
                except Exception:
                    pass
        mrp = getattr(product, "mrp", None)
        if mrp not in [None, "", "null"]:
            try:
                return Decimal(str(mrp))
            except Exception:
                pass
        return Decimal("0.00")

    def _parse_year_month(self, year_value, month_value):
        year           = datetime.now().year
        selected_month = None

        if year_value:
            try:
                year = int(year_value)
            except ValueError:
                return None, None, {"success": False, "message": "Invalid year format"}

        if month_value:
            try:
                if len(month_value) == 2:
                    selected_month = int(month_value)
                elif len(month_value) == 7:
                    parsed         = datetime.strptime(month_value, "%Y-%m")
                    year, selected_month = parsed.year, parsed.month
                elif len(month_value) == 10:
                    parsed         = datetime.strptime(month_value, "%Y-%m-%d")
                    year, selected_month = parsed.year, parsed.month
                else:
                    return None, None, {
                        "success": False,
                        "message": "Month format must be MM, YYYY-MM, or YYYY-MM-DD",
                    }
            except ValueError:
                return None, None, {"success": False, "message": "Invalid month format"}

        return year, selected_month, None

    def _calc_month_data(self, category_id, products, month_pos, year, month):
        GST = Decimal("0.18")

        category_target = Category_Target.objects.filter(
            category_id=category_id,
            month__year=year,
            month__month=month,
        ).first()
        month_target = self._to_decimal(
            category_target.target if category_target else 0
        )

        product_map   = {str(p.id): p for p in products}
        product_ids   = set(product_map.keys())
        product_total = Decimal("0.00")
        order_total   = Decimal("0.00")
        invoice_total = Decimal("0.00")

        for po in month_pos:
            for item in self._parse_items(po.product_items):
                if not isinstance(item, dict):
                    continue
                pid = self._get_item_product_id(item)
                if not pid or pid not in product_ids:
                    continue
                qty = self._to_decimal(
                    item.get("quantity", item.get("qty", 0))
                )
                if qty <= 0:
                    continue
                product        = product_map[pid]
                base_price     = self._get_base_price(item, product)
                order_amt      = qty * base_price
                product_total += qty
                order_total   += order_amt
                invoice_total += order_amt * (Decimal("1") + GST)

        return month_target, product_total, order_total, invoice_total

    # ------------------------------------------------------------------ #
    # STYLING
    # ------------------------------------------------------------------ #

    RED_FILL    = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    GRAY_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    CENTER      = Alignment(horizontal="center", vertical="center")
    LEFT        = Alignment(horizontal="left",   vertical="center")

    # ✅ Category column removed
    HEADERS = [
        "SR No",
        "Month",
        "Target",
        "Total Product",
        "Order (excl. GST)",
        "Invoice (incl. GST)",
    ]

    def _style_cell(self, cell, value, fmt=None, bold=False, fill=None,
                    align=None, font_color="000000"):
        cell.value = value
        cell.font  = Font(bold=bold, color=font_color)
        if fill:
            cell.fill = fill
        if fmt:
            cell.number_format = fmt
        cell.alignment = align or self.CENTER

    # ------------------------------------------------------------------ #
    # MAIN GET
    # ------------------------------------------------------------------ #

    def get(self, request):

        category_id = (
            request.GET.get("category")
            or request.GET.get("category_id")
        )
        month_value = request.GET.get("month")
        year_value  = request.GET.get("year")

        year, selected_month, err = self._parse_year_month(year_value, month_value)
        if err:
            return Response(err, status=status.HTTP_400_BAD_REQUEST)

        months_to_loop = (
            [selected_month] if selected_month else range(1, 13)
        )

        if category_id:
            categories = Category.objects.filter(id=category_id)
            if not categories.exists():
                return Response(
                    {"success": False, "message": "Category not found"},
                    status=status.HTTP_404_NOT_FOUND,
                )
        else:
            categories = Category.objects.all()
            if not categories.exists():
                return Response(
                    {"success": False, "message": "No categories found"},
                    status=status.HTTP_404_NOT_FOUND,
                )

        cat_list  = list(categories)
        cat_names = [c.name for c in cat_list if c.name]

        title_prefix = (
            ", ".join(cat_names)
            if len(cat_names) <= 4
            else f"All Categories ({len(cat_names)})"
        )

        report_title = (
            f"{title_prefix} | Target Report - "
            f"{calendar.month_name[selected_month]} {year}"
            if selected_month
            else f"{title_prefix} | Target Report - {year}"
        )

        purchase_orders = PurchaseOrder.objects.filter(
            Q(po_date__year=year)
            | Q(po_date__isnull=True, created_at__year=year)
        ).order_by("-id")

        # ---------------------------------------------------------------- #
        # WORKBOOK
        # ---------------------------------------------------------------- #

        workbook = Workbook()
        ws       = workbook.active
        ws.title = "Category Target Report"

        last_col = len(self.HEADERS)

        # Row 1 — Title
        ws.merge_cells(f"A1:{get_column_letter(last_col)}1")
        ws["A1"].value     = report_title
        ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill      = self.RED_FILL
        ws["A1"].alignment = self.CENTER

        # Row 2 — spacer
        ws.row_dimensions[2].height = 6

        # Row 3 — Headers
        for col, header in enumerate(self.HEADERS, 1):
            self._style_cell(
                ws.cell(row=3, column=col), header,
                bold=True, fill=self.RED_FILL,
                font_color="FFFFFF"
            )

        # ---------------------------------------------------------------- #
        # DATA ROWS
        # ---------------------------------------------------------------- #

        row_num = 4
        sr_no   = 1

        all_grand_target  = Decimal("0.00")
        all_grand_product = Decimal("0.00")
        all_grand_order   = Decimal("0.00")
        all_grand_invoice = Decimal("0.00")

        rows_written = 0

        for category_obj in cat_list:

            products = list(Product.objects.filter(category_id=category_obj.id))
            if not products:
                continue

            cat_target  = Decimal("0.00")
            cat_product = Decimal("0.00")
            cat_order   = Decimal("0.00")
            cat_invoice = Decimal("0.00")

            for month in months_to_loop:

                month_name = calendar.month_abbr[month] + f"-{str(year)[-2:]}"

                month_pos = purchase_orders.filter(
                    Q(po_date__month=month)
                    | Q(po_date__isnull=True, created_at__month=month)
                )

                m_target, m_product, m_order, m_invoice = self._calc_month_data(
                    category_obj.id, products, month_pos, year, month
                )

                cat_target  += m_target
                cat_product += m_product
                cat_order   += m_order
                cat_invoice += m_invoice

                # ✅ Col 1: SR No
                ws.cell(row=row_num, column=1).value     = sr_no
                ws.cell(row=row_num, column=1).alignment = self.CENTER

                # ✅ Col 2: Month (was col 3 before)
                ws.cell(row=row_num, column=2).value     = month_name
                ws.cell(row=row_num, column=2).alignment = self.CENTER

                # ✅ Cols 3-6: numeric values
                for col, val in zip(
                    range(3, 7),
                    [m_target, m_product, m_order, m_invoice]
                ):
                    cell               = ws.cell(row=row_num, column=col)
                    cell.value         = float(val.quantize(Decimal("0.01")))
                    cell.number_format = "#,##0.00"
                    cell.alignment     = self.CENTER

                row_num      += 1
                sr_no        += 1
                rows_written += 1

            # ✅ Subtotal row
            ws.cell(row=row_num, column=2).value     = "Subtotal"
            ws.cell(row=row_num, column=2).alignment = self.CENTER
            ws.cell(row=row_num, column=2).font      = Font(bold=True)
            ws.cell(row=row_num, column=2).fill      = self.GRAY_FILL

            for col, val in zip(
                range(3, 7),
                [cat_target, cat_product, cat_order, cat_invoice]
            ):
                cell               = ws.cell(row=row_num, column=col)
                cell.value         = float(val.quantize(Decimal("0.01")))
                cell.number_format = "#,##0.00"
                cell.font          = Font(bold=True)
                cell.fill          = self.GRAY_FILL
                cell.alignment     = self.CENTER

            for col in [1, 2]:
                ws.cell(row=row_num, column=col).fill = self.GRAY_FILL

            row_num += 1  # subtotal
            row_num += 1  # blank separator

            all_grand_target  += cat_target
            all_grand_product += cat_product
            all_grand_order   += cat_order
            all_grand_invoice += cat_invoice

        if rows_written == 0:
            return Response(
                {"success": False, "message": "No products found for the selected category"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # ✅ Grand Total row (only when all categories)
        if not category_id:
            ws.cell(row=row_num, column=2).value     = "GRAND TOTAL"
            ws.cell(row=row_num, column=2).alignment = self.CENTER
            ws.cell(row=row_num, column=2).font      = Font(bold=True)
            ws.cell(row=row_num, column=2).fill      = self.YELLOW_FILL

            for col, val in zip(
                range(3, 7),
                [all_grand_target, all_grand_product, all_grand_order, all_grand_invoice]
            ):
                cell               = ws.cell(row=row_num, column=col)
                cell.value         = float(val.quantize(Decimal("0.01")))
                cell.number_format = "#,##0.00"
                cell.font          = Font(bold=True)
                cell.fill          = self.YELLOW_FILL
                cell.alignment     = self.CENTER

            for col in [1, 2]:
                c      = ws.cell(row=row_num, column=col)
                c.fill = self.YELLOW_FILL
                c.font = Font(bold=True)

        # Auto column widths
        for col in range(1, ws.max_column + 1):
            max_len    = 0
            col_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = max_len + 5

        # ---------------------------------------------------------------- #
        # RESPONSE
        # ---------------------------------------------------------------- #

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )

        if category_id:
            safe_name = (
                str(cat_list[0].name or "category")
                .replace(" ", "_")
                .replace("/", "_")
            )
            file_name = (
                f"{safe_name}_{year}_{selected_month}_report.xlsx"
                if selected_month
                else f"{safe_name}_target_report_{year}.xlsx"
            )
        else:
            file_name = (
                f"all_categories_{year}_{selected_month}_report.xlsx"
                if selected_month
                else f"all_categories_target_report_{year}.xlsx"
            )

        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
        workbook.save(response)
        return response
    
class DistributorTargetReportAPIView(APIView):

    def _to_decimal(self, value, default="0.00"):
        try:
            if value in [None, "", "null"]:
                return Decimal(default)
            return Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    def _safe_str(self, value, default=""):
        if value in [None, "", "null"]:
            return default
        return str(value)

    def _get_item_product_id(self, item):
        for key in ["product_id", "product", "product_data", "id"]:
            value = item.get(key)
            if value not in [None, "", "null"]:
                if isinstance(value, dict):
                    return str(value.get("id") or value.get("product_id") or "")
                return str(value)
        return None

    def _get_order_amount_without_gst(self, item):
        qty = self._to_decimal(item.get("quantity", 0))

        if item.get("total_price") not in [None, "", "null"]:
            return self._to_decimal(item.get("total_price"))

        if item.get("subtotal") not in [None, "", "null"]:
            return self._to_decimal(item.get("subtotal"))

        if item.get("taxable_amount") not in [None, "", "null"]:
            return self._to_decimal(item.get("taxable_amount"))

        if item.get("total_amount") not in [None, "", "null"]:
            return self._to_decimal(item.get("total_amount"))

        if item.get("total_distributor_price") not in [None, "", "null"] and self._to_decimal(item.get("total_distributor_price")) > 0:
            return self._to_decimal(item.get("total_distributor_price"))

        if item.get("unit_distributor_price") not in [None, "", "null"] and self._to_decimal(item.get("unit_distributor_price")) > 0:
            return self._to_decimal(item.get("unit_distributor_price")) * qty

        if item.get("unit_price") not in [None, "", "null"]:
            return self._to_decimal(item.get("unit_price")) * qty

        if item.get("price") not in [None, "", "null"]:
            return self._to_decimal(item.get("price")) * qty

        if item.get("mrp") not in [None, "", "null"]:
            return self._to_decimal(item.get("mrp")) * qty

        return Decimal("0.00")

    def _get_invoice_amount_with_gst(self, item):
        qty = self._to_decimal(item.get("quantity", 0))

        if item.get("final_amount") not in [None, "", "null"]:
            return self._to_decimal(item.get("final_amount"))

        if item.get("total_with_gst") not in [None, "", "null"]:
            return self._to_decimal(item.get("total_with_gst"))

        if item.get("amount_with_gst") not in [None, "", "null"]:
            return self._to_decimal(item.get("amount_with_gst"))

        if item.get("invoice_total") not in [None, "", "null"]:
            return self._to_decimal(item.get("invoice_total"))

        taxable = Decimal("0.00")
        gst = Decimal("0.00")

        if item.get("total_price") not in [None, "", "null"]:
            taxable = self._to_decimal(item.get("total_price"))
        elif item.get("subtotal") not in [None, "", "null"]:
            taxable = self._to_decimal(item.get("subtotal"))
        elif item.get("taxable_amount") not in [None, "", "null"]:
            taxable = self._to_decimal(item.get("taxable_amount"))
        elif item.get("total_amount") not in [None, "", "null"]:
            taxable = self._to_decimal(item.get("total_amount"))
        elif item.get("total_distributor_price") not in [None, "", "null"] and self._to_decimal(item.get("total_distributor_price")) > 0:
            taxable = self._to_decimal(item.get("total_distributor_price"))
        elif item.get("unit_distributor_price") not in [None, "", "null"] and self._to_decimal(item.get("unit_distributor_price")) > 0:
            taxable = self._to_decimal(item.get("unit_distributor_price")) * qty
        elif item.get("unit_price") not in [None, "", "null"]:
            taxable = self._to_decimal(item.get("unit_price")) * qty
        elif item.get("price") not in [None, "", "null"]:
            taxable = self._to_decimal(item.get("price")) * qty
        elif item.get("mrp") not in [None, "", "null"]:
            taxable = self._to_decimal(item.get("mrp")) * qty

        if item.get("gst_amount") not in [None, "", "null"]:
            gst = self._to_decimal(item.get("gst_amount"))
        elif item.get("gst") not in [None, "", "null"]:
            gst = self._to_decimal(item.get("gst"))
        else:
            gst = (
                self._to_decimal(item.get("cgst_amount", 0)) +
                self._to_decimal(item.get("sgst_amount", 0)) +
                self._to_decimal(item.get("igst_amount", 0))
            )

        return taxable + gst

    def get(self, request):
        distributor_id = request.GET.get("distributor_id")
        month_value = request.GET.get("month")

        if not distributor_id or not month_value:
            return Response({
                "success": False,
                "message": "distributor_id and month are required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            if len(month_value) == 7:
                parsed_date = datetime.strptime(month_value, "%Y-%m")
            elif len(month_value) == 10:
                parsed_date = datetime.strptime(month_value, "%Y-%m-%d")
            else:
                return Response({
                    "success": False,
                    "message": "Month format must be YYYY-MM or YYYY-MM-DD"
                }, status=status.HTTP_400_BAD_REQUEST)
        except ValueError:
            return Response({
                "success": False,
                "message": "Invalid month format. Use YYYY-MM or YYYY-MM-DD"
            }, status=status.HTTP_400_BAD_REQUEST)

        year = parsed_date.year
        month_num = parsed_date.month

        # Find target for this distributor
        distributor_target = DistributorInformation_Target.objects.filter(
            distributor_id=distributor_id,
            month__year=year,
            month__month=month_num
        ).first()

        if not distributor_target:
            return Response({
                "success": False,
                "message": "No target found for this distributor and month"
            }, status=status.HTTP_404_NOT_FOUND)

        from django.db.models import Q
        # Deep Search: Get POs for this distributor in the specified month
        purchase_orders = PurchaseOrder.objects.filter(
            Q(distributor_id=distributor_id),
            Q(po_date__year=year, po_date__month=month_num) |
            Q(po_date__isnull=True, created_at__year=year, created_at__month=month_num)
        ).order_by("-id")

        total_order = Decimal("0.00")
        total_invoice = Decimal("0.00")
        total_products = 0
        order_list = []

        for po in purchase_orders:

            serializer_data = PurchaseOrderSerializer(po).data

            po_total_order = self._to_decimal(
                serializer_data.get("distributor_price", 0)
            )

            po_total_invoice = Decimal("0.00")
            po_total_products = 0

            for item in serializer_data.get("product_items", []):

                qty = int(self._to_decimal(item.get("quantity", 0)))
                po_total_products += qty

                product_id = item.get("product_id")

                try:
                    product = Product.objects.get(id=product_id)

                    gst_rate = self._to_decimal(
                        getattr(product, "gst", 18)
                    ) / Decimal("100")

                except Product.DoesNotExist:
                    gst_rate = Decimal("0.18")

                item_total = self._to_decimal(
                    item.get("total_distributor_price", 0)
                )

                po_total_invoice += item_total * (
                    Decimal("1") + gst_rate
                )

            total_order += po_total_order
            total_invoice += po_total_invoice
            total_products += po_total_products

            order_list.append({
                "po_number": po.po_number,
                "po_date": (po.po_date or po.created_at).strftime("%Y-%m-%d"),
                "total_products": po_total_products,
                "total_order_without_gst": str(
                    po_total_order.quantize(Decimal("0.01"))
                ),
                "total_invoice_with_gst": str(
                    po_total_invoice.quantize(Decimal("0.01"))
                )
            })

        return Response({
        "success": True,
        "distributor_id": distributor_id,
        "distributor_name": self._safe_str(distributor_target.distributor),
        "month": distributor_target.month.strftime("%Y-%m") if distributor_target.month else "",
        "target": self._safe_str(distributor_target.target, "0"),
        "summary": {
            "total_products": total_products,
            "total_order_without_gst": str(
                total_order.quantize(Decimal("0.01"))
            ),
            "total_invoice_with_gst": str(
                total_invoice.quantize(Decimal("0.01"))
            )
        },
        "orders": order_list
    }, status=status.HTTP_200_OK)
        


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# class DistributorTargetReportExcelAPIView(APIView):

#     # --------------------------------
#     # SAFE DECIMAL
#     # --------------------------------
#     def _to_decimal(self, value, default="0.00"):

#         try:

#             if value in [None, "", "null"]:
#                 return Decimal(default)

#             return Decimal(str(value))

#         except (InvalidOperation, TypeError, ValueError):

#             return Decimal(default)

#     # --------------------------------
#     # FLOAT FORMAT
#     # --------------------------------
#     def _to_float(self, value):

#         return float(
#             self._to_decimal(value).quantize(
#                 Decimal("0.01")
#             )
#         )

#     # --------------------------------
#     # MAIN API
#     # --------------------------------
#     def get(self, request):

#         distributor_id = request.GET.get("distributor_id")
#         month_value = request.GET.get("month")
#         year_value = request.GET.get("year")

#         if not distributor_id:

#             return Response(
#                 {
#                     "success": False,
#                     "message": "distributor_id is required"
#                 },
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         # --------------------------------
#         # YEAR + MONTH FILTER
#         # --------------------------------
#         current_date = datetime.now()

#         # DEFAULT YEAR
#         year = current_date.year
#         selected_month = None

#         # --------------------------------
#         # YEAR FILTER
#         # --------------------------------
#         if year_value:

#             try:

#                 year = int(year_value)

#             except ValueError:

#                 return Response(
#                     {
#                         "success": False,
#                         "message": "Invalid year format"
#                     },
#                     status=status.HTTP_400_BAD_REQUEST
#                 )

#         # --------------------------------
#         # MONTH FILTER
#         # --------------------------------
#         if month_value:

#             try:

#                 # MM
#                 if len(month_value) == 2:

#                     selected_month = int(month_value)

#                 # YYYY-MM
#                 elif len(month_value) == 7:

#                     parsed = datetime.strptime(
#                         month_value,
#                         "%Y-%m"
#                     )

#                     year = parsed.year
#                     selected_month = parsed.month

#                 # YYYY-MM-DD
#                 elif len(month_value) == 10:

#                     parsed = datetime.strptime(
#                         month_value,
#                         "%Y-%m-%d"
#                     )

#                     year = parsed.year
#                     selected_month = parsed.month

#                 else:

#                     return Response(
#                         {
#                             "success": False,
#                             "message": (
#                                 "Use MM or YYYY-MM "
#                                 "or YYYY-MM-DD"
#                             )
#                         },
#                         status=status.HTTP_400_BAD_REQUEST
#                     )

#             except ValueError:

#                 return Response(
#                     {
#                         "success": False,
#                         "message": "Invalid month format"
#                     },
#                     status=status.HTTP_400_BAD_REQUEST
#                 )

#         # --------------------------------
#         # DISTRIBUTOR
#         # --------------------------------
#         distributor_obj = DistributorInformation.objects.filter(
#             id=distributor_id
#         ).first()

#         if not distributor_obj:

#             return Response(
#                 {
#                     "success": False,
#                     "message": "Distributor not found"
#                 },
#                 status=status.HTTP_404_NOT_FOUND
#             )

#         distributor_name = (
#             getattr(
#                 distributor_obj,
#                 "distributor_name",
#                 None
#             )
#             or getattr(
#                 distributor_obj,
#                 "company_name",
#                 None
#             )
#             or str(distributor_obj)
#         )

#         # --------------------------------
#         # PURCHASE ORDERS
#         # --------------------------------
#         purchase_orders = PurchaseOrder.objects.filter(
#             distributor_id=distributor_id
#         ).filter(
#             Q(po_date__year=year) |
#             Q(
#                 po_date__isnull=True,
#                 created_at__year=year
#             )
#         )

#         # --------------------------------
#         # EXCEL
#         # --------------------------------
#         wb = Workbook()

#         ws = wb.active

#         ws.title = "Distributor Target Report"

#         # --------------------------------
#         # TITLE
#         # --------------------------------
#         ws.merge_cells("A1:F1")

#         title = ws["A1"]

#         if selected_month:

#             report_title = (
#                 f"{distributor_name} Target Report - "
#                 f"{calendar.month_name[selected_month]} "
#                 f"{year}"
#             )

#         else:

#             report_title = (
#                 f"{distributor_name} Target Report - "
#                 f"{year}"
#             )

#         title.value = report_title

#         title.font = Font(
#             bold=True,
#             size=16,
#             color="FFFFFF"
#         )

#         title.fill = PatternFill(
#             fill_type="solid",
#             start_color="C00000",
#             end_color="C00000"
#         )

#         title.alignment = Alignment(
#             horizontal="center",
#             vertical="center"
#         )

#         # --------------------------------
#         # HEADERS
#         # --------------------------------
#         headers = [
#             "SR No",
#             "Month",
#             "Target",
#             "Total Product",
#             "Order (Excl GST)",
#             "Invoice (Incl GST)"
#         ]

#         header_fill = PatternFill(
#             fill_type="solid",
#             start_color="C00000",
#             end_color="C00000"
#         )

#         for col, h in enumerate(headers, 1):

#             cell = ws.cell(
#                 row=3,
#                 column=col
#             )

#             cell.value = h

#             cell.font = Font(
#                 bold=True,
#                 color="FFFFFF"
#             )

#             cell.fill = header_fill

#             cell.alignment = Alignment(
#                 horizontal="center",
#                 vertical="center"
#             )

#         # --------------------------------
#         # DATA
#         # --------------------------------
#         row = 4
#         sr_no = 1

#         total_target = Decimal("0.00")
#         total_product = Decimal("0.00")
#         total_order = Decimal("0.00")
#         total_invoice = Decimal("0.00")

#         months = (
#             [selected_month]
#             if selected_month
#             else range(1, 13)
#         )

#         for m in months:

#             # --------------------------------
#             # MONTH NAME
#             # --------------------------------
#             month_name = (
#                 calendar.month_abbr[m]
#                 + f"-{str(year)[-2:]}"
#             )

#             # --------------------------------
#             # TARGET
#             # --------------------------------
#             target_obj = (
#                 DistributorInformation_Target.objects.filter(
#                     distributor_id=distributor_id,
#                     month__year=year,
#                     month__month=m
#                 ).first()
#             )

#             month_target = self._to_decimal(
#                 target_obj.target if target_obj else 0
#             )

#             # --------------------------------
#             # MONTH PURCHASE ORDERS
#             # --------------------------------
#             month_orders = purchase_orders.filter(
#                 Q(po_date__month=m) |
#                 Q(
#                     po_date__isnull=True,
#                     created_at__month=m
#                 )
#             )

#             month_product_total = Decimal("0.00")
#             month_order_total = Decimal("0.00")
#             month_invoice_total = Decimal("0.00")

#             # --------------------------------
#             # ORDER LOOP
#             # --------------------------------
#             for po in month_orders:

#                 for item in (po.product_items or []):

#                     if isinstance(item, dict):

#                         qty = self._to_decimal(
#                             item.get("quantity", 0)
#                         )

#                         month_product_total += qty

#                         base = self._to_decimal(
#                             item.get("unit_price")
#                             or item.get("price")
#                             or item.get("mrp")
#                             or 0
#                         )

#                         order_amt = qty * base

#                         gst = (
#                             order_amt *
#                             Decimal("0.18")
#                         )

#                         invoice_amt = order_amt + gst

#                         month_order_total += order_amt
#                         month_invoice_total += invoice_amt

#             # --------------------------------
#             # GRAND TOTALS
#             # --------------------------------
#             total_target += month_target
#             total_product += month_product_total
#             total_order += month_order_total
#             total_invoice += month_invoice_total

#             # --------------------------------
#             # WRITE ROW
#             # --------------------------------
#             ws.cell(
#                 row=row,
#                 column=1
#             ).value = sr_no

#             ws.cell(
#                 row=row,
#                 column=2
#             ).value = month_name

#             # TARGET
#             cell = ws.cell(
#                 row=row,
#                 column=3
#             )

#             cell.value = self._to_float(
#                 month_target
#             )

#             cell.number_format = "0.00"

#             # TOTAL PRODUCT
#             cell = ws.cell(
#                 row=row,
#                 column=4
#             )

#             cell.value = self._to_float(
#                 month_product_total
#             )

#             cell.number_format = "0.00"

#             # ORDER
#             cell = ws.cell(
#                 row=row,
#                 column=5
#             )

#             cell.value = self._to_float(
#                 month_order_total
#             )

#             cell.number_format = "0.00"

#             # INVOICE
#             cell = ws.cell(
#                 row=row,
#                 column=6
#             )

#             cell.value = self._to_float(
#                 month_invoice_total
#             )

#             cell.number_format = "0.00"

#             # ALIGNMENT
#             for c in range(1, 7):

#                 ws.cell(
#                     row=row,
#                     column=c
#                 ).alignment = Alignment(
#                     horizontal="center",
#                     vertical="center"
#                 )

#             row += 1
#             sr_no += 1

#         # --------------------------------
#         # TOTAL ROW
#         # --------------------------------
#         total_fill = PatternFill(
#             fill_type="solid",
#             start_color="FFF2CC",
#             end_color="FFF2CC"
#         )

#         total_label = ws.cell(
#             row=row,
#             column=2
#         )

#         total_label.value = "TOTAL"

#         total_label.font = Font(
#             bold=True
#         )

#         total_label.fill = total_fill

#         total_label.alignment = Alignment(
#             horizontal="center",
#             vertical="center"
#         )

#         # TARGET TOTAL
   
#         cell = ws.cell(
#             row=row,
#             column=3
#         )
                
#         cell.value = self._to_float(
#             total_target
#         )

#         cell.number_format = "0.00"
#         cell.font = Font(bold=True)
#         cell.fill = total_fill
#         cell.alignment = Alignment(
#             horizontal="center",
#             vertical="center"
#         )

#         # TOTAL PRODUCT
#         cell = ws.cell(
#             row=row,
#             column=4
#         )

#         cell.value = self._to_float(
#             total_product
#         )

#         cell.number_format = "0.00"
#         cell.font = Font(bold=True)
#         cell.fill = total_fill
#         cell.alignment = Alignment(
#             horizontal="center",
#             vertical="center"
#         )

#         # ORDER TOTAL
#         cell = ws.cell(
#             row=row,
#             column=5
#         )

#         cell.value = self._to_float(
#             total_order
#         )

#         cell.number_format = "0.00"
#         cell.font = Font(bold=True)
#         cell.fill = total_fill
#         cell.alignment = Alignment(
#             horizontal="center",
#             vertical="center"
#         )

#         # INVOICE TOTAL
#         cell = ws.cell(
#             row=row,
#             column=6
#         )

#         cell.value = self._to_float(
#             total_invoice
#         )

#         cell.number_format = "0.00"
#         cell.font = Font(bold=True)
#         cell.fill = total_fill
#         cell.alignment = Alignment(
#             horizontal="center",
#             vertical="center"
#         )
#             # --------------------------------
#         # AUTO WIDTH
#         # --------------------------------
#         for col in range(
#             1,
#             ws.max_column + 1
#         ):

#             max_length = 0

#             col_letter = get_column_letter(col)

#             for r in range(
#                 1,
#                 ws.max_row + 1
#             ):

#                 val = ws.cell(
#                     row=r,
#                     column=col
#                 ).value

#                 if val:

#                     max_length = max(
#                         max_length,
#                         len(str(val))
#                     )

#             ws.column_dimensions[
#                 col_letter
#             ].width = max_length + 5

#         # --------------------------------
#         # RESPONSE
#         # --------------------------------
#         response = HttpResponse(
#             content_type=(
#                 "application/vnd.openxmlformats-"
#                 "officedocument.spreadsheetml.sheet"
#             )
#         )

#         safe_name = (
#             str(distributor_name)
#             .replace(" ", "_")
#             .replace("/", "_")
#         )

#         if selected_month:

#             file_name = (
#                 f"{safe_name}_"
#                 f"{year}_{selected_month}_report.xlsx"
#             )

#         else:

#             file_name = (
#                 f"{safe_name}_target_report_{year}.xlsx"
#             )

#         response[
#             "Content-Disposition"
#         ] = (
#             f'attachment; filename="{file_name}"'
#         )

#         wb.save(response)

#         return response 

import calendar
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.db.models import Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView



class DistributorTargetReportExcelAPIView(APIView):

    # ------------------------------------------------------------------ #
    # HELPERS
    # ------------------------------------------------------------------ #

    def _to_decimal(self, value, default="0.00"):
        try:
            if value in [None, "", "null"]:
                return Decimal(default)
            return Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    def _to_float(self, value):
        return float(
            self._to_decimal(value).quantize(Decimal("0.01"))
        )

    def _get_base_price(self, item, product_map):
        """
        ✅ Same priority order as the API view:
        1. PO item-level prices first
        2. Fall back to product.mrp from product_map
        """
        for key in ["unit_distributor_price", "mrp", "unit_price", "price"]:
            val = item.get(key)
            if val not in [None, "", "null", 0, "0", "0.00"]:
                try:
                    d = Decimal(str(val))
                    if d > 0:
                        return d
                except Exception:
                    pass

        # Fall back to product.mrp via product_map
        pid = None
        for key in ["product_id", "product", "id"]:
            v = item.get(key)
            if v not in [None, "", "null"]:
                pid = str(v.get("id") if isinstance(v, dict) else v)
                break

        if pid and pid in product_map:
            mrp = getattr(product_map[pid], "mrp", None)
            if mrp not in [None, "", "null"]:
                try:
                    d = Decimal(str(mrp))
                    if d > 0:
                        return d
                except Exception:
                    pass

        return Decimal("0.00")

    # ------------------------------------------------------------------ #
    # SHARED STYLES
    # ------------------------------------------------------------------ #

    RED_FILL    = PatternFill(fill_type="solid", start_color="C00000", end_color="C00000")
    YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
    CENTER      = Alignment(horizontal="center", vertical="center")

    # ------------------------------------------------------------------ #
    # AUTO COLUMN WIDTH
    # ------------------------------------------------------------------ #

    def _auto_width(self, ws):
        for col in range(1, ws.max_column + 1):
            max_length = max(
                (
                    len(str(ws.cell(row=r, column=col).value or ""))
                    for r in range(1, ws.max_row + 1)
                ),
                default=0,
            )
            ws.column_dimensions[get_column_letter(col)].width = max_length + 5

    # ------------------------------------------------------------------ #
    # WRITE ONE DISTRIBUTOR BLOCK
    # ------------------------------------------------------------------ #

    def _write_distributor_block(
        self, ws, start_row,
        distributor_id, distributor_name,
        year, selected_month,
        purchase_orders_qs,
    ):
        row = start_row

        # ── Title ──────────────────────────────────────────────────────
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=6,
        )
        title_cell = ws.cell(row=row, column=1)
        title_cell.value = (
            f"{distributor_name} Target Report - "
            f"{calendar.month_name[selected_month]} {year}"
            if selected_month
            else f"{distributor_name} Target Report - {year}"
        )
        title_cell.font      = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill      = self.RED_FILL
        title_cell.alignment = self.CENTER
        row += 1

        # ── Column headers ─────────────────────────────────────────────
        headers = [
            "SR No",
            "Month",
            "Target",
            "Total Product",
            "Order (Excl GST)",
            "Invoice (Incl GST)",
        ]
        for col, h in enumerate(headers, 1):
            cell           = ws.cell(row=row, column=col)
            cell.value     = h
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = self.RED_FILL
            cell.alignment = self.CENTER
        row += 1

        # ── Pre-build product map for price fallback ───────────────────
        # Get all product IDs referenced in this distributor's POs
        all_product_ids = set()
        for po in purchase_orders_qs:
            for item in (po.product_items or []):
                if not isinstance(item, dict):
                    continue
                for key in ["product_id", "product", "id"]:
                    v = item.get(key)
                    if v not in [None, "", "null"]:
                        pid = str(v.get("id") if isinstance(v, dict) else v)
                        all_product_ids.add(pid)
                        break

        product_map = {
            str(p.id): p
            for p in Product.objects.filter(id__in=all_product_ids)
        } if all_product_ids else {}

        # ── Data rows ──────────────────────────────────────────────────
        sr_no         = 1
        grand_target  = Decimal("0.00")
        grand_product = Decimal("0.00")
        grand_order   = Decimal("0.00")
        grand_invoice = Decimal("0.00")

        months = [selected_month] if selected_month else range(1, 13)

        for m in months:

            month_label = calendar.month_abbr[m] + f"-{str(year)[-2:]}"

            # Target
            target_obj = DistributorInformation_Target.objects.filter(
                distributor_id=distributor_id,
                month__year=year,
                month__month=m,
            ).first()
            month_target = self._to_decimal(
                target_obj.target if target_obj else 0
            )

            # Orders
            month_orders = purchase_orders_qs.filter(
                Q(po_date__month=m) |
                Q(po_date__isnull=True, created_at__month=m)
            )

            month_product = Decimal("0.00")
            month_order   = Decimal("0.00")
            month_invoice = Decimal("0.00")

            for po in month_orders:
                for item in (po.product_items or []):
                    if not isinstance(item, dict):
                        continue

                    qty = self._to_decimal(
                        item.get("quantity", item.get("qty", 0))
                    )
                    if qty <= 0:
                        continue

                    # ✅ Use same price priority as API view
                    base = self._get_base_price(item, product_map)

                    order_amt      = qty * base
                    invoice_amt    = order_amt * Decimal("1.18")
                    month_product += qty
                    month_order   += order_amt
                    month_invoice += invoice_amt

            grand_target  += month_target
            grand_product += month_product
            grand_order   += month_order
            grand_invoice += month_invoice

            # Write cells
            ws.cell(row=row, column=1).value = sr_no
            ws.cell(row=row, column=2).value = month_label

            for col, val in [
                (3, month_target),
                (4, month_product),
                (5, month_order),
                (6, month_invoice),
            ]:
                cell               = ws.cell(row=row, column=col)
                cell.value         = self._to_float(val)
                cell.number_format = "#,##0.00"

            for col in range(1, 7):
                ws.cell(row=row, column=col).alignment = self.CENTER

            row   += 1
            sr_no += 1

        # ── Total row ──────────────────────────────────────────────────
        ws.cell(row=row, column=1).fill = self.YELLOW_FILL

        total_label           = ws.cell(row=row, column=2)
        total_label.value     = "TOTAL"
        total_label.font      = Font(bold=True)
        total_label.fill      = self.YELLOW_FILL
        total_label.alignment = self.CENTER

        for col, val in [
            (3, grand_target),
            (4, grand_product),
            (5, grand_order),
            (6, grand_invoice),
        ]:
            cell               = ws.cell(row=row, column=col)
            cell.value         = self._to_float(val)
            cell.number_format = "#,##0.00"
            cell.font          = Font(bold=True)
            cell.fill          = self.YELLOW_FILL
            cell.alignment     = self.CENTER

        return row + 2

    # ------------------------------------------------------------------ #
    # MAIN GET
    # ------------------------------------------------------------------ #

    def get(self, request):

        distributor_id = request.GET.get("distributor_id")
        month_value    = request.GET.get("month")
        year_value     = request.GET.get("year")
        sheet_mode     = request.GET.get("mode", "multi")

        # ── Year ───────────────────────────────────────────────────────
        year           = datetime.now().year
        selected_month = None

        if year_value:
            try:
                year = int(year_value)
            except ValueError:
                return Response(
                    {"success": False, "message": "Invalid year format"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # ── Month ──────────────────────────────────────────────────────
        if month_value:
            try:
                if len(month_value) == 2:
                    selected_month = int(month_value)
                elif len(month_value) == 7:
                    parsed         = datetime.strptime(month_value, "%Y-%m")
                    year           = parsed.year
                    selected_month = parsed.month
                elif len(month_value) == 10:
                    parsed         = datetime.strptime(month_value, "%Y-%m-%d")
                    year           = parsed.year
                    selected_month = parsed.month
                else:
                    return Response(
                        {"success": False, "message": "Use MM, YYYY-MM, or YYYY-MM-DD"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            except ValueError:
                return Response(
                    {"success": False, "message": "Invalid month format"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # ── Distributor queryset ────────────────────────────────────────
        if distributor_id:
            distributors = DistributorInformation.objects.filter(id=distributor_id)
            if not distributors.exists():
                return Response(
                    {"success": False, "message": "Distributor not found"},
                    status=status.HTTP_404_NOT_FOUND,
                )
        else:
            distributors = DistributorInformation.objects.all().order_by("id")

        if not distributors.exists():
            return Response(
                {"success": False, "message": "No distributors found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # ── Build workbook ─────────────────────────────────────────────
        wb = Workbook()
        wb.remove(wb.active)

        if sheet_mode == "single":
            ws          = wb.create_sheet(title="All Distributors Report")
            current_row = 1

            for dist in distributors:
                dist_name = (
                    getattr(dist, "distributor_name", None)
                    or getattr(dist, "company_name", None)
                    or str(dist)
                )
                pos = PurchaseOrder.objects.filter(
                    distributor_id=dist.id
                ).filter(
                    Q(po_date__year=year) |
                    Q(po_date__isnull=True, created_at__year=year)
                )
                current_row = self._write_distributor_block(
                    ws, current_row,
                    dist.id, dist_name,
                    year, selected_month, pos,
                )

            self._auto_width(ws)

        else:
            for dist in distributors:
                dist_name = (
                    getattr(dist, "distributor_name", None)
                    or getattr(dist, "company_name", None)
                    or str(dist)
                )
                safe_title = (
                    str(dist_name)[:31]
                    .replace("/", "_").replace("\\", "_")
                    .replace("*", "_").replace("?",  "_")
                    .replace("[", "_").replace("]",  "_")
                )
                ws = wb.create_sheet(title=safe_title)

                pos = PurchaseOrder.objects.filter(
                    distributor_id=dist.id
                ).filter(
                    Q(po_date__year=year) |
                    Q(po_date__isnull=True, created_at__year=year)
                )
                self._write_distributor_block(
                    ws, 1,
                    dist.id, dist_name,
                    year, selected_month, pos,
                )
                self._auto_width(ws)

        # ── HTTP response ───────────────────────────────────────────────
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )

        if distributor_id and distributors.count() == 1:
            dist      = distributors.first()
            dist_name = (
                getattr(dist, "distributor_name", None)
                or getattr(dist, "company_name", None)
                or str(dist)
            )
            safe_name = str(dist_name).replace(" ", "_").replace("/", "_")
            file_name = (
                f"{safe_name}_{year}_{selected_month:02d}_report.xlsx"
                if selected_month
                else f"{safe_name}_target_report_{year}.xlsx"
            )
        else:
            file_name = f"all_distributors_target_report_{year}.xlsx"

        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
        wb.save(response)
        return response
 
class ProductTargetReportAPIView(APIView):

    def _to_decimal(self, value, default="0.00"):
        try:
            if value in [None, "", "null", "0", "0.0", 0, 0.0]:
                return Decimal(default)
            return Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    def _safe_str(self, value, default=""):
        if value in [None, "", "null"]:
            return default
        return str(value)

    def _get_item_product_id(self, item):
        for key in ["product_id", "product", "product_data", "id"]:
            value = item.get(key)
            if value not in [None, "", "null"]:
                if isinstance(value, dict):
                    return str(value.get("id") or value.get("product_id") or "")
                return str(value)
        return None

    def _get_order_amount_without_gst(self, item):
        qty = self._to_decimal(item.get("quantity", 0))
        if item.get("total_price") not in [None, "", "null"]:
            return self._to_decimal(item.get("total_price"))
        if item.get("subtotal") not in [None, "", "null"]:
            return self._to_decimal(item.get("subtotal"))
        if item.get("taxable_amount") not in [None, "", "null"]:
            return self._to_decimal(item.get("taxable_amount"))
        if item.get("total_amount") not in [None, "", "null"]:
            return self._to_decimal(item.get("total_amount"))
        if item.get("total_distributor_price") not in [None, "", "null"] and self._to_decimal(item.get("total_distributor_price")) > 0:
            return self._to_decimal(item.get("total_distributor_price"))
        if item.get("unit_distributor_price") not in [None, "", "null"] and self._to_decimal(item.get("unit_distributor_price")) > 0:
            return self._to_decimal(item.get("unit_distributor_price")) * qty
        if item.get("unit_price") not in [None, "", "null"]:
            return self._to_decimal(item.get("unit_price")) * qty
        if item.get("price") not in [None, "", "null"]:
            return self._to_decimal(item.get("price")) * qty
        if item.get("mrp") not in [None, "", "null"]:
            return self._to_decimal(item.get("mrp")) * qty
        return Decimal("0.00")

    def _get_invoice_amount_with_gst(self, item):

        qty = self._to_decimal(item.get("quantity", 0))

        unit_price = self._to_decimal(
            item.get("unit_distributor_price", item.get("price", 0))
        )

        product_id = item.get("product_id")

        gst_percentage = Decimal("0")

        if product_id:
            try:
                product = Product.objects.get(id=product_id)

                if product.GST:
                    gst_percentage = self._to_decimal(
                        str(product.GST).replace("%", "")
                    )

            except Product.DoesNotExist:
                pass

        total_distributor_price = qty * unit_price

        gst_amount = (
            total_distributor_price * gst_percentage / Decimal("100")
        )

        total_with_gst = total_distributor_price + gst_amount

        return total_with_gst.quantize(
            Decimal("0.01")
        )

    def get(self, request):
        product_id = request.GET.get("product_id")
        month_value = request.GET.get("month")
        if not product_id or not month_value:
            return Response({
                "success": False,
                "message": "product_id and month are required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            if len(month_value) == 7:
                parsed_date = datetime.strptime(month_value, "%Y-%m")
            elif len(month_value) == 10:
                parsed_date = datetime.strptime(month_value, "%Y-%m-%d")
            else:
                return Response({
                    "success": False,
                    "message": "Month format must be YYYY-MM or YYYY-MM-DD"
                }, status=status.HTTP_400_BAD_REQUEST)
        except ValueError:
            return Response({
                "success": False,
                "message": "Invalid month format. Use YYYY-MM or YYYY-MM-DD"
            }, status=status.HTTP_400_BAD_REQUEST)

        year = parsed_date.year
        month_num = parsed_date.month

        from .models import Product_Target
        product_target = Product_Target.objects.filter(
            product_id=product_id,
            month__year=year,
            month__month=month_num
        ).first()

        if not product_target:
            return Response({
                "success": False,
                "message": "No target found for this product and month"
            }, status=status.HTTP_404_NOT_FOUND)

        from django.db.models import Q
        purchase_orders = PurchaseOrder.objects.filter(
            Q(po_date__year=year, po_date__month=month_num) |
            Q(po_date__isnull=True, created_at__year=year, created_at__month=month_num)
        ).order_by("-id")

        try:
            from .models import Product
            prod_obj = Product.objects.get(pk=product_id)
            self.prod_obj = prod_obj
        except:
            self.prod_obj = None

        total_order = Decimal("0.00")
        total_products = Decimal("0")
        total_invoice = Decimal("0.00")
        order_list = []

        for po in purchase_orders:
            items = po.product_items or []

            po_total_order = Decimal("0.00")
            po_total_invoice = Decimal("0.00")
            po_total_products = Decimal("0")
            item_match_count = 0

            if isinstance(items, list):
                for item in items:
                    if not isinstance(item, dict):
                        continue

                    item_product_id = self._get_item_product_id(item)

                    if item_product_id and item_product_id == str(product_id).strip():
                        qty = self._to_decimal(item.get("quantity", item.get("qty", 0)))

                        without_gst = self._get_order_amount_without_gst(item)
                        with_gst = self._get_invoice_amount_with_gst(item)

                        po_total_products += qty
                        po_total_order += without_gst
                        po_total_invoice += with_gst

                        item_match_count += 1

            if item_match_count > 0:
                total_order += po_total_order
                total_invoice += po_total_invoice
                total_products += po_total_products

                order_list.append({
                    "po_number": po.po_number,
                    "po_date": po.po_date.strftime("%Y-%m-%d") if po.po_date else "",
                    "total_products": str(po_total_products.quantize(Decimal("1"))),
                    "total_order_without_gst": str(po_total_order.quantize(Decimal("0.01"))),
                    "total_invoice_with_gst": str(po_total_invoice.quantize(Decimal("0.01")))
                })

        return Response({
            "success": True,
            "product_id": product_id,
            "product_name": self.prod_obj.product_name if self.prod_obj else "",
            "month": product_target.month.strftime("%Y-%m") if product_target.month else "",
            "target": self._safe_str(product_target.target, "0"),
            "summary": {
                "total_order_without_gst": str(total_order.quantize(Decimal("0.01"))),
                "total_products": str(total_products.quantize(Decimal("1"))),
                "total_invoice_with_gst": str(total_invoice.quantize(Decimal("0.01")))
            },
            "orders": order_list
        }, status=status.HTTP_200_OK)

         


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class ProductTargetReportExcelAPIView(APIView):

    # --------------------------------
    # SAFE DECIMAL
    # --------------------------------
    def _to_decimal(self, value, default="0.00"):
        try:
            if value in [None, "", "null"]:
                return Decimal(default)
            return Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    # --------------------------------
    # FLOAT FORMAT
    # --------------------------------
    def _to_float(self, value):
        return float(
            self._to_decimal(value).quantize(Decimal("0.01"))
        )

    # --------------------------------
    # ✅ BASE PRICE — same priority as API view
    # --------------------------------
    def _get_base_price(self, item, product_obj):
        for key in ["unit_distributor_price", "mrp", "unit_price", "price"]:
            val = item.get(key)
            if val not in [None, "", "null", 0, "0", "0.00"]:
                try:
                    d = Decimal(str(val))
                    if d > 0:
                        return d
                except Exception:
                    pass

        # Fall back to product.mrp
        mrp = getattr(product_obj, "mrp", None)
        if mrp not in [None, "", "null"]:
            try:
                d = Decimal(str(mrp))
                if d > 0:
                    return d
            except Exception:
                pass

        return Decimal("0.00")

    # --------------------------------
    # MAIN API
    # --------------------------------
    def get(self, request):

        product_id  = request.GET.get("product_id")
        month_value = request.GET.get("month")
        year_value  = request.GET.get("year")

        if not product_id:
            return Response(
                {"success": False, "message": "product_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # --------------------------------
        # YEAR
        # --------------------------------
        year           = datetime.now().year
        selected_month = None

        if year_value:
            try:
                year = int(year_value)
            except ValueError:
                return Response(
                    {"success": False, "message": "Invalid year format"},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # --------------------------------
        # MONTH
        # --------------------------------
        if month_value:
            try:
                if len(month_value) == 2:
                    selected_month = int(month_value)
                elif len(month_value) == 7:
                    parsed         = datetime.strptime(month_value, "%Y-%m")
                    year           = parsed.year
                    selected_month = parsed.month
                elif len(month_value) == 10:
                    parsed         = datetime.strptime(month_value, "%Y-%m-%d")
                    year           = parsed.year
                    selected_month = parsed.month
                else:
                    return Response(
                        {"success": False, "message": "Use MM or YYYY-MM or YYYY-MM-DD"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except ValueError:
                return Response(
                    {"success": False, "message": "Invalid month format"},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # --------------------------------
        # PRODUCT
        # --------------------------------
        product_obj = Product.objects.filter(id=product_id).first()

        if not product_obj:
            return Response(
                {"success": False, "message": "Product not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        product_name = getattr(product_obj, "product_name", None) or str(product_obj)

        # --------------------------------
        # PURCHASE ORDERS
        # --------------------------------
        purchase_orders = PurchaseOrder.objects.filter(
            Q(po_date__year=year) |
            Q(po_date__isnull=True, created_at__year=year)
        )

        # --------------------------------
        # EXCEL
        # --------------------------------
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Target Report"

        RED_FILL    = PatternFill(fill_type="solid", start_color="C00000", end_color="C00000")
        YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
        CENTER      = Alignment(horizontal="center", vertical="center")

        # --------------------------------
        # TITLE
        # --------------------------------
        ws.merge_cells("A1:F1")
        title       = ws["A1"]
        title.value = (
            f"{product_name} Target Report - "
            f"{calendar.month_name[selected_month]} {year}"
            if selected_month
            else f"{product_name} Target Report - {year}"
        )
        title.font      = Font(bold=True, size=16, color="FFFFFF")
        title.fill      = RED_FILL
        title.alignment = CENTER

        # Row 2 — spacer
        ws.row_dimensions[2].height = 6

        # --------------------------------
        # HEADERS
        # --------------------------------
        headers = [
            "SR No",
            "Month",
            "Target",
            "Total Product",
            "Order (Excl GST)",
            "Invoice (Incl GST)",
        ]

        for col, h in enumerate(headers, 1):
            cell           = ws.cell(row=3, column=col)
            cell.value     = h
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = RED_FILL
            cell.alignment = CENTER

        # --------------------------------
        # DATA
        # --------------------------------
        row           = 4
        sr_no         = 1
        total_target  = Decimal("0.00")
        total_product = Decimal("0.00")
        total_order   = Decimal("0.00")
        total_invoice = Decimal("0.00")

        months = [selected_month] if selected_month else range(1, 13)

        for m in months:

            month_name = calendar.month_abbr[m] + f"-{str(year)[-2:]}"

            # TARGET
            target_obj   = Product_Target.objects.filter(
                product_id=product_id,
                month__year=year,
                month__month=m
            ).first()
            month_target = self._to_decimal(
                target_obj.target if target_obj else 0
            )

            # ORDERS
            month_orders        = purchase_orders.filter(
                Q(po_date__month=m) |
                Q(po_date__isnull=True, created_at__month=m)
            )
            month_product_total = Decimal("0.00")
            month_order_total   = Decimal("0.00")
            month_invoice_total = Decimal("0.00")

            for po in month_orders:
                for item in (po.product_items or []):
                    if not isinstance(item, dict):
                        continue

                    if str(item.get("product_id")) != str(product_id):
                        continue

                    # ✅ qty with fallback + guard
                    qty = self._to_decimal(
                        item.get("quantity", item.get("qty", 0))
                    )
                    if qty <= 0:
                        continue

                    # ✅ price with correct priority
                    base      = self._get_base_price(item, product_obj)
                    order_amt = qty * base
                    inv_amt   = order_amt * Decimal("1.18")

                    month_product_total += qty
                    month_order_total   += order_amt
                    month_invoice_total += inv_amt

            total_target  += month_target
            total_product += month_product_total
            total_order   += month_order_total
            total_invoice += month_invoice_total

            # WRITE ROW
            ws.cell(row=row, column=1).value = sr_no
            ws.cell(row=row, column=2).value = month_name

            for col, val in [
                (3, month_target),
                (4, month_product_total),
                (5, month_order_total),
                (6, month_invoice_total),
            ]:
                cell               = ws.cell(row=row, column=col)
                cell.value         = self._to_float(val)
                cell.number_format = "#,##0.00"

            for c in range(1, 7):
                ws.cell(row=row, column=c).alignment = CENTER

            row   += 1
            sr_no += 1

        # --------------------------------
        # TOTAL ROW
        # --------------------------------
        ws.cell(row=row, column=1).fill = YELLOW_FILL

        total_label           = ws.cell(row=row, column=2)
        total_label.value     = "TOTAL"
        total_label.font      = Font(bold=True)
        total_label.fill      = YELLOW_FILL
        total_label.alignment = CENTER

        for col, val in [
            (3, total_target),
            (4, total_product),
            (5, total_order),
            (6, total_invoice),
        ]:
            cell               = ws.cell(row=row, column=col)
            cell.value         = self._to_float(val)
            cell.number_format = "#,##0.00"
            cell.font          = Font(bold=True)
            cell.fill          = YELLOW_FILL
            cell.alignment     = CENTER

        # --------------------------------
        # AUTO WIDTH
        # --------------------------------
        for col in range(1, ws.max_column + 1):
            max_length = max(
                (
                    len(str(ws.cell(row=r, column=col).value or ""))
                    for r in range(1, ws.max_row + 1)
                ),
                default=0,
            )
            ws.column_dimensions[get_column_letter(col)].width = max_length + 5

        # --------------------------------
        # RESPONSE
        # --------------------------------
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )

        safe_name = str(product_name).replace(" ", "_").replace("/", "_")
        file_name = (
            f"{safe_name}_{year}_{selected_month}_report.xlsx"
            if selected_month
            else f"{safe_name}_target_report_{year}.xlsx"
        )

        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
        wb.save(response)
        return response

class RegionTargetReportAPIView(APIView):

    def _to_decimal(self, value, default="0.00"):
        try:
            if value in [None, "", "null"]:
                return Decimal(default)
            return Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    def _safe_str(self, value, default=""):
        if value in [None, "", "null"]:
            return default
        return str(value)

    def _get_order_amount_without_gst(self, item):
        qty = self._to_decimal(item.get("quantity", 0))

        if item.get("total_price") not in [None, "", "null"]:
            return self._to_decimal(item.get("total_price"))
        if item.get("subtotal") not in [None, "", "null"]:
            return self._to_decimal(item.get("subtotal"))
        if item.get("taxable_amount") not in [None, "", "null"]:
            return self._to_decimal(item.get("taxable_amount"))
        if item.get("total_amount") not in [None, "", "null"]:
            return self._to_decimal(item.get("total_amount"))
        if item.get("total_distributor_price") not in [None, "", "null"] and self._to_decimal(item.get("total_distributor_price")) > 0:
            return self._to_decimal(item.get("total_distributor_price"))
        if item.get("unit_distributor_price") not in [None, "", "null"] and self._to_decimal(item.get("unit_distributor_price")) > 0:
            return self._to_decimal(item.get("unit_distributor_price")) * qty
        if item.get("unit_price") not in [None, "", "null"]:
            return self._to_decimal(item.get("unit_price")) * qty
        if item.get("price") not in [None, "", "null"]:
            return self._to_decimal(item.get("price")) * qty
        if item.get("mrp") not in [None, "", "null"]:
            return self._to_decimal(item.get("mrp")) * qty

        return Decimal("0.00")

    def _get_invoice_amount_with_gst(self, item):
        qty = self._to_decimal(
            item.get("quantity", item.get("qty", 0))
        )

        unit_price = self._to_decimal(
            item.get(
                "unit_distributor_price",
                item.get("price", 0)
            )
        )

        product_id = item.get("product_id")

        gst_percentage = Decimal("0.00")

        if product_id:
            try:
                product = Product.objects.get(id=product_id)

                if product.GST:
                    gst_percentage = self._to_decimal(
                        str(product.GST).replace("%", "")
                    )

            except Product.DoesNotExist:
                pass

        taxable = qty * unit_price

        gst_amount = (
            taxable * gst_percentage / Decimal("100")
        )

        return (taxable + gst_amount).quantize(
            Decimal("0.01")
        )

    def get(self, request):
        region_id = request.GET.get("region_id")
        region_name = request.GET.get("region")
        month_value = request.GET.get("month")

        if not month_value:
            return Response({
                "success": False,
                "message": "month is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            if len(month_value) == 7:
                parsed_date = datetime.strptime(month_value, "%Y-%m")
            elif len(month_value) == 10:
                parsed_date = datetime.strptime(month_value, "%Y-%m-%d")
            else:
                return Response({
                    "success": False,
                    "message": "Month format must be YYYY-MM or YYYY-MM-DD"
                }, status=status.HTTP_400_BAD_REQUEST)
        except ValueError:
            return Response({
                "success": False,
                "message": "Invalid month format. Use YYYY-MM or YYYY-MM-DD"
            }, status=status.HTTP_400_BAD_REQUEST)

        year = parsed_date.year
        month_num = parsed_date.month

        region_obj = None

        if region_id:
            region_obj = Region.objects.filter(id=region_id).first()
        elif region_name:
            region_obj = Region.objects.filter(name__iexact=region_name).first()

        if not region_obj:
            return Response({
                "success": False,
                "message": "Region not found"
            }, status=status.HTTP_404_NOT_FOUND)

        region_id = region_obj.id

        region_target = Region_Target.objects.filter(
            region_id=region_id,
            month__year=year,
            month__month=month_num
        ).first()

        if not region_target:
            return Response({
                "success": False,
                "message": "No target found for this region and month"
            }, status=status.HTTP_404_NOT_FOUND)

        purchase_orders = PurchaseOrder.objects.filter(
            distributor_id__region_id=region_obj.id
        ).filter(
            Q(
                po_date__year=year,
                po_date__month=month_num
            ) |
            Q(
                po_date__isnull=True,
                created_at__year=year,
                created_at__month=month_num
            )
        ).order_by("-id")

        total_order = Decimal("0.00")
        total_invoice = Decimal("0.00")   # ✅ fixed: removed duplicate line
        total_products = 0
        order_list = []

        print("Region ID:", region_obj.id)
        print("Region Name:", region_obj.name)

        print(
            "PO Count:",
            purchase_orders.count()
        )

        for po in purchase_orders:
            items = po.product_items or []

            po_total_order = Decimal("0.00")
            po_total_invoice = Decimal("0.00")
            po_total_products = 0

            if isinstance(items, list):
                for item in items:
                    if not isinstance(item, dict):
                        continue

                    qty = int(self._to_decimal(item.get("quantity", item.get("qty", 0))))

                    po_total_order += self._get_order_amount_without_gst(item)
                    po_total_invoice += self._get_invoice_amount_with_gst(item)
                    po_total_products += qty

            total_order += po_total_order
            total_invoice += po_total_invoice
            total_products += po_total_products

            order_list.append({
                "po_number": po.po_number,
                "po_date": po.po_date.strftime("%Y-%m-%d") if po.po_date else None,
                "distributor_id": po.distributor_id.id if po.distributor_id else None,
                "total_products": po_total_products,
                "total_order_without_gst": str(po_total_order.quantize(Decimal("0.01"))),
                "total_invoice_with_gst": str(po_total_invoice.quantize(Decimal("0.01")))
            })

        return Response({
            "success": True,
            "region_id": region_obj.id,
            "region_name": region_obj.name,
            "month": region_target.month.strftime("%Y-%m") if region_target.month else "",
            "target": self._safe_str(region_target.target, "0"),
            "summary": {
                "total_products": total_products,
                "total_order_without_gst": str(total_order.quantize(Decimal("0.01"))),
                "total_invoice_with_gst": str(total_invoice.quantize(Decimal("0.01")))
            },
            "orders": order_list
        }, status=status.HTTP_200_OK)

        
from decimal import Decimal, InvalidOperation
from datetime import datetime
import calendar

from django.http import HttpResponse
from django.db.models import Q

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import (
    PurchaseOrder,
    Region,
    Region_Target,
)

class RegionTargetReportExcelAPIView(APIView):

    GST_RATE = Decimal("0.18")

    # ------------------------------------------------------------------ #
    # HELPERS
    # ------------------------------------------------------------------ #

    def _to_decimal(self, value, default="0.00"):
        try:
            if value in [None, "", "null"]:
                return Decimal(default)
            return Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    def _to_float(self, value):
        return float(
            self._to_decimal(value).quantize(Decimal("0.01"))
        )

    def _get_base_price(self, item, product_obj):
        """
        ✅ Same priority as all other Excel/API views:
        1. PO item-level prices first
        2. Fall back to product.mrp
        """
        for key in ["unit_distributor_price", "mrp", "unit_price", "price"]:
            val = item.get(key)
            if val not in [None, "", "null", 0, "0", "0.00"]:
                try:
                    d = Decimal(str(val))
                    if d > 0:
                        return d
                except Exception:
                    pass

        # Fall back to product model
        if product_obj:
            for attr in ["unit_distributor_price", "mrp"]:
                val = getattr(product_obj, attr, None)
                if val not in [None, "", "null"]:
                    try:
                        d = Decimal(str(val))
                        if d > 0:
                            return d
                    except Exception:
                        pass

        return Decimal("0.00")

    def _get_item_product_id(self, item):
        for key in ["product_id", "product", "product_data", "id"]:
            value = item.get(key)
            if value not in [None, "", "null"]:
                if isinstance(value, dict):
                    return str(value.get("id") or value.get("product_id") or "")
                return str(value)
        return None

    # ------------------------------------------------------------------ #
    # MAIN GET
    # ------------------------------------------------------------------ #

    def get(self, request):

        region_id   = request.GET.get("region_id")
        month_value = request.GET.get("month")

        if not region_id:
            return Response(
                {"success": False, "message": "region_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ── Date parse ─────────────────────────────────────────────────
        parsed_date = None
        if month_value:
            try:
                parsed_date = datetime.strptime(month_value, "%Y-%m")
            except ValueError:
                try:
                    parsed_date = datetime.strptime(month_value, "%Y-%m-%d")
                except ValueError:
                    return Response(
                        {"success": False, "message": "Use YYYY-MM or YYYY-MM-DD"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

        year   = parsed_date.year  if parsed_date else datetime.now().year
        months = [parsed_date.month] if parsed_date else range(1, 13)

        # ── Region ─────────────────────────────────────────────────────
        region_obj = Region.objects.filter(id=region_id).first()
        if not region_obj:
            return Response(
                {"success": False, "message": "Region not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        # ── Distributors & POs ─────────────────────────────────────────
        distributors = DistributorInformation.objects.filter(
            Q(region=region_obj) |
            Q(sales_region=str(region_obj.id)) |
            Q(sales_region__iexact=region_obj.name)
        )
        distributor_ids = list(distributors.values_list("id", flat=True))

        purchase_orders = PurchaseOrder.objects.filter(
            distributor_id__in=distributor_ids
        ).filter(
            Q(po_date__year=year) |
            Q(po_date__isnull=True, created_at__year=year)
        )

        # ── Pre-fetch all products referenced in POs ───────────────────
        all_product_ids = set()
        for po in purchase_orders:
            for item in (po.product_items or []):
                if not isinstance(item, dict):
                    continue
                pid = self._get_item_product_id(item)
                if pid:
                    all_product_ids.add(pid)

        product_map = {
            str(p.id): p
            for p in Product.objects.filter(id__in=all_product_ids)
        } if all_product_ids else {}

        # ---------------------------------------------------------------- #
        # EXCEL SETUP
        # ---------------------------------------------------------------- #

        wb = Workbook()
        ws = wb.active
        ws.title = "Region Target Report"

        RED_FILL    = PatternFill(fill_type="solid", start_color="C00000", end_color="C00000")
        YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
        CENTER      = Alignment(horizontal="center", vertical="center")

        # Title
        ws.merge_cells("A1:F1")
        title_cell           = ws["A1"]
        title_cell.value     = (
            f"{region_obj.name} Target Report - "
            f"{calendar.month_name[months[0]]} {year}"
            if len(months) == 1
            else f"{region_obj.name} Target Report - {year}"
        )
        title_cell.font      = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill      = RED_FILL
        title_cell.alignment = CENTER

        # Spacer
        ws.row_dimensions[2].height = 6

        # Headers
        headers = [
            "SR No",
            "Month",
            "Target",
            "Total Products",
            "Order (Excl GST)",
            "Invoice (Incl GST)",
        ]
        for col, h in enumerate(headers, 1):
            cell           = ws.cell(row=3, column=col)
            cell.value     = h
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = RED_FILL
            cell.alignment = CENTER

        # Column widths
        for col, width in {1: 10, 2: 18, 3: 20, 4: 20, 5: 25, 6: 25}.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[3].height = 30

        # ---------------------------------------------------------------- #
        # DATA ROWS
        # ---------------------------------------------------------------- #

        row                  = 4
        sr_no                = 1
        total_target         = Decimal("0.00")
        total_order          = Decimal("0.00")
        total_invoice        = Decimal("0.00")
        grand_total_products = Decimal("0.00")

        for m in months:

            # Target
            target_obj = Region_Target.objects.filter(
                region_id=region_id,
                month__year=year,
                month__month=m
            ).first()
            target = self._to_decimal(
                target_obj.target if target_obj else 0
            )

            month_order          = Decimal("0.00")
            month_total_products = Decimal("0.00")

            month_orders = purchase_orders.filter(
                Q(po_date__month=m) |
                Q(po_date__isnull=True, created_at__month=m)
            )

            for po in month_orders:
                for item in (po.product_items or []):
                    if not isinstance(item, dict):
                        continue

                    pid = self._get_item_product_id(item)
                    if not pid:
                        continue

                    # ✅ qty with fallback + guard
                    qty = self._to_decimal(
                        item.get("quantity", item.get("qty", 0))
                    )
                    if qty <= 0:
                        continue

                    # ✅ price with correct priority
                    product_obj = product_map.get(pid)
                    base_price  = self._get_base_price(item, product_obj)

                    month_order          += qty * base_price
                    month_total_products += qty

            month_invoice = month_order * (Decimal("1") + self.GST_RATE)

            total_target         += target
            total_order          += month_order
            total_invoice        += month_invoice
            grand_total_products += month_total_products

            # Write row
            ws.cell(row=row, column=1).value = sr_no
            ws.cell(row=row, column=2).value = (
                f"{calendar.month_abbr[m]}-{str(year)[-2:]}"
            )

            for col, val in [
                (3, target),
                (4, month_total_products),
                (5, month_order),
                (6, month_invoice),
            ]:
                cell               = ws.cell(row=row, column=col)
                cell.value         = self._to_float(val)
                cell.number_format = "#,##0.00"
                cell.alignment     = CENTER

            for c in range(1, 3):
                ws.cell(row=row, column=c).alignment = CENTER

            ws.row_dimensions[row].height = 25
            row   += 1
            sr_no += 1

        # ---------------------------------------------------------------- #
        # TOTAL ROW
        # ---------------------------------------------------------------- #

        ws.cell(row=row, column=1).fill = YELLOW_FILL

        total_label           = ws.cell(row=row, column=2)
        total_label.value     = "TOTAL"
        total_label.font      = Font(bold=True)
        total_label.fill      = YELLOW_FILL
        total_label.alignment = CENTER

        for col, val in [
            (3, total_target),
            (4, grand_total_products),
            (5, total_order),
            (6, total_invoice),
        ]:
            cell               = ws.cell(row=row, column=col)
            cell.value         = self._to_float(val)
            cell.number_format = "#,##0.00"
            cell.font          = Font(bold=True)
            cell.fill          = YELLOW_FILL
            cell.alignment     = CENTER

        ws.row_dimensions[row].height = 28

        # ---------------------------------------------------------------- #
        # RESPONSE
        # ---------------------------------------------------------------- #

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )

        safe_name = (
            str(region_obj.name)
            .replace(" ", "_")
            .replace("/", "_")
        )
        file_name = (
            f"{safe_name}_{year}_{months[0]}_region_report.xlsx"
            if len(months) == 1
            else f"{safe_name}_region_target_report_{year}.xlsx"
        )

        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
        wb.save(response)
        return response
 
    
class WarrantyCardPDFView(APIView):
    """
    GET /warranty-card/<serial_id>/pdf/
    Generate and stream a warranty card PDF for a single InventorySerial.
    """

    def get(self, request, serial_id):
        try:
            serial_obj = InventorySerial.objects.select_related('product_id').get(id=serial_id)
        except InventorySerial.DoesNotExist:
            return Response(
                {"success": False, "message": "Inventory serial not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        product = serial_obj.product_id

        # Card dimensions (landscape label ~10" x 3.5")
        card_width = 10 * inch
        card_height = 3.5 * inch
        page_w = card_width + 0.5 * inch
        page_h = card_height + 0.5 * inch

        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=(page_w, page_h))

        x = 0.25 * inch
        y = 0.25 * inch

        _draw_warranty_card(c, x, y, card_width, card_height, product, serial_obj)

        c.showPage()
        c.save()
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'inline; filename="warranty_card_{serial_obj.serial_number}.pdf"'
        )
        return response

class WarrantyCardBulkPDFView(APIView):
    """
    GET /warranty-card/bulk-pdf/?po_number=<po_number>
    Generate warranty cards for all inventory serials associated with a Purchase Order.
    One card per page.
    """

    def get(self, request):
        po_number = request.query_params.get("po_number")
        product_id = request.query_params.get("product_id")

        if po_number:
            try:
                po = PurchaseOrder.objects.get(po_number=po_number)
            except PurchaseOrder.DoesNotExist:
                return Response(
                    {"success": False, "message": "Purchase Order not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Fetch serials reserved for this PO via DistributerOrders
            orders = DistributerOrders.objects.filter(purchase_order_id=po).select_related(
                'inventory_serial_id',
                'inventory_serial_id__product_id'
            )

            if not orders.exists():
                return Response(
                    {"success": False, "message": f"No reserved serials found for PO {po_number}"},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Prepare list of (product, serial_obj) tuples
            items_to_print = []
            for order in orders:
                if order.inventory_serial_id:
                    items_to_print.append((order.inventory_serial_id.product_id, order.inventory_serial_id))

            filename = f"warranty_cards_po_{po_number}.pdf"

        elif product_id:
            try:
                product = Product.objects.get(id=product_id)
            except Product.DoesNotExist:
                return Response(
                    {"success": False, "message": "Product not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            serials = InventorySerial.objects.filter(product_id=product).order_by("id")
            if not serials.exists():
                return Response(
                    {"success": False, "message": "No inventory serials found for this product"},
                    status=status.HTTP_404_NOT_FOUND
                )

            items_to_print = [(product, s) for s in serials]
            filename = f"warranty_cards_product_{product_id}.pdf"

        else:
            return Response(
                {"success": False, "message": "po_number or product_id query parameter is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Card dimensions
        card_width = 10 * inch
        card_height = 3.5 * inch
        page_w = card_width + 0.5 * inch
        page_h = card_height + 0.5 * inch

        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=(page_w, page_h))

        x = 0.25 * inch
        y = 0.25 * inch

        for product_obj, serial_obj in items_to_print:
            _draw_warranty_card(c, x, y, card_width, card_height, product_obj, serial_obj)
            c.showPage()

        c.save()
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response


class GalleryAPIView(APIView):

    # ---------- GET ----------
    def get(self, request, pk=None):
        if pk:
            gallery = get_object_or_404(Gallery, pk=pk)
            serializer = GallerySerializer(gallery, context={'request': request})
            return Response({
                "success": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        name = request.GET.get('name')

        galleries = Gallery.objects.all()

        # 🔍 Filter by name
        if name:
            galleries = galleries.filter(name__icontains=name)

        # 🔽 Optional: order by sequence
        galleries = galleries.order_by('gallery_sequence')

        serializer = GallerySerializer(galleries, many=True, context={'request': request})

        return Response({
            "success": True,
            "message":"Gallery list fetched successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)


    # ---------- POST ----------
    def post(self, request):
        serializer = GallerySerializer(data=request.data, context={'request': request})

        if serializer.is_valid():
            gallery = serializer.save()
            return Response({
                "success": True,
                "message": "Gallery created successfully",
                "data": {
                    "id": gallery.id,
                    "name": gallery.name,
                    "image": gallery.image.url if gallery.image else None,
                    "gallery_sequence": gallery.gallery_sequence
                }
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": 'Invalid data',
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- PATCH ----------
    def patch(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "pk is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        gallery = get_object_or_404(Gallery, pk=pk)
        serializer = GallerySerializer(gallery, data=request.data, partial=True, context={'request': request})

        if serializer.is_valid():
            serializer.save()
            return Response({
                "success": True,
                "message": "Gallery updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # ---------- DELETE ----------
    def delete(self, request, pk=None):
        if not pk:
            return Response({
                "success": False,
                "message": "Id is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        gallery = get_object_or_404(Gallery, pk=pk)
        gallery.delete()

        return Response({
            "success": True,
            "message": "Gallery deleted successfully"
        }, status=status.HTTP_200_OK)


class DashboardAPIView(APIView):
    def get(self, request):

        role_name = request.GET.get("role")
        role_not = request.GET.get("role_not")  # 👈 new param

        employees = Employee.objects.all()

        # 🔥 INCLUDE filter
        if role_name:
            employees = employees.filter(role_id__name__iexact=role_name)

        # 🔥 EXCLUDE filter (HR remove)
        if role_not:
            employees = employees.exclude(role_id__name__iexact=role_not)

        data = {
            "department": Department.objects.count(),
            "roles": Roles.objects.count(),
            "office_branches": Office_branch.objects.count(),

            "employee": employees.count(),

            "users": User.objects.count(),

            "personal_details": Employee_personal_details.objects.filter(
                employee_id__in=employees
            ).count(),

            "employee_salary": EmployeeSalary.objects.filter(
                employee_id__in=employees
            ).count(),

            "employee_documents": EmployeeDocuments.objects.filter(
                employee_id__in=employees
            ).count(),

            "leave_balance": LeaveBalance.objects.filter(
                employee_id__in=employees
            ).count(),

            "leave_request": LeaveRequest.objects.filter(
                employee_id__in=employees
            ).count(),

            "employee_attendance": Expenses_login.objects.filter(
                employee_id__in=employees
            ).count(),

            "salary_payout": Salary_payment.objects.filter(
                employee_id__in=employees
            ).count(),

            "leads": Lead.objects.count(),
            "visits": Visit.objects.count(),
            "lead_followups": LeadFollowUp.objects.count(),
            "travel_plan": TravelPlan.objects.count(),
            "expense": Expense.objects.count(),
            "holiday": Holiday.objects.count(),
        }

        return Response({
            "success": True,
            "filter": {
                "role": role_name,
                "role_not": role_not
            },
            "data": data
        })
        
        
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

class ReplacementRequestAPI(APIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get(self, request, pk=None):

        if pk:
            try:
                replacement = ReplacementRequest.objects.get(id=pk)

            except ReplacementRequest.DoesNotExist:
                return Response(
                    {
                        "success": False,
                        "message": "Replacement request not found"
                    },
                    status=status.HTTP_404_NOT_FOUND
                )

            serializer = ReplacementRequestSerializer(replacement)

            return Response(
                {
                    "success": True,
                    "message": "Replacement request fetched successfully",
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )

        replacements = ReplacementRequest.objects.all().order_by("-id")

        # distributor filter
        distributor_id = request.query_params.get("distributor_id")

        if distributor_id:
            replacements = replacements.filter(
                Q(
                    old_serial__distributororders__distributor_id_id=distributor_id
                ) |
                Q(
                    new_distributor_id=distributor_id
                )
            ).distinct()

        # status filter
        replacement_status = request.query_params.get("status")

        if replacement_status:
            replacements = replacements.filter(
                status=replacement_status
            )

        # old serial filter
        old_serial = request.query_params.get("old_serial")

        if old_serial:
            replacements = replacements.filter(
                old_serial__serial_number__icontains=old_serial
            )

        # new serial filter
        new_serial = request.query_params.get("new_serial")

        if new_serial:
            replacements = replacements.filter(
                new_serial__serial_number__icontains=new_serial
            )

        serializer = ReplacementRequestSerializer(
            replacements,
            many=True
        )

        return Response(
            {
                "success": True,
                "message": "Replacement request list fetched successfully",
                "count": replacements.count(),
                "data": serializer.data
            },
            status=status.HTTP_200_OK
        )

    def post(self, request):

        serializer = ReplacementRequestSerializer(
            data=request.data,
            context={"request": request}
        )

        if not serializer.is_valid():
            return Response({
                "success": False,
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        replacement = serializer.save()

        try:
            from django.core.mail import send_mail

            # ===========================
            # SERIAL NUMBER
            # ===========================
            serial_number = replacement.old_serial.serial_number

            # ===========================
            # WARRANTY GET (FIXED)
            # ===========================
            warranty = Warranty.objects.filter(
                serial_id__serial_number=serial_number
            ).first()

            owner_email = None
            distributor_email = None
            customer_name = "Customer"
            vehicle_number = "-"
            product_name = "-"

            if warranty:
                owner_email = warranty.owner_email
                distributor_email = warranty.email
                customer_name = warranty.owner_name or "Customer"
                vehicle_number = warranty.car_registration_number or "-"

                # ✅ FIXED PRODUCT ACCESS
                if warranty.product_id:
                    product_name = warranty.product_id.product_name

            subject = "Warranty Claim Request Submitted Successfully"

            message = f"""
    Dear {customer_name},

    Your warranty claim request has been successfully submitted and is currently under review by our support team.

    Claim Details:

    Claim ID: {replacement.id}
    Product: {product_name}
    Serial Number: {serial_number}
    Vehicle Number: {vehicle_number}
    Claim Date: {replacement.requested_at.strftime('%d-%m-%Y %H:%M')}

    Our team will verify the submitted details and update the claim status shortly.

    If additional information is required, our team will contact you.

    Thank you for choosing HOGO AUTOFILMS India Pvt. Ltd.

    Regards,
    HOGO AUTOFILMS India Pvt. Ltd.
    Customer Support Team
    """

            recipient_list = []

            if owner_email:
                recipient_list.append(owner_email)

            if distributor_email:
                recipient_list.append(distributor_email)

            if recipient_list:
                print("EMAIL LIST:", recipient_list)

                send_mail(
                    subject,
                    message,
                    settings.EMAIL_HOST_USER,
                    recipient_list,
                    fail_silently=False,
                )

        except Exception as e:
            print("EMAIL ERROR:", str(e))

            return Response({
                "success": True,
                "message": "Replacement created but email failed",
                "error": str(e),
                "data": ReplacementRequestSerializer(replacement).data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "success": True,
            "message": "Replacement request created & emails sent successfully",
            "data": ReplacementRequestSerializer(replacement).data
        }, status=status.HTTP_201_CREATED)

    def patch(self, request, pk=None):

        if not pk:
            return Response({
                "success": False,
                "message": "Replacement request id is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            replacement = ReplacementRequest.objects.get(id=pk)
        except ReplacementRequest.DoesNotExist:
            return Response({
                "success": False,
                "message": "Replacement request not found"
            }, status=status.HTTP_404_NOT_FOUND)

        old_status = replacement.status

        serializer = ReplacementRequestSerializer(
            replacement,
            data=request.data,
            partial=True
        )

        if not serializer.is_valid():
            return Response({
                "success": False,
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        replacement = serializer.save()

        # =========================================
        # WARRANTY FETCH (SERIAL BASED)
        # =========================================
        try:
            from django.core.mail import send_mail

            serial_number = replacement.old_serial.serial_number

            warranty = Warranty.objects.filter(
                serial_id__serial_number=serial_number
            ).first()

            owner_email = None
            distributor_email = None
            customer_name = "Customer"
            vehicle_number = "-"
            product_name = "-"
            rejection_reason = "-"

            if warranty:
                owner_email = warranty.owner_email
                distributor_email = warranty.email
                customer_name = warranty.owner_name or "Customer"
                vehicle_number = warranty.car_registration_number or "-"

                if warranty.product_id:
                    product_name = warranty.product_id.product_name

            recipient_list = []

            if owner_email:
                recipient_list.append(owner_email)

            if distributor_email:
                recipient_list.append(distributor_email)

            # =========================================
            # ✅ APPROVED EMAIL
            # =========================================
            if str(replacement.status).upper() == "APPROVED" and str(old_status).upper() != "APPROVED":

                subject = "Warranty Claim Approved Successfully"

                message = f"""
    Dear {customer_name},

    We are pleased to inform you that your warranty claim request has been approved successfully.

    Approved Claim Details:

    Claim ID: {replacement.id}
    Product: {product_name}
    Serial Number: {serial_number}
    Vehicle Number: {vehicle_number}
    Approval Date: {replacement.approved_at.strftime('%d-%m-%Y %H:%M') if replacement.approved_at else '-'}

    Our support team will coordinate the next process regarding replacement/service assistance.

    Thank you for trusting HOGO AUTOFILMS India Pvt. Ltd.

    Regards,
    Warranty Support Team
    """

                if recipient_list:
                    send_mail(
                        subject,
                        message,
                        settings.EMAIL_HOST_USER,
                        recipient_list,
                        fail_silently=False,
                    )

            # =========================================
            # ❌ REJECTED EMAIL
            # =========================================
            elif str(replacement.status).upper() == "REJECTED" and str(old_status).upper() != "REJECTED":

                rejection_reason = getattr(replacement, "rejection_reason", "-")

                subject = "Update Regarding Your Warranty Claim Request"

                message = f"""
    Dear {customer_name},

    We regret to inform you that your warranty claim request could not be approved after verification.

    Claim Details:

    Claim ID: {replacement.id}
    Product: {product_name}
    Serial Number: {serial_number}
    Vehicle Number: {vehicle_number}

    Reason for Rejection:
    {rejection_reason}

    If you believe additional clarification or documents are required, please contact our support team or your distributor.

    Thank you for your understanding.

    Regards,
    HOGO AUTOFILMS India Pvt. Ltd.
    Warranty Support Team
    """
                if recipient_list:
                    send_mail(
                        subject,
                        message,
                        settings.EMAIL_HOST_USER,
                        recipient_list,
                        fail_silently=False,
                    )

        except Exception as e:
            print("PATCH EMAIL ERROR:", str(e))

            return Response({
                "success": True,
                "message": "Replacement updated but email failed",
                "error": str(e),
                "data": ReplacementRequestSerializer(replacement).data
            }, status=status.HTTP_200_OK)

        return Response({
            "success": True,
            "message": "Replacement updated successfully",
            "data": ReplacementRequestSerializer(replacement).data
        }, status=status.HTTP_200_OK)

    def delete(self, request, pk=None):

        if not pk:
            return Response({
                "success": False,
                "message": "Replacement request id is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            replacement = ReplacementRequest.objects.get(id=pk)
        except ReplacementRequest.DoesNotExist:
            return Response({
                "success": False,
                "message": "Replacement request not found"
            }, status=status.HTTP_404_NOT_FOUND)

        replacement.delete()

        return Response({
            "success": True,
            "message": "Replacement request deleted successfully"
        }, status=status.HTTP_200_OK)


class ResetPasswordAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        serializer = ResetPasswordSerializer(data=request.data)

        if serializer.is_valid():

            user = request.user

            old_password = serializer.validated_data["old_password"]
            new_password = serializer.validated_data["new_password"]

            # Check old password
            if not user.check_password(old_password):
                return Response(
                    {
                        "success": False,
                        "message": "Old password is incorrect"
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Set new password
            user.set_password(new_password)
            user.save()

            return Response(
                {
                    "success": True,
                    "message": "Password reset successfully"
                },
                status=status.HTTP_200_OK
            )
        return Response(
            {
                "success": False,
                "errors": serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )
        
class EmployeeForgotPasswordAPIView(APIView):

    # -----------------------------------
    # SEND OTP
    # -----------------------------------
    def post(self, request):

        email = request.data.get("email")
        employee_id = request.data.get("employee")

        if not email or not employee_id:
            return Response(
                {"status": "error", "message": "Email and Employee ID required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            employee_obj = Employee.objects.get(
                email=email,
                id=employee_id
            )

        except Employee.DoesNotExist:
            return Response(
                {"status": "error", "message": "Employee not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        otp = str(randint(100000, 999999))

        # delete old otp
        EmployeeForgotPassword.objects.filter(
            employee=employee_obj
        ).delete()

        # create new otp
        EmployeeForgotPassword.objects.create(
            employee=employee_obj,
            email=email,
            otp=otp
        )

        # send mail
        send_mail(
            subject="Employee Password Reset OTP",
            message=f"Your OTP is: {otp}",
            from_email="yourgmail@gmail.com",
            recipient_list=[email],
            fail_silently=False
        )

        return Response(
            {"status": "success", "message": "OTP sent successfully"},
            status=status.HTTP_200_OK
        )

    # -----------------------------------
    # VERIFY OTP + RESET PASSWORD
    # -----------------------------------
    def patch(self, request):

        email = request.data.get("email")
        employee_id = request.data.get("employee")
        otp = request.data.get("otp")
        new_password = request.data.get("new_password")
        confirm_password = request.data.get("confirm_password")

        # validate fields
        if not all([email, employee_id, otp, new_password, confirm_password]):
            return Response(
                {"status": "error", "message": "All fields are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # password match check
        if new_password != confirm_password:
            return Response(
                {"status": "error", "message": "Passwords do not match"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # get employee
        try:
            employee_obj = Employee.objects.get(
                email=email,
                id=employee_id
            )

        except Employee.DoesNotExist:
            return Response(
                {"status": "error", "message": "Employee not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        # verify otp
        try:
            reset_obj = EmployeeForgotPassword.objects.get(
                employee=employee_obj,
                email=email,
                otp=otp,
                is_used=False
            )

        except EmployeeForgotPassword.DoesNotExist:
            return Response(
                {"status": "error", "message": "Invalid OTP"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # check otp expiry
        if reset_obj.is_expired():
            return Response(
                {"status": "error", "message": "OTP expired"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # mark otp used
        reset_obj.is_used = True
        reset_obj.save()

        # store plain password
        employee_obj.password = new_password
        employee_obj.save()

        return Response(
            {
                "status": "success",
                "message": "Password reset successfully"
            },
            status=status.HTTP_200_OK
        )


class DistributorForgotPasswordAPIView(APIView):

    # -----------------------------------
    # SEND OTP
    # -----------------------------------
    def post(self, request):

        email = request.data.get("email")
        distributor_id = request.data.get("distributor")

        if not email or not distributor_id:
            return Response(
                {"status": "error", "message": "Email and Distributor ID required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            distributor_obj = DistributorInformation.objects.get(
            email_id=email,
            id=distributor_id
        )
        except DistributorInformation.DoesNotExist:
            return Response(
                {"status": "error", "message": "Distributor not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        otp = str(randint(100000, 999999))

        DistriutorForgotPassword.objects.filter(
            distributor=distributor_obj
        ).delete()

        DistriutorForgotPassword.objects.create(
            distributor=distributor_obj,
            email=email,
            otp=otp
        )

        send_mail(
            subject="Distributor Password Reset OTP",
            message=f"Your OTP is: {otp}",
            from_email="yourgmail@gmail.com",
            recipient_list=[email],
            fail_silently=False
        )

        return Response(
            {"status": "success", "message": "OTP sent successfully"},
            status=status.HTTP_200_OK
        )

    # -----------------------------------
    # VERIFY OTP + RESET PASSWORD
    # -----------------------------------
    def patch(self, request):

        email = request.data.get("email")
        distributor_id = request.data.get("distributor")  # ✅ FIXED NAME
        otp = request.data.get("otp")
        new_password = request.data.get("new_password")
        confirm_password = request.data.get("confirm_password")

        if not all([email, distributor_id, otp, new_password, confirm_password]):
            return Response(
                {"status": "error", "message": "All fields are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if new_password != confirm_password:
            return Response(
                {"status": "error", "message": "Passwords do not match"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ----------------------------
        # GET DISTRIBUTOR (FIXED)
        # ----------------------------
        try:
            distributor_obj = DistributorInformation.objects.get(
                email_id=email,
                id=distributor_id   # ✅ correct field
            )
        except DistributorInformation.DoesNotExist:
            return Response(
                {"status": "error", "message": "Distributor not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        # ----------------------------
        # VERIFY OTP (FIXED MODEL NAME)
        # ----------------------------
        try:
            reset_obj = DistriutorForgotPassword.objects.get(
                distributor=distributor_obj,
                email=email,
                otp=otp,
                is_used=False
            )
        except DistriutorForgotPassword.DoesNotExist:
            return Response(
                {"status": "error", "message": "Invalid OTP"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # expiry check
        if reset_obj.is_expired():
            return Response(
                {"status": "error", "message": "OTP expired"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # mark OTP used
        reset_obj.is_used = True
        reset_obj.save()

        # update password
        distributor_obj.password = make_password(new_password)
        distributor_obj.save()

        return Response(
            {"status": "success", "message": "Password reset successfully"},
            status=status.HTTP_200_OK
        )
        
        
from rest_framework_simplejwt.exceptions import InvalidToken
from rest_framework_simplejwt.tokens import UntypedToken


class DistributorResetPasswordAPIView(APIView):

    permission_classes = [AllowAny]

    def post(self, request):

        # get token
        auth_header = request.headers.get("Authorization")

        if not auth_header:
            return Response({
                "message": "Authorization token required"
            }, status=status.HTTP_401_UNAUTHORIZED)

        try:
            token = auth_header.split(" ")[1]

            # decode jwt token
            decoded_token = UntypedToken(token)

            distributor_id = decoded_token.payload.get("distributor_id")

            distributor = DistributorInformation.objects.get(
                id=distributor_id
            )

        except InvalidToken:
            return Response({
                "message": "Invalid token"
            }, status=status.HTTP_401_UNAUTHORIZED)

        except DistributorInformation.DoesNotExist:
            return Response({
                "message": "Distributor not found"
            }, status=status.HTTP_404_NOT_FOUND)

        old_password = request.data.get("old_password")
        new_password = request.data.get("new_password")
        confirm_password = request.data.get("confirm_password")

        if not old_password:
            return Response({
                "old_password": "Old password is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        if not new_password:
            return Response({
                "new_password": "New password is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        if not confirm_password:
            return Response({
                "confirm_password": "Confirm password is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        db_password = distributor.password

        # ✅ hashed password
        if db_password.startswith("pbkdf2"):

            if not check_password(old_password, db_password):
                return Response({
                    "old_password": "Old password is incorrect"
                }, status=status.HTTP_400_BAD_REQUEST)

        else:
            # ✅ plain text password
            if db_password != old_password:
                return Response({
                    "old_password": "Old password is incorrect"
                }, status=status.HTTP_400_BAD_REQUEST)

        # confirm password check
        if new_password != confirm_password:
            return Response({
                "confirm_password": "Passwords do not match"
            }, status=status.HTTP_400_BAD_REQUEST)

        # save new password as HASHED
        distributor.password = make_password(new_password)
        distributor.save()

        return Response({
            "message": "Password reset successfully"
        }, status=status.HTTP_200_OK)




import jwt
from django.conf import settings        
class EmployeeResetPasswordAPIView(APIView):
 
    def post(self, request):
 
        auth_header = request.headers.get("Authorization")
 
        if not auth_header:
            return Response(
                {
                    "success": False,
                    "message": "Token required"
                },
                status=status.HTTP_401_UNAUTHORIZED
            )
 
        try:
            token = auth_header.split(" ")[1]
 
            decoded_data = jwt.decode(
                token,
                settings.SECRET_KEY,
                algorithms=["HS256"]
            )
 
            employee_id = decoded_data.get("employee_id")
 
            employee = Employee.objects.get(id=employee_id)
 
        except jwt.ExpiredSignatureError:
            return Response(
                {
                    "success": False,
                    "message": "Token expired"
                },
                status=status.HTTP_401_UNAUTHORIZED
            )
 
        except jwt.InvalidTokenError:
            return Response(
                {
                    "success": False,
                    "message": "Invalid token"
                },
                status=status.HTTP_401_UNAUTHORIZED
            )
 
        old_password = request.data.get("old_password")
        new_password = request.data.get("new_password")
        confirm_password = request.data.get("confirm_password")
 
        # Validation
        if employee.password != old_password:
            return Response(
                {
                    "success": False,
                    "message": "Old password incorrect"
                },
                status=status.HTTP_400_BAD_REQUEST
            )
 
        if new_password != confirm_password:
            return Response(
                {
                    "success": False,
                    "message": "Passwords do not match"
                },
                status=status.HTTP_400_BAD_REQUEST
            )
 
        # Update Password
        employee.password = new_password
        employee.save()
 
        return Response(
            {
                "success": True,
                "message": "Password changed successfully"
            }
        )
 
                       
class TestimonialView(APIView):

    # 🔹 GET (single / all)
    def get(self, request, id=None):
        if id:
            try:
                testimonial = Testimonial.objects.get(id=id)
                serializer = TestimonialSerializer(testimonial)
                return Response({
                    "status": "success",
                    "data": serializer.data
                })
            except Testimonial.DoesNotExist:
                return Response({"status": "invalid id"}, status=status.HTTP_404_NOT_FOUND)

        testimonials = Testimonial.objects.all().order_by("-created_at")
        serializer = TestimonialSerializer(testimonials, many=True)
        return Response({
            "status": "success",
            "message": "Testimonials Fetched Successfully",
            "data": serializer.data
        })

    # 🔹 POST
    def post(self, request):
        serializer = TestimonialSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "status": "success",
                "message": "Testimonial Created Successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": "invalid data",
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # 🔹 PATCH (partial update)
    def patch(self, request, id=None):
        if not id:
            return Response({"status": "invalid id"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            testimonial = Testimonial.objects.get(id=id)
        except Testimonial.DoesNotExist:
            return Response({"status": "invalid id"}, status=status.HTTP_404_NOT_FOUND)

        serializer = TestimonialSerializer(testimonial, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "status": "success",
                "message": "Testimonial Updated Successfully",
                "data": serializer.data
            })

        return Response({
            "status": "invalid data",
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # 🔹 DELETE
    def delete(self, request, id=None):
        if not id:
            return Response({"status": "invalid id"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            testimonial = Testimonial.objects.get(id=id)
            testimonial.delete()
            return Response({
                "status": "success",
                "message": "Testimonial Deleted Successfully"
            })
        except Testimonial.DoesNotExist:
            return Response({"status": "invalid id"}, status=status.HTTP_404_NOT_FOUND)



from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter

import os
from datetime import datetime

from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView

from .models import Shipment_product


import os
from datetime import datetime

from django.conf import settings
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView

from .models import Shipment_product


import os
from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView

from .models import Shipment_product


class ShipmentProductExcelDownloadAPI(APIView):

    def get(self, request):

        # =====================================
        # ✅ Get Filters
        # =====================================

        batch_id     = request.GET.get("batch_id", "").strip()
        product_name = request.GET.get("product_name", "").strip()

        # =====================================
        # ✅ Fetch Base Queryset
        # =====================================

        shipment_products = Shipment_product.objects.select_related(
            'product_id'
        ).all().order_by('-id')

        # =====================================
        # ✅ Filter By Batch
        # =====================================

        if batch_id:
            shipment_products = shipment_products.filter(
                batch_data=batch_id
            )

        # =====================================
        # ✅ Filter By Product Name
        # =====================================

        if product_name:
            shipment_products = shipment_products.filter(
                product_id__product_name__icontains=product_name
            )

        # =====================================
        # ✅ Evaluate Queryset Once
        # =====================================

        shipment_products = list(shipment_products)

        # =====================================
        # ✅ Create Workbook
        # =====================================

        wb = Workbook()
        ws = wb.active
        ws.title = "Shipment Product Report"

        # =====================================
        # ✅ Styles
        # =====================================

        blue_fill = PatternFill(
            start_color="000080",
            end_color="000080",
            fill_type="solid"
        )

        white_font   = Font(bold=True, color="FFFFFF", name="Arial")
        heading_font = Font(bold=True, color="FFFFFF", size=16, name="Arial")
        data_font    = Font(name="Arial", size=10)

        center_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # =====================================
        # ✅ Main Heading
        # =====================================

        ws.merge_cells('A1:I2')
        heading_cell           = ws['A1']
        heading_cell.value     = "SHIPMENT PRODUCT REPORT"
        heading_cell.fill      = blue_fill
        heading_cell.font      = heading_font
        heading_cell.alignment = center_alignment
        heading_cell.border    = thin_border

        for row in ws['A1:I2']:
            for cell in row:
                cell.fill   = blue_fill
                cell.border = thin_border

        # =====================================
        # ✅ Filter Info Row
        # =====================================

        ws.merge_cells('A3:I3')
        filter_text = (
            f"Batch ID : {batch_id or 'All'}"
            f"    |    "
            f"Product Name : {product_name or 'All'}"
        )
        filter_cell           = ws['A3']
        filter_cell.value     = filter_text
        filter_cell.font      = Font(bold=True, size=11, name="Arial")
        filter_cell.alignment = center_alignment
        filter_cell.border    = thin_border

        # =====================================
        # ✅ Table Headers
        # =====================================

        headers = [
            "Sr",
            "Date",
            "Product",
            "Batch",
            "Landed Cost",
            "INR / Unit",
            "USD / Unit",
            "Allocation Basis",
            "Qty",
        ]

        for col_num, header in enumerate(headers, 1):
            cell           = ws.cell(row=4, column=col_num, value=header)
            cell.fill      = blue_fill
            cell.font      = white_font
            cell.alignment = center_alignment
            cell.border    = thin_border

        # =====================================
        # ✅ Data Rows
        # =====================================

        for row_num, item in enumerate(shipment_products, start=5):

            item_date = item.created_at.strftime("%d-%m-%Y") if item.created_at else "N/A"

            row_data = [
                row_num - 4,
                item_date,
                str(item.product_id),
                item.batch_data,
                float(item.landed_cost_allocated or 0),
                float(item.per_unit_cost_inr or 0),
                float(item.per_unit_cost_usd or 0),
                item.allocation_basis,
                item.quantity,
            ]

            for col_num, value in enumerate(row_data, start=1):
                cell           = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = center_alignment
                cell.border    = thin_border
                cell.font      = data_font

        # =====================================
        # ✅ Column Widths
        # =====================================

        column_widths = {
            1: 8,   # Sr
            2: 18,  # Date
            3: 40,  # Product
            4: 25,  # Batch
            5: 18,  # Landed Cost
            6: 18,  # INR / Unit
            7: 18,  # USD / Unit
            8: 25,  # Allocation Basis
            9: 15,  # Qty
        }

        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # =====================================
        # ✅ Row Heights
        # =====================================

        for row_num, height in {1: 30, 2: 30, 3: 25, 4: 25}.items():
            ws.row_dimensions[row_num].height = height

        # =====================================
        # ✅ Save To Memory & Return Response
        # =====================================

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        file_name = f"shipment_product_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        return response
    
class AllCategoryTargetReportExcelAPIView(APIView):

    RED_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    CENTER   = Alignment(horizontal="center", vertical="center")
    LEFT     = Alignment(horizontal="left", vertical="center")

    def _to_decimal(self, value, default="0.00"):
        try:
            if value in [None, "", "null"]:
                return Decimal(default)
            return Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    def _get_item_product_id(self, item):
        for key in ["product_id", "product", "id"]:
            value = item.get(key)
            if value not in [None, "", "null"]:
                if isinstance(value, dict):
                    return str(value.get("id") or value.get("product_id") or "")
                return str(value)
        return None

    def _parse_year_month(self, year_value, month_value):
        year = datetime.now().year
        selected_month = None

        if year_value:
            try:
                year = int(year_value)
            except ValueError:
                return None, None, {"success": False, "message": "Invalid year format"}

        if month_value:
            try:
                if len(month_value) == 2:
                    selected_month = int(month_value)
                elif len(month_value) == 7:
                    parsed = datetime.strptime(month_value, "%Y-%m")
                    year, selected_month = parsed.year, parsed.month
                elif len(month_value) == 10:
                    parsed = datetime.strptime(month_value, "%Y-%m-%d")
                    year, selected_month = parsed.year, parsed.month
                else:
                    return None, None, {"success": False, "message": "Month format must be MM, YYYY-MM, or YYYY-MM-DD"}
            except ValueError:
                return None, None, {"success": False, "message": "Invalid month format"}

        return year, selected_month, None

    def _style_cell(self, cell, value, bold=False, fill=None, align=None, font_color="000000"):
        cell.value = value
        cell.font  = Font(bold=bold, color=font_color)
        if fill:
            cell.fill = fill
        cell.alignment = align or self.CENTER

    def _calc_category_month_data(self, category_obj, purchase_orders, year, month):
        """
        Returns (target, total_product_qty) for a category in a given month.
        - target        : from Category_Target model
        - total_product : sum of quantities from PO product_items for this category's products
        """
        # ── Target ────────────────────────────────────────────────────────
        category_target = Category_Target.objects.filter(
            category_id=category_obj.id,
            month__year=year,
            month__month=month
        ).first()
        m_target = self._to_decimal(category_target.target) if category_target else Decimal("0.00")

        # ── Total Product qty from PurchaseOrder.product_items ────────────
        product_ids = set(
            str(pid) for pid in
            Product.objects.filter(category_id=category_obj.id)
                           .values_list("id", flat=True)
        )

        if not product_ids:
            return m_target, Decimal("0.00")

        # Filter POs for this month
        month_pos = purchase_orders.filter(
            Q(po_date__month=month) | Q(po_date__isnull=True, created_at__month=month)
        )

        total_qty = 0
        for po in month_pos:
            items = po.product_items or []
            if not isinstance(items, list):
                continue
            for item in items:
                if not isinstance(item, dict):
                    continue
                item_product_id = self._get_item_product_id(item)
                if item_product_id in product_ids:
                    qty = item.get("quantity", item.get("qty", 0))
                    try:
                        total_qty += int(self._to_decimal(qty))
                    except Exception:
                        pass

        return m_target, Decimal(str(total_qty))

    def get(self, request):
        month_value = request.GET.get("month")
        year_value  = request.GET.get("year")

        year, selected_month, err = self._parse_year_month(year_value, month_value)
        if err:
            return Response(err, status=status.HTTP_400_BAD_REQUEST)

        months_to_loop = [selected_month] if selected_month else range(1, 13)

        categories = Category.objects.all()
        if not categories.exists():
            return Response({"success": False, "message": "No categories found"},
                            status=status.HTTP_404_NOT_FOUND)

        # All POs for the year — month filtering done inside _calc_category_month_data
        purchase_orders = PurchaseOrder.objects.filter(
            Q(po_date__year=year) | Q(po_date__isnull=True, created_at__year=year)
        ).order_by("-id")

        workbook = Workbook()
        ws = workbook.active
        ws.title = "All Categories Target Report"

        total_cols = len(categories) * 2 + 1

        # # ── Title row ─────────────────────────────────────────────────────
        # title_text = f"All Category Target Report - {year}"
        # ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        # self._style_cell(ws["A1"], title_text, bold=True, fill=self.RED_FILL, font_color="FFFFFF")
        # ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")

        # ── Header row ────────────────────────────────────────────────────
        headers = ["Month"]
        for category_obj in categories:
            cat_name = category_obj.name or "Category"
            headers.extend([f"{cat_name} Target", f"{cat_name} Total Product"])

        for col, header in enumerate(headers, 1):
            self._style_cell(
                ws.cell(row=2, column=col), header,
                bold=True, fill=self.RED_FILL, font_color="FFFFFF"
            )

        # ── Month rows ────────────────────────────────────────────────────
        row_num = 3
        for month in months_to_loop:
            month_name = calendar.month_abbr[month] + f"-{str(year)[-2:]}"
            ws.cell(row=row_num, column=1).value     = month_name
            ws.cell(row=row_num, column=1).alignment = self.CENTER

            col_num = 2
            for category_obj in categories:
                m_target, m_product = self._calc_category_month_data(
                    category_obj, purchase_orders, year, month
                )

                # Target cell
                t_cell                = ws.cell(row=row_num, column=col_num)
                t_cell.value          = float(m_target.quantize(Decimal("0.01")))
                t_cell.number_format  = "#,##0.00"
                t_cell.alignment      = self.CENTER

                # Total Product cell
                p_cell                = ws.cell(row=row_num, column=col_num + 1)
                p_cell.value          = float(m_product.quantize(Decimal("0.01")))
                p_cell.number_format  = "#,##0.00"
                p_cell.alignment      = self.CENTER

                col_num += 2

            row_num += 1

        # ── Auto column width ─────────────────────────────────────────────
        for col in range(1, ws.max_column + 1):
            max_len    = 0
            col_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = max_len + 5

        # ── HTTP response ─────────────────────────────────────────────────
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        file_name = f"all_categories_target_product_report_{year}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
        workbook.save(response)
        return response


class AllProductTargetReportExcelAPIView(APIView):

    RED_FILL  = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    CENTER    = Alignment(horizontal="center", vertical="center")

    def _to_decimal(self, value, default="0.00"):
        try:
            if value in [None, "", "null"]:
                return Decimal(default)
            cleaned = str(value).strip()
            float(cleaned)
            return Decimal(cleaned)
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    def _safe_str(self, value, default=""):
        if value in [None, "", "null"]:
            return default
        return str(value)

    def _get_product_name(self, product):
        for field in ["product_name", "name", "title", "item_name"]:
            value = getattr(product, field, None)
            if value not in [None, "", "null"]:
                return str(value)
        return str(product) if product else ""

    def _get_category_id(self, product):
        cat = getattr(product, "category_id", None)
        if cat is None:
            return None
        if hasattr(cat, "id"):
            return cat.id
        try:
            return int(cat)
        except Exception:
            return None

    def _get_item_product_id(self, item):
        for key in ["product_id", "product", "id"]:
            value = item.get(key)
            if value not in [None, "", "null"]:
                if isinstance(value, dict):
                    return str(value.get("id") or value.get("product_id") or "")
                return str(value)
        return None

    def _parse_items(self, product_items):
        if not product_items:
            return []
        if isinstance(product_items, list):
            return product_items
        if isinstance(product_items, str):
            try:
                import json
                parsed = json.loads(product_items)
                return parsed if isinstance(parsed, list) else []
            except Exception:
                return []
        return []

    def _parse_year_month(self, year_value, month_value):
        year = datetime.now().year
        selected_month = None
        if year_value:
            try:
                year = int(year_value)
            except ValueError:
                return None, None, {"success": False, "message": "Invalid year format"}
        if month_value:
            try:
                if len(month_value) == 2:
                    selected_month = int(month_value)
                elif len(month_value) == 7:
                    parsed = datetime.strptime(month_value, "%Y-%m")
                    year, selected_month = parsed.year, parsed.month
                elif len(month_value) == 10:
                    parsed = datetime.strptime(month_value, "%Y-%m-%d")
                    year, selected_month = parsed.year, parsed.month
                else:
                    return None, None, {"success": False, "message": "Invalid month format"}
            except ValueError:
                return None, None, {"success": False, "message": "Invalid month format"}
        return year, selected_month, None

    def _style_cell(self, cell, value, bold=False, fill=None, font_color="000000"):
        cell.value     = value
        cell.font      = Font(bold=bold, color=font_color)
        cell.alignment = self.CENTER
        if fill:
            cell.fill = fill

    def _get_product_target(self, product, year, month, target_cache, count_cache):
        cache_key = (product.id, year, month)
        if cache_key in target_cache:
            return target_cache[cache_key]

        # Try Product_Target first
        try:
            from .models import Product_Target
            pt = Product_Target.objects.filter(
                product_id=product.id,
                month__year=year,
                month__month=month
            ).first()
            if pt:
                try:
                    val = Decimal(str(pt.target).strip())
                    float(val)
                    target_cache[cache_key] = val
                    return val
                except Exception:
                    pass
        except Exception:
            pass

        # Fall back to Category_Target ÷ product count
        cat_id  = self._get_category_id(product)
        cat_key = (cat_id, year, month)

        if cat_key not in target_cache:
            ct = Category_Target.objects.filter(
                category_id=cat_id,
                month__year=year,
                month__month=month
            ).first()
            if ct:
                try:
                    t = Decimal(str(ct.target).strip())
                    float(t)
                    target_cache[cat_key] = t
                except Exception:
                    target_cache[cat_key] = Decimal("0.00")
            else:
                target_cache[cat_key] = Decimal("0.00")

        if cat_id not in count_cache:
            count_cache[cat_id] = Product.objects.filter(
                category_id=cat_id
            ).count()

        cat_target    = target_cache[cat_key]
        product_count = count_cache[cat_id]
        per_target    = (
            cat_target / Decimal(product_count)
            if (product_count and cat_target > 0)
            else Decimal("0.00")
        )
        target_cache[cache_key] = per_target
        return per_target

    def _get_product_qty(self, product_id, month_pos):
        total_qty = 0
        for po in month_pos:
            items = self._parse_items(po.product_items)
            for item in items:
                if not isinstance(item, dict):
                    continue
                item_pid = self._get_item_product_id(item)
                if item_pid and str(item_pid) == str(product_id):
                    qty = self._to_decimal(
                        item.get("quantity", item.get("qty", 0))
                    )
                    if qty > 0:
                        total_qty += int(qty)
        return total_qty

    def get(self, request):
        month_value = request.GET.get("month")
        year_value  = request.GET.get("year")
        category_id = request.GET.get("category")

        year, selected_month, err = self._parse_year_month(year_value, month_value)
        if err:
            return Response(err, status=status.HTTP_400_BAD_REQUEST)

        months_to_loop = [selected_month] if selected_month else range(1, 13)

        # ── Products ──────────────────────────────────────────────────────
        products = Product.objects.select_related("category_id").order_by(
            "category_id", "id"
        )
        if category_id:
            products = products.filter(category_id=category_id)

        if not products.exists():
            return Response(
                {"success": False, "message": "No products found"},
                status=status.HTTP_404_NOT_FOUND
            )

        product_list = list(products)

        # ── All POs for the year ──────────────────────────────────────────
        purchase_orders = PurchaseOrder.objects.filter(
            Q(po_date__year=year) | Q(created_at__year=year)
        ).distinct().order_by("-id")

        # Pre-build month → PO list
        month_po_map = {}
        for month in months_to_loop:
            month_po_map[month] = list(
                purchase_orders.filter(
                    Q(po_date__month=month) | Q(created_at__month=month)
                ).distinct()
            )

        # ── Excel setup ───────────────────────────────────────────────────
        workbook   = Workbook()
        ws         = workbook.active
        ws.title   = "All Products Target Report"
        total_cols = len(product_list) * 2 + 1

        # ── Row 1: Title ──────────────────────────────────────────────────
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1,   end_column=total_cols
        )
        self._style_cell(
            ws["A1"], f"All Product Target Report - {year}",
            bold=True, fill=self.RED_FILL, font_color="FFFFFF"
        )
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")

        # ── Row 2: Headers ────────────────────────────────────────────────
        self._style_cell(
            ws.cell(row=2, column=1), "Month",
            bold=True, fill=self.RED_FILL, font_color="FFFFFF"
        )
        col_num = 2
        for product in product_list:
            p_name = self._safe_str(self._get_product_name(product), "Product")
            self._style_cell(
                ws.cell(row=2, column=col_num),
                f"{p_name} Target",
                bold=True, fill=self.RED_FILL, font_color="FFFFFF"
            )
            self._style_cell(
                ws.cell(row=2, column=col_num + 1),
                f"{p_name} Total Product",
                bold=True, fill=self.RED_FILL, font_color="FFFFFF"
            )
            col_num += 2

        # ── Data rows ─────────────────────────────────────────────────────
        target_cache = {}
        count_cache  = {}
        row_num      = 3

        for month in months_to_loop:
            month_name = calendar.month_abbr[month] + f"-{str(year)[-2:]}"
            ws.cell(row=row_num, column=1).value     = month_name
            ws.cell(row=row_num, column=1).alignment = self.CENTER

            month_pos = month_po_map[month]
            col_num   = 2

            for product in product_list:
                p_target = self._get_product_target(
                    product, year, month, target_cache, count_cache
                )
                p_qty = self._get_product_qty(product.id, month_pos)

                # Target cell
                t_cell               = ws.cell(row=row_num, column=col_num)
                t_cell.value         = float(p_target.quantize(Decimal("0.01")))
                t_cell.number_format = "#,##0.00"
                t_cell.alignment     = self.CENTER
                if p_target == Decimal("0.00"):
                    t_cell.fill = self.GRAY_FILL

                # Total Product cell
                p_cell               = ws.cell(row=row_num, column=col_num + 1)
                p_cell.value         = p_qty
                p_cell.number_format = "#,##0"
                p_cell.alignment     = self.CENTER

                col_num += 2

            row_num += 1

        # ── Auto column width ─────────────────────────────────────────────
        for col in range(1, ws.max_column + 1):
            max_len    = 0
            col_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = max_len + 5

        # ── HTTP response ─────────────────────────────────────────────────
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        file_name = f"all_products_target_report_{year}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
        workbook.save(response)
        return response
    
    
class AllDistributorTargetReportExcelAPIView(APIView):
    
    RED_FILL    = PatternFill(fill_type="solid", start_color="C00000", end_color="C00000")
    CENTER      = Alignment(horizontal="center", vertical="center")

    def _to_decimal(self, value, default="0.00"):
        try:
            if value in [None, "", "null"]:
                return Decimal(default)
            return Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    def _to_float(self, value):
        return float(self._to_decimal(value).quantize(Decimal("0.01")))

    def _parse_year_month(self, year_value, month_value):
        year = datetime.now().year
        selected_month = None
        if year_value:
            try:
                year = int(year_value)
            except ValueError:
                return None, None, {"success": False, "message": "Invalid year format"}
        if month_value:
            try:
                if len(month_value) == 2:
                    selected_month = int(month_value)
                elif len(month_value) == 7:
                    parsed = datetime.strptime(month_value, "%Y-%m")
                    year, selected_month = parsed.year, parsed.month
                elif len(month_value) == 10:
                    parsed = datetime.strptime(month_value, "%Y-%m-%d")
                    year, selected_month = parsed.year, parsed.month
                else:
                    return None, None, {"success": False, "message": "Month format must be MM, YYYY-MM, or YYYY-MM-DD"}
            except ValueError:
                return None, None, {"success": False, "message": "Invalid month format"}
        return year, selected_month, None

    def get(self, request):
        month_value = request.GET.get("month")
        year_value  = request.GET.get("year")

        year, selected_month, err = self._parse_year_month(year_value, month_value)
        if err:
            return Response(err, status=status.HTTP_400_BAD_REQUEST)

        months_to_loop = [selected_month] if selected_month else range(1, 13)

        distributors = DistributorInformation.objects.all().order_by("id")
        if not distributors.exists():
            return Response({"success": False, "message": "No distributors found"},
                            status=status.HTTP_404_NOT_FOUND)

        purchase_orders = PurchaseOrder.objects.filter(
            Q(po_date__year=year) | Q(po_date__isnull=True, created_at__year=year)
        ).order_by("-id")

        wb = Workbook()
        ws = wb.active
        ws.title = "All Distributors Target Report"

        # ---- Title row ----
        title_text = f"All Distributor Target Report - {year}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(distributors)*2+1)
        title = ws["A1"]
        title.value = title_text
        title.font = Font(bold=True, size=14, color="FFFFFF")
        title.fill = self.RED_FILL
        title.alignment = self.CENTER

        # ---- Header row ----
        headers = ["Month"]
        for dist in distributors:
            dname = getattr(dist, "distributor_name", None) or getattr(dist, "company_name", None) or str(dist)
            headers.extend([f"{dname} Target", f"{dname} Total Product"])
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = self.RED_FILL
            cell.alignment = self.CENTER

        # ---- Fill month rows ----
        row_num = 3
        for m in months_to_loop:
            month_name = calendar.month_abbr[m] + f"-{str(year)[-2:]}"
            ws.cell(row=row_num, column=1).value = month_name
            ws.cell(row=row_num, column=1).alignment = self.CENTER

            col_num = 2
            for dist in distributors:
                # Target
                target_obj = DistributorInformation_Target.objects.filter(
                    distributor_id=dist.id,
                    month__year=year,
                    month__month=m
                ).first()
                month_target = self._to_decimal(target_obj.target if target_obj else 0)

                # Orders
                month_orders = purchase_orders.filter(
                    distributor_id=dist.id
                ).filter(
                    Q(po_date__month=m) | Q(po_date__isnull=True, created_at__month=m)
                )
                month_product = Decimal("0.00")
                for po in month_orders:
                    for item in (po.product_items or []):
                        if isinstance(item, dict):
                            qty = self._to_decimal(item.get("quantity", 0))
                            month_product += qty

                # Write Target + Product
                for val in [month_target, month_product]:
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = self._to_float(val)
                    cell.number_format = "0.00"
                    cell.alignment = self.CENTER
                    col_num += 1

            row_num += 1

        # ---- Auto column width ----
        for col in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for r in range(1, ws.max_row + 1):
                val = ws.cell(row=r, column=col).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = max_len + 5

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        file_name = f"all_distributors_target_report_{year}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
        wb.save(response)
        return response

class AllRegionTargetReportExcelAPIView(APIView):
    
    RED_FILL    = PatternFill(fill_type="solid", start_color="C00000", end_color="C00000")
    CENTER      = Alignment(horizontal="center", vertical="center")

    gst_rate = Decimal("0.18")

    def _to_decimal(self, value, default="0.00"):
        try:
            if value in [None, "", "null"]:
                return Decimal(default)
            return Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    def _to_float(self, value):
        return float(self._to_decimal(value).quantize(Decimal("0.01")))

    def _parse_year_month(self, year_value, month_value):
        year = datetime.now().year
        selected_month = None
        if year_value:
            try:
                year = int(year_value)
            except ValueError:
                return None, None, {"success": False, "message": "Invalid year format"}
        if month_value:
            try:
                if len(month_value) == 2:
                    selected_month = int(month_value)
                elif len(month_value) == 7:
                    parsed = datetime.strptime(month_value, "%Y-%m")
                    year, selected_month = parsed.year, parsed.month
                elif len(month_value) == 10:
                    parsed = datetime.strptime(month_value, "%Y-%m-%d")
                    year, selected_month = parsed.year, parsed.month
                else:
                    return None, None, {"success": False, "message": "Month format must be MM, YYYY-MM, or YYYY-MM-DD"}
            except ValueError:
                return None, None, {"success": False, "message": "Invalid month format"}
        return year, selected_month, None

    def get(self, request):
        month_value = request.GET.get("month")
        year_value  = request.GET.get("year")

        year, selected_month, err = self._parse_year_month(year_value, month_value)
        if err:
            return Response(err, status=status.HTTP_400_BAD_REQUEST)

        months_to_loop = [selected_month] if selected_month else range(1, 13)

        regions = Region.objects.all().order_by("id")
        if not regions.exists():
            return Response({"success": False, "message": "No regions found"},
                            status=status.HTTP_404_NOT_FOUND)

        wb = Workbook()
        ws = wb.active
        ws.title = "All Regions Target Report"

        # ---- Title row ----
        title_text = f"All Region Target Report - {year}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(regions)*2+1)
        title = ws["A1"]
        title.value = title_text
        title.font = Font(bold=True, size=14, color="FFFFFF")
        title.fill = self.RED_FILL
        title.alignment = self.CENTER

        # ---- Header row ----
        headers = ["Month"]
        for region_obj in regions:
            headers.extend([f"{region_obj.name} Target", f"{region_obj.name} Total Product"])
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = self.RED_FILL
            cell.alignment = self.CENTER

        # ---- Fill month rows ----
        row_num = 3
        for m in months_to_loop:
            month_name = calendar.month_abbr[m] + f"-{str(year)[-2:]}"
            ws.cell(row=row_num, column=1).value = month_name
            ws.cell(row=row_num, column=1).alignment = self.CENTER

            col_num = 2
            for region_obj in regions:
                # Target
                target_obj = Region_Target.objects.filter(
                    region_id=region_obj.id,
                    month__year=year,
                    month__month=m
                ).first()
                month_target = self._to_decimal(target_obj.target if target_obj else 0)

                # Orders
                distributors = DistributorInformation.objects.filter(
                    Q(region=region_obj) |
                    Q(sales_region=str(region_obj.id)) |
                    Q(sales_region__iexact=region_obj.name)
                )
                distributor_ids = list(distributors.values_list("id", flat=True))
                month_orders = PurchaseOrder.objects.filter(
                    distributor_id__in=distributor_ids
                ).filter(
                    Q(po_date__year=year, po_date__month=m) |
                    Q(po_date__isnull=True, created_at__year=year, created_at__month=m)
                )
                month_product_total = Decimal("0.00")
                for po in month_orders:
                    items = po.product_items or []
                    for item in items:
                        if not isinstance(item, dict):
                            continue
                        qty = self._to_decimal(item.get("quantity", 0))
                        month_product_total += qty

                # Write Target + Product
                for val in [month_target, month_product_total]:
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = self._to_float(val)
                    cell.number_format = "0.00"
                    cell.alignment = self.CENTER
                    col_num += 1

            row_num += 1
        # ---- Auto column width ----
        for col in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for r in range(1, ws.max_row + 1):
                val = ws.cell(row=r, column=col).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = max_len + 5

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        file_name = f"all_regions_target_report_{year}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
        wb.save(response)
        return response
# --------------------------------------------------------------------------
# ADD THESE IMPORTS to the top of your existing views.py (alongside the
# imports already used by WarrantyAPI)
# --------------------------------------------------------------------------
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse, JsonResponse
from datetime import datetime


# --------------------------------------------------------------------------
# NEW VIEW: exports Warranty records as a downloadable .xlsx file, with
# column names matching the Warranty Management table in the admin panel
# (Serial Number, Car Details, Detailer, Mobile, Install Date, Warranty,
# Warranty Status, Hold Reason).
#
#   GET /warranty/excel/                                  -> full list
#   GET /warranty/excel/?warranty_status=ACTIVE            -> by status
#   GET /warranty/excel/?start_date=2026-01-01&end_date=2026-06-01
#                                                           -> by date range
#   (status + date filters can be combined)
#
# Drop this class into views.py near WarrantyAPI.
# --------------------------------------------------------------------------
class WarrantyExcelExportAPIView(APIView):

    # Confirmed against the Warranty model: created_at is a
    # DateTimeField(auto_now_add=True), filtered via the __date lookup below.
    DATE_FIELD = "created_at"

    RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    HEADING_FONT = Font(bold=True, color="000000", size=14)

    # Matches the admin table columns exactly (minus "Action", which has
    # nothing to export since it's just the view/edit icons in the UI)
    HEADERS = [
        "Serial Number",
        "Car Details",
        "Detailer",
        "Mobile",
        "Install Date",
        "Warranty",
        "Warranty Status",
        "Hold Reason",
    ]

    def get(self, request):
        warranties = Warranty.objects.select_related("serial_id").order_by("-id")

        warranty_status = request.query_params.get("warranty_status")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        if warranty_status:
            warranties = warranties.filter(warranty_status__iexact=warranty_status)

        if start_date:
            try:
                parsed_start = datetime.strptime(start_date, "%Y-%m-%d").date()
                warranties = warranties.filter(**{f"{self.DATE_FIELD}__date__gte": parsed_start})
            except ValueError:
                return JsonResponse(
                    {"success": False, "errors": "start_date must be in YYYY-MM-DD format"},
                    status=400
                )

        if end_date:
            try:
                parsed_end = datetime.strptime(end_date, "%Y-%m-%d").date()
                warranties = warranties.filter(**{f"{self.DATE_FIELD}__date__lte": parsed_end})
            except ValueError:
                return JsonResponse(
                    {"success": False, "errors": "end_date must be in YYYY-MM-DD format"},
                    status=400
                )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Warranties"

        # Row 1: "Warranty" heading, merged across all columns, red background
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.HEADERS))
        heading_cell = ws.cell(row=1, column=1, value="Warranty")
        heading_cell.fill = self.RED_FILL
        heading_cell.font = self.HEADING_FONT
        heading_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        # Row 2: column headers (bold)
        ws.append(self.HEADERS)
        for cell in ws[2]:
            cell.font = cell.font.copy(bold=True)

        if not warranties.exists():
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(self.HEADERS))
            ws.cell(row=3, column=1, value="No warranty records found").alignment = Alignment(horizontal="center")
        else:
            for w in warranties:
                serial_number = w.serial_id.serial_number if w.serial_id else "-"

                car_details = " ".join(filter(None, [w.car_brand, w.car_model]))
                if w.car_registration_number:
                    car_details = f"{car_details} ({w.car_registration_number})" if car_details else w.car_registration_number
                if not car_details:
                    car_details = "-"

                install_date = w.installation_date.strftime("%d %b %Y") if w.installation_date else "-"
                warranty_period = f"{w.warranty_period} year" if w.warranty_period else "-"
                hold_reason = w.hold_reason or "-"

                ws.append([
                    serial_number,
                    car_details,
                    w.detailer_name or "-",
                    w.detailer_mobile or "-",
                    install_date,
                    warranty_period,
                    w.warranty_status,
                    hold_reason,
                ])

        # Reasonable fixed column widths matching each column's content type
        widths = [16, 35, 22, 16, 16, 12, 16, 28]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Build a descriptive filename based on whatever filters were applied
        filename_parts = ["warranty_list"]
        if warranty_status:
            filename_parts.append(warranty_status.lower())
        if start_date:
            filename_parts.append(start_date)
        if end_date:
            filename_parts.append(end_date)
        filename = "_".join(filename_parts) + ".xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response












































































































































































































































































































